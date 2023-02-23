from django.core.paginator import Paginator
from decimal import Decimal
import json
from openpyxl import Workbook
from openpyxl.styles import *
from urllib import request
from django.shortcuts import render
from applications.clientes.models import Cliente
from applications.cobranza.funciones import eliminarDeuda, generarDeuda
from applications.comprobante_venta.forms import BoletaVentaAnularForm, BoletaVentaBuscarForm, BoletaVentaSerieForm, DescargarComprobantesForm, FacturaVentaAnularForm, FacturaVentaBuscarForm, FacturaVentaDetalleForm, FacturaVentaObservacionForm, FacturaVentaSerieForm
from applications.comprobante_venta.funciones import anular_nubefact, boleta_nubefact, consultar_documento, factura_nubefact
from applications.cotizacion.models import ConfirmacionVenta
from applications.datos_globales.models import NubefactRespuesta, SeriesComprobante, TipoCambio, Unidad
from applications.funciones import calculos_linea, consulta_ruc, get_datetime, igv, numeroXn, obtener_totales, registrar_excepcion, slug_aleatorio, tipo_de_cambio
from applications.importaciones import *

from . models import(
    BoletaVenta,
    BoletaVentaDetalle,
    FacturaVenta,
    FacturaVentaDetalle,
)

class FacturaVentaListView(PermissionRequiredMixin, FormView):
    permission_required = ('comprobante_venta.view_facturaventa')
    template_name = 'comprobante_venta/factura_venta/inicio.html'
    form_class = FacturaVentaBuscarForm
    success_url = '.'

    def get_form_kwargs(self):
        kwargs = super(FacturaVentaListView, self).get_form_kwargs()
        kwargs['filtro_numero_factura'] = self.request.GET.get('numero_factura')
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_cliente'] = self.request.GET.get('cliente')
        kwargs['filtro_fecha_emision'] = self.request.GET.get('fecha_emision')
        kwargs['filtro_estado'] = self.request.GET.get('estado')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaListView,self).get_context_data(**kwargs)
        factura_venta = FacturaVenta.objects.all()

        filtro_numero_factura = self.request.GET.get('numero_factura')
        filtro_sociedad = self.request.GET.get('sociedad')
        filtro_cliente = self.request.GET.get('cliente')
        filtro_fecha_emision = self.request.GET.get('fecha_emision')
        filtro_estado = self.request.GET.get('estado')

        contexto_filtro = []

        if filtro_cliente:
            condicion = Q(cliente = filtro_cliente)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("cliente=" + filtro_cliente)

        if filtro_fecha_emision:
            condicion = Q(fecha_emision = filtro_fecha_emision)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("fecha_emision=" + filtro_fecha_emision)

        if filtro_sociedad:
            condicion = Q(sociedad = filtro_sociedad)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("sociedad=" + filtro_sociedad)

        if filtro_numero_factura:
            condicion = Q(numero_factura = filtro_numero_factura)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("numero_factura=" + filtro_numero_factura)

        if filtro_estado:
            condicion = Q(estado = filtro_estado)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("estado=" + filtro_estado)
        
        context['contexto_filtro'] = "&".join(contexto_filtro)

        context['pagina_filtro'] = ""
        if self.request.GET.get('page'):
            if context['contexto_filtro']:
                context['pagina_filtro'] = f'&page={self.request.GET.get("page")}'
            else:
                context['pagina_filtro'] = f'page={self.request.GET.get("page")}'
        context['contexto_filtro'] = '?' + context['contexto_filtro']

        objectsxpage = 25 # Show 25 objects per page.

        if len(factura_venta) > objectsxpage:
            paginator = Paginator(factura_venta, objectsxpage)
            page_number = self.request.GET.get('page')
            factura_venta = paginator.get_page(page_number)
   
        context['contexto_pagina'] = factura_venta
        return context

def FacturaVentaTabla(request):
    data = dict()
    if request.method == 'GET':
        template = 'comprobante_venta/factura_venta/inicio_tabla.html'
        context = {}
        factura_venta = FacturaVenta.objects.all()
        filtro_numero_factura = request.GET.get('numero_factura')
        filtro_sociedad = request.GET.get('sociedad')
        filtro_cliente = request.GET.get('cliente')
        filtro_fecha_emision = request.GET.get('fecha_emision')
        filtro_estado = request.GET.get('estado')

        contexto_filtro = []

        if filtro_cliente:
            condicion = Q(cliente = filtro_cliente)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("cliente=" + filtro_cliente)

        if filtro_fecha_emision:
            condicion = Q(fecha_emision = filtro_fecha_emision)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("fecha_emision=" + filtro_fecha_emision)

        if filtro_sociedad:
            condicion = Q(sociedad = filtro_sociedad)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("sociedad=" + filtro_sociedad)

        if filtro_numero_factura:
            condicion = Q(numero_factura = filtro_numero_factura)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("numero_factura=" + filtro_numero_factura)

        if filtro_estado:
            condicion = Q(estado = filtro_estado)
            factura_venta = factura_venta.filter(condicion)
            contexto_filtro.append("estado=" + filtro_estado)
            
        context['contexto_filtro'] = "&".join(contexto_filtro)

        context['pagina_filtro'] = ""
        if request.GET.get('page'):
            if context['contexto_filtro']:
                context['pagina_filtro'] = f'&page={request.GET.get("page")}'
            else:
                context['pagina_filtro'] = f'page={request.GET.get("page")}'
        context['contexto_filtro'] = '?' + context['contexto_filtro']
        
        objectsxpage = 25 # Show 25 objects per page.

        if len(factura_venta) > objectsxpage:
            paginator = Paginator(factura_venta, objectsxpage)
            page_number = request.GET.get('page')
            factura_venta = paginator.get_page(page_number)
        
        context['contexto_pagina'] = factura_venta 

        data['table'] = render_to_string(
            template,
            context,
            request=request
        )
        return JsonResponse(data)

class FacturaVentaDetalleView(PermissionRequiredMixin, TemplateView):
    permission_required = ('comprobante_venta.view_facturaventa')
    template_name = "comprobante_venta/factura_venta/detalle.html"

    def get_context_data(self, **kwargs):
        obj = FacturaVenta.objects.get(id = kwargs['id_factura_venta'])

        materiales = None
        try:
            materiales = obj.FacturaVentaDetalle_factura_venta.all()

            for material in materiales:
                material.material = material.content_type.get_object_for_this_type(id = material.id_registro)
        except:
            pass
        
        tipo_cambio_hoy = TipoCambio.objects.tipo_cambio_venta(date.today())
        if obj.confirmacion:
            tipo_cambio = obj.confirmacion.tipo_cambio.tipo_cambio_venta
        else:
            tipo_cambio = Decimal('1.00')
        context = super(FacturaVentaDetalleView, self).get_context_data(**kwargs)
        context['factura'] = obj
        context['confirmacion'] = obj.confirmacion
        context['materiales'] = materiales
        context['tipo_cambio_hoy'] = tipo_cambio_hoy
        context['tipo_cambio'] = tipo_cambio
        tipo_cambio = tipo_de_cambio(tipo_cambio, tipo_cambio_hoy)
        context['totales'] = obtener_totales(FacturaVenta.objects.get(id=self.kwargs['id_factura_venta']))
        if obj.serie_comprobante:
            context['nubefact_acceso'] = obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj))
        url_nubefact = NubefactRespuesta.objects.respuesta(obj)
        if url_nubefact:
            context['url_nubefact'] = url_nubefact
        if obj.nubefact:
            context['url_nubefact'] = obj.nubefact
        context['respuestas_nubefact'] = NubefactRespuesta.objects.respuestas(obj)

        return context

def FacturaVentaDetalleVerTabla(request, id_factura_venta):
    data = dict()
    if request.method == 'GET':
        template = 'comprobante_venta/factura_venta/detalle_tabla.html'
        obj = FacturaVenta.objects.get(id=id_factura_venta)

        tipo_cambio_hoy = TipoCambio.objects.tipo_cambio_venta(date.today())
        if obj.confirmacion:
            tipo_cambio = obj.confirmacion.tipo_cambio.tipo_cambio_venta
        else:
            tipo_cambio = Decimal('1.00')
        materiales = None
        try:
            materiales = obj.FacturaVentaDetalle_factura_venta.all()

            for material in materiales:
                material.material = material.content_type.get_object_for_this_type(id = material.id_registro)
        except:
            pass

        context = {}
        context['factura'] = obj
        context['materiales'] = materiales
        context['tipo_cambio_hoy'] = tipo_cambio_hoy
        context['tipo_cambio'] = tipo_cambio
        tipo_cambio = tipo_de_cambio(tipo_cambio, tipo_cambio_hoy)
        context['totales'] = obtener_totales(obj)
        if obj.serie_comprobante:
            context['nubefact_acceso'] = obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj))
        url_nubefact = NubefactRespuesta.objects.respuesta(obj)
        if url_nubefact:
            context['url_nubefact'] = url_nubefact
        if obj.nubefact:
            context['url_nubefact'] = obj.nubefact
        context['respuestas_nubefact'] = NubefactRespuesta.objects.respuestas(obj)

        data['table'] = render_to_string(
            template,
            context,
            request=request
        )
        return JsonResponse(data)

class FacturaVentaCrearView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.add_facturaventa')
    model = ConfirmacionVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_codigo_sunat = False
        error_cuotas = False
        context['titulo'] = 'Error de guardar'
        detalles = self.get_object().ConfirmacionVentaDetalle_confirmacion_venta.all()
        for detalle in detalles:
            producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
            if not producto.producto_sunat:
                error_codigo_sunat = True
        
        if self.get_object().tipo_venta == 2 and not self.get_object().ConfirmacionVentaCuota_confirmacion_venta.all():
            error_cuotas = True

        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat.'
            return render(request, 'includes/modal sin permiso.html', context)
        if error_cuotas:
            context['texto'] = 'Falta ingresar las cuotas.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        try:
            return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['factura_venta'].id})
        except:
            return reverse_lazy('comprobante_venta_app:factura_venta_inicio')

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            self.object = self.get_object()

            detalles = self.object.ConfirmacionVentaDetalle_confirmacion_venta.all()

            serie_comprobante = SeriesComprobante.objects.por_defecto(ContentType.objects.get_for_model(FacturaVenta))

            factura_venta = FacturaVenta.objects.create(
                confirmacion=self.object,
                sociedad = self.object.sociedad,
                serie_comprobante = serie_comprobante,
                cliente = self.object.cliente,
                cliente_interlocutor = self.object.cliente_interlocutor,
                moneda = self.object.moneda,
                tipo_cambio = self.object.tipo_cambio,
                tipo_venta = self.object.tipo_venta,
                condiciones_pago = self.object.condiciones_pago,
                descuento_global = self.object.descuento_global,
                total_descuento = self.object.total_descuento,
                total_anticipo = self.object.total_anticipo,
                total_gravada = self.object.total_gravada,
                total_inafecta = self.object.total_inafecta,
                total_exonerada = self.object.total_exonerada,
                total_igv = self.object.total_igv,
                total_gratuita = self.object.total_gratuita,
                total_otros_cargos = self.object.otros_cargos,
                total = self.object.total,
                observaciones = self.object.observacion,
                slug = slug_aleatorio(FacturaVenta),
                created_by=self.request.user,
                updated_by=self.request.user,
            )

            for detalle in detalles:
                producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
                factura_venta_detalle = FacturaVentaDetalle.objects.create(
                    item=detalle.item,
                    content_type=detalle.content_type,
                    id_registro=detalle.id_registro,
                    unidad=producto.unidad_base,
                    descripcion_documento=producto.descripcion_documento,
                    cantidad=detalle.cantidad_confirmada,
                    precio_unitario_sin_igv=detalle.precio_unitario_sin_igv,
                    precio_unitario_con_igv=detalle.precio_unitario_con_igv,
                    precio_final_con_igv=detalle.precio_final_con_igv,
                    descuento=detalle.descuento,
                    sub_total=detalle.sub_total,
                    tipo_igv=detalle.tipo_igv,
                    igv=detalle.igv,
                    total=detalle.total,
                    codigo_producto_sunat=producto.producto_sunat.codigo,
                    factura_venta=factura_venta,
                    created_by=self.request.user,
                    updated_by=self.request.user,
                )
            self.kwargs['factura_venta'] = factura_venta

            registro_guardar(self.object, self.request)
            self.object.save()

            messages.success(request, MENSAJE_GENERAR_FACTURA)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaCrearView, self).get_context_data(**kwargs)
        context['accion'] = 'Generar'
        context['titulo'] = 'Factura de venta'
        context['texto'] = '¿Seguro que desea generar la Factura de venta?'
        context['item'] = str(self.object.cliente) 
        return context


class FacturaVentaAnticipoCrearView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.add_facturaventa')
    model = ConfirmacionVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_codigo_sunat = False
        context['titulo'] = 'Error de guardar'
        detalles = self.get_object().ConfirmacionVentaDetalle_confirmacion_venta.all()
        for detalle in detalles:
            producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
            if not producto.producto_sunat:
                error_codigo_sunat = True

        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        try:
            return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['factura_venta'].id})
        except:
            return reverse_lazy('comprobante_venta_app:factura_venta_inicio')

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            self.object = self.get_object()

            serie_comprobante = SeriesComprobante.objects.por_defecto(ContentType.objects.get_for_model(FacturaVenta))

            factura_venta = FacturaVenta.objects.create(
                confirmacion=self.object,
                sociedad = self.object.sociedad,
                serie_comprobante = serie_comprobante,
                cliente = self.object.cliente,
                cliente_interlocutor = self.object.cliente_interlocutor,
                moneda = self.object.moneda,
                tipo_cambio = self.object.tipo_cambio,
                tipo_venta = self.object.tipo_venta,
                condiciones_pago = self.object.condiciones_pago,
                descuento_global = self.object.descuento_global,
                total_descuento = self.object.total_descuento,
                total_anticipo = self.object.total_anticipo,
                total_gravada = self.object.total_gravada,
                total_inafecta = self.object.total_inafecta,
                total_exonerada = self.object.total_exonerada,
                total_igv = self.object.total_igv,
                total_gratuita = self.object.total_gratuita,
                total_otros_cargos = self.object.otros_cargos,
                total = self.object.total,
                observaciones = self.object.observacion,
                slug = slug_aleatorio(FacturaVenta),
                created_by=self.request.user,
                updated_by=self.request.user,
            )
        
            FacturaVentaDetalle.objects.create(
                item=1,
                unidad=Unidad.objects.get(unidad_sunat='NIU'),
                descripcion_documento="DETALLE DEL PRIMER ANTICIPO",
                cantidad=1,
                precio_unitario_sin_igv=self.object.total_gravada,
                precio_unitario_con_igv=self.object.total,
                precio_final_con_igv=self.object.total,
                sub_total=self.object.total_gravada,
                igv=self.object.total_igv,
                total=self.object.total,
                codigo_producto_sunat='10000000',
                factura_venta=factura_venta,
                created_by=self.request.user,
                updated_by=self.request.user,
            )
            self.kwargs['factura_venta'] = factura_venta

            registro_guardar(self.object, self.request)
            self.object.save()

            messages.success(request, MENSAJE_GENERAR_FACTURA)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaAnticipoCrearView, self).get_context_data(**kwargs)
        context['accion'] = 'Generar'
        context['titulo'] = 'Factura de Venta Anticipada'
        context['texto'] = '¿Seguro que desea generar la Factura de Venta Anticipada?'
        context['item'] = str(self.object.cliente) 
        return context


class FacturaVentaAnticipoRegularizarCrearView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.add_facturaventa')
    model = ConfirmacionVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_codigo_sunat = False
        context['titulo'] = 'Error de guardar'
        detalles = self.get_object().ConfirmacionVentaDetalle_confirmacion_venta.all()
        for detalle in detalles:
            producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
            if not producto.producto_sunat:
                error_codigo_sunat = True

        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        try:
            return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['factura_venta'].id})
        except:
            return reverse_lazy('comprobante_venta_app:factura_venta_inicio')

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            self.object = self.get_object()

            factura_anticipada = self.object.FacturaVenta_confirmacion.get(
                estado = 4,
            )
            
            detalles = self.object.ConfirmacionVentaDetalle_confirmacion_venta.all()

            serie_comprobante = SeriesComprobante.objects.por_defecto(ContentType.objects.get_for_model(FacturaVenta))

            factura_venta = FacturaVenta.objects.create(
                confirmacion=self.object,
                sociedad = self.object.sociedad,
                serie_comprobante = serie_comprobante,
                cliente = self.object.cliente,
                cliente_interlocutor = self.object.cliente_interlocutor,
                moneda = self.object.moneda,
                tipo_cambio = self.object.tipo_cambio,
                tipo_venta = self.object.tipo_venta,
                condiciones_pago = self.object.condiciones_pago,
                descuento_global = self.object.descuento_global,
                total_descuento = self.object.total_descuento,
                total_anticipo = self.object.total_anticipo,
                total_gravada = self.object.total_gravada,
                total_inafecta = self.object.total_inafecta,
                total_exonerada = self.object.total_exonerada,
                total_igv = self.object.total_igv,
                total_gratuita = self.object.total_gratuita,
                total_otros_cargos = self.object.otros_cargos,
                total = self.object.total,
                observaciones = self.object.observacion,
                slug = slug_aleatorio(FacturaVenta),
                created_by=self.request.user,
                updated_by=self.request.user,
            )

            contador = 0
            for detalle in detalles:
                contador += 1
                producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
                factura_venta_detalle = FacturaVentaDetalle.objects.create(
                    item=detalle.item,
                    content_type=detalle.content_type,
                    id_registro=detalle.id_registro,
                    unidad=producto.unidad_base,
                    descripcion_documento=producto.descripcion_documento,
                    cantidad=detalle.cantidad_confirmada,
                    precio_unitario_sin_igv=detalle.precio_unitario_sin_igv,
                    precio_unitario_con_igv=detalle.precio_unitario_con_igv,
                    precio_final_con_igv=detalle.precio_final_con_igv,
                    descuento=detalle.descuento,
                    sub_total=detalle.sub_total,
                    tipo_igv=detalle.tipo_igv,
                    igv=detalle.igv,
                    total=detalle.total,
                    codigo_producto_sunat=producto.producto_sunat.codigo,
                    factura_venta=factura_venta,
                    created_by=self.request.user,
                    updated_by=self.request.user,
                )

            FacturaVentaDetalle.objects.create(
                item=contador + 1,
                unidad=Unidad.objects.get(unidad_sunat='NIU'),
                descripcion_documento="FACTURA ANTICIPADA %s-%s" % (factura_anticipada.serie_comprobante.serie, numeroXn(factura_anticipada.numero_factura, 6)),
                cantidad=1,
                precio_unitario_sin_igv=factura_anticipada.total_gravada,
                precio_unitario_con_igv=factura_anticipada.total,
                precio_final_con_igv=factura_anticipada.total,
                sub_total=factura_anticipada.total_gravada,
                igv=factura_anticipada.total_igv,
                total=factura_anticipada.total,
                codigo_producto_sunat='20000000',
                anticipo_regularizacion = True,
                anticipo_documento_serie = factura_anticipada.serie_comprobante,
                anticipo_documento_numero = factura_anticipada.numero_factura,
                factura_venta=factura_venta,
                created_by=self.request.user,
                updated_by=self.request.user,
            )
            self.kwargs['factura_venta'] = factura_venta

            registro_guardar(self.object, self.request)
            self.object.save()

            messages.success(request, MENSAJE_GENERAR_FACTURA)
            # return HttpResponseRedirect(reverse_lazy('cotizacion_app:confirmacion_ver', kwargs={'id_confirmacion':self.object.id}))
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaAnticipoRegularizarCrearView, self).get_context_data(**kwargs)
        context['accion'] = 'Regularizar'
        context['titulo'] = 'Factura de Venta'
        context['texto'] = '¿Seguro que desea regularizar la Factura de Venta?'
        context['item'] = str(self.object.cliente) 
        return context


class FacturaVentaSerieUpdateView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/formulario generico.html"
    form_class = FacturaVentaSerieForm
    success_url = '.'
    
    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaSerieUpdateView, self).get_context_data(**kwargs)
        context['accion'] = 'Seleccionar'
        context['titulo'] = 'Serie'
        return context


class FacturaVentaDireccionView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('clientes.change_cliente')
    model = Cliente
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_tipo_documento = False
        context['titulo'] = 'Error de dirección'
        if self.get_object().tipo_documento!='6':
            error_tipo_documento = True

        if error_tipo_documento:
            context['texto'] = 'El cliente debe tener RUC.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)
    
    def get_success_url(self, **kwargs):
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['id_factura']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            cliente = self.get_object()
            consulta = cliente.consulta_direccion
            cliente.direccion_fiscal = consulta['direccion']
            cliente.ubigeo = consulta['ubigeo']
            cliente.save()
            messages.success(request, 'Operación exitosa: Dirección actualizada')
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaDireccionView, self).get_context_data(**kwargs)
        context['accion'] = 'Actualizar'
        context['titulo'] = 'Dirección'
        context['texto'] = f'Dirección anterior: {self.get_object().direccion_anterior}'
        context['item'] = f'Nueva Dirección: {self.get_object().direccion_nueva}'
        return context


class FacturaVentaGuardarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_tipo_cambio = False
        context['titulo'] = 'Error de guardar'
        if len(TipoCambio.objects.filter(fecha=datetime.today()))==0:
            error_tipo_cambio = True

        if error_tipo_cambio:
            context['texto'] = 'Ingrese un tipo de cambio para hoy.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            fecha_vencimiento = generarDeuda(obj, self.request)

            obj.fecha_emision = date.today()
            obj.fecha_vencimiento = fecha_vencimiento
            obj.estado = 2
            obj.numero_factura = FacturaVenta.objects.nuevo_numero(obj)
            if obj.confirmacion.estado == 1:
                obj.confirmacion.estado = 2
            elif obj.confirmacion.estado == 5:
                obj.confirmacion.estado = 4
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaGuardarView, self).get_context_data(**kwargs)
        context['accion'] = 'Guardar'
        context['titulo'] = 'Factura de Venta'
        context['texto'] = '¿Seguro de guardar la Factura de Venta?'
        context['item'] = self.get_object()
        return context


class FacturaVentaAnularView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/form generico.html"
    
    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('cotizacion_app:confirmacion_ver', kwargs={'id_confirmacion':self.request.session['id_confirmacion']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            eliminar = eliminarDeuda(obj)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            obj.estado = 3
            if obj.confirmacion.estado == 2:
                obj.confirmacion.estado = 1
            elif obj.confirmacion.estado == 4:
                obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            if obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj)) == 'MANUAL':
                obj.save()
                return HttpResponseRedirect(self.get_success_url())

            return super().delete(request, *args, **kwargs)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaAnularView, self).get_context_data(**kwargs)
        obj = self.get_object()
        self.request.session['id_confirmacion'] = obj.confirmacion.id
        context['accion'] = 'Anular'
        context['titulo'] = 'Factura de Venta'
        context['texto'] = '¿Seguro de anular la Factura de Venta?'
        context['item'] = self.get_object()
        return context


class FacturaVentaNubeFactEnviarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_nubefact = False
        error_codigo_sunat = False
        context['titulo'] = 'Error de guardar'
        if self.get_object().serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(self.get_object().sociedad, ContentType.objects.get_for_model(self.get_object())) == 'MANUAL':
            error_nubefact = True
        for detalle in self.get_object().FacturaVentaDetalle_factura_venta.all():
            if not detalle.codigo_producto_sunat:
                error_codigo_sunat = True

        if error_nubefact:
            context['texto'] = 'No hay una ruta para envío a NubeFact'
            return render(request, 'includes/modal sin permiso.html', context)
        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            respuesta = factura_nubefact(obj, self.request.user)
            if respuesta.error:
                obj.estado = 6
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            elif respuesta.aceptado:
                obj.estado = 4
                if obj.confirmacion.estado == 1:
                    obj.confirmacion.estado = 2
                elif obj.confirmacion.estado == 5:
                    obj.confirmacion.estado = 4
            else:
                obj.estado = 5
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaNubeFactEnviarView, self).get_context_data(**kwargs)
        context['accion'] = 'Enviar'
        context['titulo'] = 'Factura de Venta a NubeFact'
        context['texto'] = '¿Seguro de enviar la Factura de Venta a NubeFact?'
        context['item'] = self.get_object()
        return context


class FacturaVentaNubeFactAnularView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/formulario generico.html"
    form_class = FacturaVentaAnularForm

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_fecha = False
        context['titulo'] = 'Error de guardar'
        if (date.today() - self.get_object().fecha_emision).days > 0:
            error_fecha = True

        if error_fecha:
            context['texto'] = 'No se puede anular, realizar nota de crédito.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['pk']})

    @transaction.atomic
    def form_valid(self, form):
        sid = transaction.savepoint()
        try:
            respuesta = anular_nubefact(form.instance, self.request.user)
            if respuesta.error:
                form.instance.estado = 6
            else:
                form.instance.estado = 3
            registro_guardar(form.instance, self.request)
            eliminar = eliminarDeuda(form.instance)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            return super().form_valid(form)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaNubeFactAnularView, self).get_context_data(**kwargs)
        context['accion'] = 'Anular'
        context['titulo'] = 'Factura de Venta a NubeFact'
        return context


class FacturaVentaNubefactRespuestaDetailView(PermissionRequiredMixin, BSModalReadView):
    permission_required = ('comprobante_venta.view_facturaventa')
    model = FacturaVenta
    template_name = "comprobante_venta/nubefact_respuesta.html"

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(FacturaVentaNubefactRespuestaDetailView, self).get_context_data(**kwargs)
        context['titulo'] = 'Movimientos Nubefact'
        context['movimientos'] = NubefactRespuesta.objects.respuestas(self.get_object())
        return context


class FacturaVentaNubefactConsultarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/form generico.html"
    
    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            respuesta = consultar_documento(obj, self.request.user)
            if respuesta.error:
                obj.estado = 6
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            elif respuesta.aceptado:
                obj.estado = 4
                if obj.confirmacion.estado == 1:
                    obj.confirmacion.estado = 2
                elif obj.confirmacion.estado == 5:
                    obj.confirmacion.estado = 4
            else:
                obj.estado = 5
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())
    
    def get_context_data(self, **kwargs):
        context = super(FacturaVentaNubefactConsultarView, self).get_context_data(**kwargs)
        context['accion'] = 'Consultar'
        context['titulo'] = 'Factura de Venta a NubeFact'
        context['texto'] = '¿Seguro de consultar la Factura de Venta a NubeFact?'
        context['item'] = self.get_object()
        return context


class FacturaVentaEliminarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.delete_facturaventa')
    model = FacturaVenta
    template_name = "includes/form generico.html"
    
    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('cotizacion_app:confirmacion_ver', kwargs={'id_confirmacion':self.request.session['id_confirmacion']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            eliminar = eliminarDeuda(obj)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            return super().delete(request, *args, **kwargs)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaEliminarView, self).get_context_data(**kwargs)
        obj = self.get_object()
        self.request.session['id_confirmacion'] = obj.confirmacion.id
        context['accion'] = 'Eliminar'
        context['titulo'] = 'Factura de Venta'
        context['texto'] = '¿Seguro de eliminar la Factura de Venta?'
        context['item'] = self.get_object()
        return context


class FacturaVentaObservacionUpdateView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVenta
    template_name = "includes/formulario generico.html"
    form_class = FacturaVentaObservacionForm

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:factura_venta_detalle', kwargs={'id_factura_venta':self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaObservacionUpdateView, self).get_context_data(**kwargs)
        context['titulo'] = "Observaciones"
        context['accion'] = "Actualizar"
        return context


class FacturaVentaDetalleUpdateView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_facturaventa')
    model = FacturaVentaDetalle
    template_name = "includes/formulario generico.html"
    form_class = FacturaVentaDetalleForm
    success_url = '.'

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.descripcion_documento = "%s (Cotización %s%s)" % (form.instance.descripcion_documento, form.instance.factura_venta.confirmacion.sociedad.abreviatura, numeroXn(form.instance.factura_venta.confirmacion.cotizacion_venta.numero_cotizacion, 6))
        respuesta = calculos_linea(form.instance.cantidad, form.instance.total, form.instance.total, igv(), form.instance.tipo_igv)
        form.instance.precio_unitario_sin_igv = respuesta['precio_unitario_sin_igv']
        form.instance.precio_unitario_con_igv = form.instance.total
        form.instance.precio_final_con_igv = form.instance.total
        form.instance.descuento = respuesta['descuento']
        form.instance.sub_total = respuesta['subtotal']
        form.instance.igv = respuesta['igv']
        form.instance.total = respuesta['total']
        registro_guardar(form.instance, self.request)
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(FacturaVentaDetalleUpdateView, self).get_context_data(**kwargs)
        context['accion'] = 'Actualizar'
        context['titulo'] = 'Descripción Contingencia'
        return context


def FacturaVentaJsonView(request):
    if request.is_ajax():
        term = request.GET.get('term')
        data = []
        filtro_razon_social = Q(cliente__razon_social__unaccent__icontains=term.split(' ')[0])
        if term.split(' ')[0].isnumeric():
            filtro_numero = Q(numero_factura=term.split(' ')[0])
        else:
            filtro_numero = False
        filtro_serie = Q(serie_comprobante__serie=term.split(' ')[0])
        for palabra in term.split(' ')[1:]:
            filtro_razon_social = filtro_razon_social | Q(cliente__razon_social__unaccent__icontains=palabra)
            if palabra.isnumeric():
                if filtro_numero:
                    filtro_numero = filtro_numero | Q(numero_factura=palabra)
                else:
                    filtro_numero = Q(numero_factura=palabra)
            filtro_serie = filtro_serie | Q(serie_comprobante__serie=palabra)
        buscar = FacturaVenta.objects.filter(
                estado=4,
            ).filter(
                filtro_razon_social | filtro_serie
            )
        if filtro_numero:
            buscar = buscar.filter(filtro_numero)

        for factura in buscar:
            data.append({
                'id' : factura.id,
                'nombre' : factura.__str__(),
                })
        return JsonResponse(data, safe=False)


###########################################################################################

class BoletaVentaListView(PermissionRequiredMixin, FormView):
    permission_required = ('comprobante_venta.view_boletaventa')
    # model = BoletaVenta
    template_name = 'comprobante_venta/boleta_venta/inicio.html'
    form_class = BoletaVentaBuscarForm
    success_url = '.'

    def get_form_kwargs(self):
        kwargs = super(BoletaVentaListView, self).get_form_kwargs()
        kwargs['filtro_numero_boleta'] = self.request.GET.get('numero_boleta')
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_cliente'] = self.request.GET.get('cliente')
        kwargs['filtro_fecha_emision'] = self.request.GET.get('fecha_emision')
        kwargs['filtro_estado'] = self.request.GET.get('estado')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaListView,self).get_context_data(**kwargs)
        boleta_venta = BoletaVenta.objects.all()

        filtro_numero_boleta = self.request.GET.get('numero_boleta')
        filtro_sociedad = self.request.GET.get('sociedad')
        filtro_cliente = self.request.GET.get('cliente')
        filtro_fecha_emision = self.request.GET.get('fecha_emision')
        filtro_estado = self.request.GET.get('estado')

        contexto_filtro = []

        if filtro_cliente:
            condicion = Q(cliente = filtro_cliente)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("cliente=" + filtro_cliente)

        if filtro_fecha_emision:
            condicion = Q(fecha_emision = datetime.strptime(filtro_fecha_emision, "%Y-%m-%d").date())
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("fecha_emision=" + filtro_fecha_emision)

        if filtro_sociedad:
            condicion = Q(sociedad = filtro_sociedad)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("sociedad=" + filtro_sociedad)

        if filtro_numero_boleta:
            condicion = Q(numero_boleta = filtro_numero_boleta)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("numero_boleta=" + filtro_numero_boleta)

        if filtro_estado:
            condicion = Q(estado = filtro_estado)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("estado=" + filtro_estado)
        
        context['contexto_filtro'] = "&".join(contexto_filtro)

        context['pagina_filtro'] = ""
        if self.request.GET.get('page'):
            if context['contexto_filtro']:
                context['pagina_filtro'] = f'&page={self.request.GET.get("page")}'
            else:
                context['pagina_filtro'] = f'page={self.request.GET.get("page")}'
        context['contexto_filtro'] = '?' + context['contexto_filtro']

        objectsxpage = 25 # Show 25 objects per page.

        if len(boleta_venta) > objectsxpage:
            paginator = Paginator(boleta_venta, objectsxpage)
            page_number = self.request.GET.get('page')
            boleta_venta = paginator.get_page(page_number)
   
        context['contexto_pagina'] = boleta_venta
        return context


def BoletaVentaTabla(request):
    data = dict()
    if request.method == 'GET':
        template = 'comprobante_venta/boleta_venta/inicio_tabla.html'
        context = {}
        boleta_venta = BoletaVenta.objects.all()

        filtro_numero_boleta = request.GET.get('numero_boleta')
        filtro_sociedad = request.GET.get('sociedad')
        filtro_cliente = request.GET.get('cliente')
        filtro_fecha_emision = request.GET.get('fecha_emision')
        filtro_estado = request.GET.get('estado')

        contexto_filtro = []

        if filtro_cliente:
            condicion = Q(cliente = filtro_cliente)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("cliente=" + filtro_cliente)

        if filtro_fecha_emision:
            condicion = Q(fecha_emision = datetime.strptime(filtro_fecha_emision, "%Y-%m-%d").date())
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("fecha_emision=" + filtro_fecha_emision)

        if filtro_sociedad:
            condicion = Q(sociedad = filtro_sociedad)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("sociedad=" + filtro_sociedad)

        if filtro_numero_boleta:
            condicion = Q(numero_boleta = filtro_numero_boleta)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("numero_boleta=" + filtro_numero_boleta)

        if filtro_estado:
            condicion = Q(estado = filtro_estado)
            boleta_venta = boleta_venta.filter(condicion)
            contexto_filtro.append("estado=" + filtro_estado)

        context['contexto_filtro'] = "&".join(contexto_filtro)

        context['pagina_filtro'] = ""
        if request.GET.get('page'):
            if context['contexto_filtro']:
                context['pagina_filtro'] = f'&page={request.GET.get("page")}'
            else:
                context['pagina_filtro'] = f'page={request.GET.get("page")}'
        context['contexto_filtro'] = '?' + context['contexto_filtro']

        objectsxpage = 25 # Show 25 objects per page.

        if len(boleta_venta) > objectsxpage:
            paginator = Paginator(boleta_venta, objectsxpage)
            page_number = request.GET.get('page')
            boleta_venta = paginator.get_page(page_number)
        
        context['contexto_pagina'] = boleta_venta 

        data['table'] = render_to_string(
            template,
            context,
            request=request
        )
        return JsonResponse(data)


class BoletaVentaDetalleView(PermissionRequiredMixin, TemplateView):
    permission_required = ('comprobante_venta.view_boletaventa')
    template_name = "comprobante_venta/boleta_venta/detalle.html"

    def get_context_data(self, **kwargs):
        obj = BoletaVenta.objects.get(id = kwargs['id_boleta_venta'])

        materiales = None
        try:
            materiales = obj.BoletaVentaDetalle_boleta_venta.all()

            for material in materiales:
                material.material = material.content_type.get_object_for_this_type(id = material.id_registro)
        except:
            pass

        tipo_cambio_hoy = TipoCambio.objects.tipo_cambio_venta(date.today())
        if obj.confirmacion:
            tipo_cambio = obj.confirmacion.tipo_cambio.tipo_cambio_venta
        else:
            tipo_cambio = Decimal('1.00')

        context = super(BoletaVentaDetalleView, self).get_context_data(**kwargs)
        context['boleta'] = obj
        context['materiales'] = materiales
        context['tipo_cambio_hoy'] = tipo_cambio_hoy
        context['tipo_cambio'] = tipo_cambio
        tipo_cambio = tipo_de_cambio(tipo_cambio, tipo_cambio_hoy)
        context['totales'] = obtener_totales(obj)
        if obj.serie_comprobante:
            context['nubefact_acceso'] = obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj))
        url_nubefact = NubefactRespuesta.objects.respuesta(obj)
        if url_nubefact:
            context['url_nubefact'] = url_nubefact
        if obj.nubefact:
            context['url_nubefact'] = obj.nubefact
        context['respuestas_nubefact'] = NubefactRespuesta.objects.respuestas(obj)
      
        return context

def BoletaVentaDetalleVerTabla(request, id_boleta_venta):
    data = dict()
    if request.method == 'GET':
        template = 'comprobante_venta/boleta_venta/detalle_tabla.html'
        obj = BoletaVenta.objects.get(id=id_boleta_venta)
        
        tipo_cambio_hoy = TipoCambio.objects.tipo_cambio_venta(date.today())
        if obj.confirmacion:
            tipo_cambio = obj.confirmacion.tipo_cambio.tipo_cambio_venta
        else:
            tipo_cambio = Decimal('1.00')
 
        materiales = None
        try:
            materiales = obj.BoletaVentaDetalle_boleta_venta.all()

            for material in materiales:
                material.material = material.content_type.get_object_for_this_type(id = material.id_registro)
        except:
            pass

        context = {}
        context['boleta'] = obj
        context['materiales'] = materiales
        context['tipo_cambio_hoy'] = tipo_cambio_hoy
        context['tipo_cambio'] = tipo_cambio
        tipo_cambio = tipo_de_cambio(tipo_cambio, tipo_cambio_hoy)
        context['totales'] = obtener_totales(obj)
        if obj.serie_comprobante:
            context['nubefact_acceso'] = obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj))
        url_nubefact = NubefactRespuesta.objects.respuesta(obj)
        if url_nubefact:
            context['url_nubefact'] = url_nubefact
        if obj.nubefact:
            context['url_nubefact'] = obj.nubefact
        context['respuestas_nubefact'] = NubefactRespuesta.objects.respuestas(obj)

        data['table'] = render_to_string(
            template,
            context,
            request=request
        )
        return JsonResponse(data)

class BoletaVentaCrearView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.add_boletaventa')
    model = ConfirmacionVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_codigo_sunat = False
        error_cuotas = False
        context['titulo'] = 'Error de guardar'
        detalles = self.get_object().ConfirmacionVentaDetalle_confirmacion_venta.all()
        for detalle in detalles:
            producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
            if not producto.producto_sunat:
                error_codigo_sunat = True
        
        if self.get_object().tipo_venta == 2 and not self.get_object().ConfirmacionVentaCuota_confirmacion_venta.all():
            error_cuotas = True

        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat.'
            return render(request, 'includes/modal sin permiso.html', context)
        if error_cuotas:
            context['texto'] = 'Falta ingresar las cuotas.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        try:
            return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['boleta_venta'].id})
        except:
            return reverse_lazy('comprobante_venta_app:boleta_venta_inicio')

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            self.object = self.get_object()

            detalles = self.object.ConfirmacionVentaDetalle_confirmacion_venta.all()

            serie_comprobante = SeriesComprobante.objects.por_defecto(ContentType.objects.get_for_model(BoletaVenta))

            boleta_venta = BoletaVenta.objects.create(
                confirmacion=self.object,
                sociedad = self.object.sociedad,
                serie_comprobante = serie_comprobante,
                cliente = self.object.cliente,
                cliente_interlocutor = self.object.cliente_interlocutor,
                moneda = self.object.moneda,
                tipo_cambio = self.object.tipo_cambio,
                tipo_venta = self.object.tipo_venta,
                condiciones_pago = self.object.condiciones_pago,
                descuento_global = self.object.descuento_global,
                total_descuento = self.object.total_descuento,
                total_anticipo = self.object.total_anticipo,
                total_gravada = self.object.total_gravada,
                total_inafecta = self.object.total_inafecta,
                total_exonerada = self.object.total_exonerada,
                total_igv = self.object.total_igv,
                total_gratuita = self.object.total_gratuita,
                total_otros_cargos = self.object.otros_cargos,
                total = self.object.total,
                observaciones = self.object.observacion,
                slug = slug_aleatorio(BoletaVenta),
                created_by=self.request.user,
                updated_by=self.request.user,
            )

            for detalle in detalles:
                producto = detalle.content_type.get_object_for_this_type(id = detalle.id_registro)
                boleta_venta_detalle = BoletaVentaDetalle.objects.create(
                    item=detalle.item,
                    content_type=detalle.content_type,
                    id_registro=detalle.id_registro,
                    unidad=producto.unidad_base,
                    descripcion_documento=producto.descripcion_documento,
                    cantidad=detalle.cantidad_confirmada,
                    precio_unitario_sin_igv=detalle.precio_unitario_sin_igv,
                    precio_unitario_con_igv=detalle.precio_unitario_con_igv,
                    precio_final_con_igv=detalle.precio_final_con_igv,
                    descuento=detalle.descuento,
                    sub_total=detalle.sub_total,
                    tipo_igv=detalle.tipo_igv,
                    igv=detalle.igv,
                    total=detalle.total,
                    codigo_producto_sunat=producto.producto_sunat.codigo,
                    boleta_venta=boleta_venta,
                    created_by=self.request.user,
                    updated_by=self.request.user,
                )
            self.kwargs['boleta_venta'] = boleta_venta
                
            registro_guardar(self.object, self.request)
            self.object.save()

            messages.success(request, MENSAJE_GENERAR_BOLETA)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaCrearView, self).get_context_data(**kwargs)
        context['accion'] = 'Generar'
        context['titulo'] = 'Boleta de venta'
        context['texto'] = '¿Seguro que desea generar la Boleta de venta?'
        context['item'] = str(self.object.cliente) 
        return context


class BoletaVentaSerieUpdateView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/formulario generico.html"
    form_class = BoletaVentaSerieForm
    success_url = '.'

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(BoletaVentaSerieUpdateView, self).get_context_data(**kwargs)
        context['accion'] = 'Seleccionar'
        context['titulo'] = 'Serie'
        return context


class BoletaVentaDireccionView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('clientes.change_cliente')
    model = Cliente
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_tipo_documento = False
        context['titulo'] = 'Error de dirección'
        if self.get_object().tipo_documento!='6':
            error_tipo_documento = True

        if error_tipo_documento:
            context['texto'] = 'El cliente debe tener RUC.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)
    
    def get_success_url(self, **kwargs):
        return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['id_boleta']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            cliente = self.get_object()
            consulta = cliente.consulta_direccion
            cliente.direccion_fiscal = consulta['direccion']
            cliente.ubigeo = consulta['ubigeo']
            cliente.save()
            messages.success(request, 'Operación exitosa: Dirección actualizada')
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaDireccionView, self).get_context_data(**kwargs)
        context['accion'] = 'Actualizar'
        context['titulo'] = 'Dirección'
        context['texto'] = f'Dirección anterior: {self.get_object().direccion_anterior}'
        context['item'] = f'Nueva Dirección: {self.get_object().direccion_nueva}'
        return context
    

class BoletaVentaGuardarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_tipo_cambio = False
        context['titulo'] = 'Error de guardar'
        if len(TipoCambio.objects.filter(fecha=datetime.today()))==0:
            error_tipo_cambio = True

        if error_tipo_cambio:
            context['texto'] = 'Ingrese un tipo de cambio para hoy.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            fecha_vencimiento = generarDeuda(obj, self.request)

            obj.fecha_emision = date.today()
            obj.fecha_vencimiento = fecha_vencimiento
            obj.estado = 2
            obj.numero_boleta = BoletaVenta.objects.nuevo_numero(obj)
            if obj.confirmacion.estado == 1:
                obj.confirmacion.estado = 2
            elif obj.confirmacion.estado == 5:
                obj.confirmacion.estado = 4
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaGuardarView, self).get_context_data(**kwargs)
        context['accion'] = 'Guardar'
        context['titulo'] = 'Boleta de Venta'
        context['texto'] = '¿Seguro de guardar la Boleta de Venta?'
        context['item'] = self.get_object()
        return context
    

class BoletaVentaAnularView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/form generico.html"
    form_class = BoletaVentaAnularForm

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('cotizacion_app:confirmacion_ver', kwargs={'id_confirmacion':self.request.session['id_confirmacion']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            eliminar = eliminarDeuda(obj)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            obj.estado = 3
            if obj.confirmacion.estado == 2:
                obj.confirmacion.estado = 1
            elif obj.confirmacion.estado == 4:
                obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            if obj.serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(obj.sociedad, ContentType.objects.get_for_model(obj)) == 'MANUAL':
                obj.save()
                return HttpResponseRedirect(self.get_success_url())

            return super().delete(request, *args, **kwargs)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())
    
    def get_context_data(self, **kwargs):
        context = super(BoletaVentaAnularView, self).get_context_data(**kwargs)
        obj = self.get_object()
        self.request.session['id_confirmacion'] = obj.confirmacion.id
        context['accion'] = 'Anular'
        context['titulo'] = 'Boleta de Venta'
        context['texto'] = '¿Seguro de anular la Boleta de Venta?'
        context['item'] = self.get_object()
        return context
    

class BoletaVentaNubeFactEnviarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_nubefact = False
        error_codigo_sunat = False
        context['titulo'] = 'Error de guardar'
        if self.get_object().serie_comprobante.NubefactSerieAcceso_serie_comprobante.acceder(self.get_object().sociedad, ContentType.objects.get_for_model(self.get_object())) == 'MANUAL':
            error_nubefact = True
        for detalle in self.get_object().BoletaVentaDetalle_boleta_venta.all():
            if not detalle.codigo_producto_sunat:
                error_codigo_sunat = True

        if error_nubefact:
            context['texto'] = 'No hay una ruta para envío a NubeFact'
            return render(request, 'includes/modal sin permiso.html', context)
        if error_codigo_sunat:
            context['texto'] = 'Hay productos sin Código de Sunat'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            respuesta = boleta_nubefact(obj, self.request.user)
            if respuesta.error:
                obj.estado = 6
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            elif respuesta.aceptado:
                obj.estado = 4
                if obj.confirmacion.estado == 1:
                    obj.confirmacion.estado = 2
                elif obj.confirmacion.estado == 5:
                    obj.confirmacion.estado = 4
            else:
                obj.estado = 5
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaNubeFactEnviarView, self).get_context_data(**kwargs)
        context['accion'] = 'Enviar'
        context['titulo'] = 'Boleta de Venta a NubeFact'
        context['texto'] = '¿Seguro de enviar la Boleta de Venta a NubeFact?'
        context['item'] = self.get_object()
        return context
    

class BoletaVentaNubeFactAnularView(PermissionRequiredMixin, BSModalUpdateView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/formulario generico.html"
    form_class = BoletaVentaAnularForm

    def dispatch(self, request, *args, **kwargs):
        context = {}
        error_fecha = False
        context['titulo'] = 'Error de guardar'
        if (date.today() - self.get_object().fecha_emision).days > 0:
            error_fecha = True

        if error_fecha:
            context['texto'] = 'No se puede anular, realizar nota de crédito.'
            return render(request, 'includes/modal sin permiso.html', context)
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['pk']})

    @transaction.atomic
    def form_valid(self, form):
        sid = transaction.savepoint()
        try:
            respuesta = anular_nubefact(form.instance, self.request.user)
            if respuesta.error:
                form.instance.estado = 6
            else:
                form.instance.estado = 3
            registro_guardar(form.instance, self.request)
            eliminar = eliminarDeuda(form.instance)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            return super().form_valid(form)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaNubeFactAnularView, self).get_context_data(**kwargs)
        context['accion'] = 'Anular'
        context['titulo'] = 'Boleta de Venta a NubeFact'
        return context


class BoletaVentaNubefactRespuestaDetailView(PermissionRequiredMixin, BSModalReadView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "comprobante_venta/nubefact_respuesta.html"

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(BoletaVentaNubefactRespuestaDetailView, self).get_context_data(**kwargs)
        context['titulo'] = 'Movimientos Nubefact'
        context['movimientos'] = NubefactRespuesta.objects.respuestas(self.get_object())
        return context


class BoletaVentaNubefactConsultarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('comprobante_venta_app:boleta_venta_detalle', kwargs={'id_boleta_venta':self.kwargs['pk']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            respuesta = consultar_documento(obj, self.request.user)
            if respuesta.error:
                obj.estado = 6
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            elif respuesta.aceptado:
                obj.estado = 4
                if obj.confirmacion.estado == 1:
                    obj.confirmacion.estado = 2
                elif obj.confirmacion.estado == 5:
                    obj.confirmacion.estado = 4
            else:
                obj.estado = 5
                if obj.confirmacion.estado == 2:
                    obj.confirmacion.estado = 1
                elif obj.confirmacion.estado == 4:
                    obj.confirmacion.estado = 5
            registro_guardar(obj.confirmacion, self.request)
            obj.confirmacion.save()
            registro_guardar(obj, self.request)
            obj.save()
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())
    
    def get_context_data(self, **kwargs):
        context = super(BoletaVentaNubefactConsultarView, self).get_context_data(**kwargs)
        context['accion'] = 'Consultar'
        context['titulo'] = 'Boleta de Venta a NubeFact'
        context['texto'] = '¿Seguro de consultar la Boleta de Venta a NubeFact?'
        context['item'] = self.get_object()
        return context


class BoletaVentaEliminarView(PermissionRequiredMixin, BSModalDeleteView):
    permission_required = ('comprobante_venta.change_boletaventa')
    model = BoletaVenta
    template_name = "includes/form generico.html"

    def dispatch(self, request, *args, **kwargs):
        if not self.has_permission():
            return render(request, 'includes/modal sin permiso.html')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self) -> str:
        return reverse_lazy('cotizacion_app:confirmacion_ver', kwargs={'id_confirmacion':self.request.session['id_confirmacion']})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        sid = transaction.savepoint()
        try:
            obj = self.get_object()
            eliminar = eliminarDeuda(obj)
            if eliminar:
                messages.success(self.request, MENSAJE_ELIMINAR_DEUDA)
            else:
                messages.warning(self.request, MENSAJE_ERROR_ELIMINAR_DEUDA)
            return super().delete(request, *args, **kwargs)
        except Exception as ex:
            transaction.savepoint_rollback(sid)
            registrar_excepcion(self, ex, __file__)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super(BoletaVentaEliminarView, self).get_context_data(**kwargs)
        obj = self.get_object()
        self.request.session['id_confirmacion'] = obj.confirmacion.id
        context['accion'] = 'Eliminar'
        context['titulo'] = 'Boleta de Venta'
        context['texto'] = '¿Seguro de eliminar la Boleta de Venta?'
        context['item'] = self.get_object()
        return context


def BoletaVentaJsonView(request):
    if request.is_ajax():
        term = request.GET.get('term')
        data = []
        filtro_razon_social = Q(cliente__razon_social__unaccent__icontains=term.split(' ')[0])
        if term.split(' ')[0].isnumeric():
            filtro_numero = Q(numero_boleta=term.split(' ')[0])
        else:
            filtro_numero = False
        filtro_serie = Q(serie_comprobante__serie=term.split(' ')[0])
        for palabra in term.split(' ')[1:]:
            filtro_razon_social = filtro_razon_social | Q(cliente__razon_social__unaccent__icontains=palabra)
            if palabra.isnumeric():
                if filtro_numero:
                    filtro_numero = filtro_numero | Q(numero_factura=palabra)
                else:
                    filtro_numero = Q(numero_factura=palabra)
            filtro_serie = filtro_serie | Q(serie_comprobante__serie=palabra)
        buscar = BoletaVenta.objects.filter(
                estado=4,
            ).filter(
                filtro_razon_social | filtro_serie
            )
        if filtro_numero:
            buscar = buscar.filter(filtro_numero)

        for boleta in buscar:
            data.append({
                'id' : boleta.id,
                'nombre' : boleta.__str__(),
                })
        return JsonResponse(data, safe=False)


###################################################################################################

class DescargarComprobantes(FormView):
    template_name = 'comprobante_venta/descargar_comprobantes.html'
    form_class = DescargarComprobantesForm
    success_url = '.'

    def get(self, request, *args, **kwargs):
        if request.method == 'GET':
            inicio = request.GET.get('inicio')
            final = request.GET.get('final')
            if inicio and final:
                filtro_inicio = get_datetime(inicio)
                filtro_final = get_datetime(final)

                facturas = FacturaVenta.objects.filter(
                    Q(created_at__gte=filtro_inicio) & Q(created_at__lte=filtro_final)
                )
                boletas = BoletaVenta.objects.filter(
                    Q(created_at__gte=filtro_inicio) & Q(created_at__lte=filtro_final)
                )

                wb = Workbook()
                hoja = wb.active
                hoja.title = 'sheet'

                fila = 1
                for factura in facturas:
                    hoja.cell(fila, 1).value = factura.descripcion
                    hoja.cell(fila, 2).value = factura.cliente.__str__()
                    fila += 1
                for boleta in boletas:
                    hoja.cell(fila, 1).value = boleta.descripcion
                    hoja.cell(fila, 2).value = boleta.cliente.__str__()
                    fila += 1
                
                nombre_archivo = "Comprobantes.xlsx"
                respuesta = HttpResponse(content_type='application/ms-excel')
                content = "attachment; filename ={0}".format(nombre_archivo)
                respuesta['content-disposition']= content
                wb.save(respuesta)
                return respuesta
        return super().get(request, *args, **kwargs)