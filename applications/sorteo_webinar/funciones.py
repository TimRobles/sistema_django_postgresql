import openpyxl

from applications.sorteo_webinar.models import Participante

def llenar_datos(excel):
    wb = openpyxl.load_workbook(excel)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row = 2, min_col = 2):
        if row[0].value and row[1].value and row[2].value and row[3].value:
            try:
                Participante.objects.create(
                    nombre_completo = row[0].value,
                    documento = str(row[1].value).replace('.0', ''),
                    telefono = str(row[2].value).replace('.0', ''),
                    correo = row[3].value,
                )
            except:
                pass