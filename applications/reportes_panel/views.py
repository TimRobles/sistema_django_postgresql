from django.shortcuts import render
from applications.importaciones import *
from django.db.models.functions import Substr
from applications.reportes_panel.forms import ReportesPanelFiltrosForm, ReporteProductoClienteVentasFiltrosForm, ReporteUbigeoVentasFiltrosForm
from applications.material.forms import Material
from applications.comprobante_venta.models import FacturaVentaDetalle, FacturaVenta
from applications.reportes.funciones import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import Reference, Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
import pandas as pd
import requests
from plotly.offline import plot
import plotly.express as px
import plotly.graph_objects as go

def consultar_marca_ventas(filtro_sociedad, filtro_marca, filtro_fecha_inicio, filtro_fecha_fin):
    sql_base = '''SELECT
            MAX(mm.id) as id,
            MAX(mm.descripcion_corta) as material_texto,
            ROUND(SUM(mam.cantidad*mam.signo_factor_multiplicador),2) as total_stock
        FROM movimiento_almacen_movimientosalmacen mam
        LEFT JOIN material_material mm
            ON mm.id=mam.id_registro_producto AND mam.content_type_producto_id='52'
        LEFT JOIN movimiento_almacen_tipostock mats
            ON mam.tipo_stock_id=mats.id
        WHERE mats.codigo NOT IN (
            1, 2, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25) '''
    sql_where_marca = '''AND mm.marca_id='%s' ''' %(filtro_marca) if filtro_marca else ''
    sql_final = ''' GROUP BY mm.id ORDER BY mm.descripcion_corta ; '''
    sql_stock = sql_base + sql_where_marca + sql_final
    query_info = Material.objects.raw(sql_stock)
    dict_stock = {}
    for fila in query_info:
        dict_stock[fila.id] = fila.total_stock
        
    sql_base = ''' SELECT 
            MAX(mm.id) AS id,
            MAX(mma.nombre) AS descrip_marca,
            MAX(mm.descripcion_corta) AS material_texto,
            ROUND(SUM(cvfd.cantidad),2) AS venta_total,
            MIN(cvf.fecha_emision) AS primera_venta,
            MAX(cvf.fecha_emision) AS ultima_venta,
            COUNT(cvf.id) AS nro_ventas,
            '' AS stock_total
        FROM material_material mm
        LEFT JOIN material_marca mma
            ON mm.marca_id=mma.id
        LEFT JOIN comprobante_venta_facturaventadetalle cvfd
            ON cvfd.content_type_id='52' AND mm.id=cvfd.id_registro
        LEFT JOIN comprobante_venta_facturaventa cvf
            ON cvf.id=cvfd.factura_venta_id
        LEFT JOIN sociedad_sociedad ss
            ON ss.id=cvf.sociedad_id
        WHERE (cvf.estado='4' OR cvf.estado IS NULL) '''
    sql_where_sociedad = '''AND (cvf.sociedad_id IS NULL OR cvf.sociedad_id='%s') ''' %(filtro_sociedad) if filtro_sociedad else ''
    sql_where_marca = '''AND mm.marca_id='%s' ''' %(filtro_marca) if filtro_marca else ''
    sql_where_fechainicio = '''AND cvf.fecha_emision>='%s' ''' %(filtro_fecha_inicio) if filtro_fecha_inicio else ''
    sql_where_fechafin = '''AND cvf.fecha_emision<='%s' ''' %(filtro_fecha_fin) if filtro_fecha_fin else ''
    sql_final = '''GROUP BY mm.id ORDER BY 3 ; '''
    sql_1 = sql_base + sql_where_sociedad + sql_where_marca + sql_where_fechainicio + sql_where_fechafin + sql_final
    
    query_info = Material.objects.raw(sql_1)
    list_info_facturas = []
    dict_info_facturas = {}
    for fila in query_info:
        list_temp = []
        stock_total = dict_stock[fila.id] if fila.id in dict_stock else 0.00
        venta_total = float(fila.venta_total) if fila.venta_total else float(0.00)
        ultima_venta = fila.ultima_venta if fila.ultima_venta else '--'
        list_temp.extend([
            fila.descrip_marca,
            fila.material_texto,
            venta_total,
            ultima_venta,
            stock_total
            ])
        dict_info_facturas[fila.id] = list_temp
        list_info_facturas.append(list_temp)

    sql_base = ''' SELECT 
            MAX(mm.id) AS id,
            MAX(mma.nombre) AS descrip_marca,
            MAX(mm.descripcion_corta) AS material_texto,
            ROUND(SUM(cvbd.cantidad),2) AS venta_total,
            MIN(cvb.fecha_emision) AS primera_venta,
            MAX(cvb.fecha_emision) AS ultima_venta,
            COUNT(cvb.id) AS nro_ventas,
            '' AS stock_total
        FROM material_material mm
        LEFT JOIN material_marca mma
            ON mm.marca_id=mma.id
        LEFT JOIN comprobante_venta_boletaventadetalle cvbd
            ON cvbd.content_type_id='52' AND mm.id=cvbd.id_registro
        LEFT JOIN comprobante_venta_boletaventa cvb
            ON cvb.id=cvbd.boleta_venta_id
        LEFT JOIN sociedad_sociedad ss
            ON ss.id=cvb.sociedad_id
        WHERE (cvb.estado='4' OR cvb.estado IS NULL) '''
    sql_where_sociedad = '''AND (cvb.sociedad_id IS NULL OR cvb.sociedad_id='%s') ''' %(filtro_sociedad) if filtro_sociedad else ''
    sql_where_marca = '''AND mm.marca_id='%s' ''' %(filtro_marca) if filtro_marca else ''
    sql_where_fechainicio = '''AND cvb.fecha_emision>='%s' ''' %(filtro_fecha_inicio) if filtro_fecha_inicio else ''
    sql_where_fechafin = '''AND cvb.fecha_emision<='%s' ''' %(filtro_fecha_fin) if filtro_fecha_fin else ''
    sql_final = '''GROUP BY mm.id ORDER BY 3 ; '''
    sql_2 = sql_base + sql_where_sociedad + sql_where_marca + sql_where_fechainicio + sql_where_fechafin + sql_final

    # sql = sql_1 + sql_2
    query_info = Material.objects.raw(sql_2)
    list_info_boletas = []
    for fila in query_info:
        list_temp = []
        stock_total = dict_stock[fila.id] if fila.id in dict_stock else 0.00
        venta_total = float(fila.venta_total) if fila.venta_total else float(0.00)
        ultima_venta = fila.ultima_venta if fila.ultima_venta else '--'
        list_temp.extend([
            fila.descrip_marca,
            fila.material_texto,
            venta_total,
            ultima_venta,
            stock_total
            ])
        if fila.id in dict_info_facturas:
            dict_info_facturas[fila.id][2] += venta_total
            if dict_info_facturas[fila.id][3] == '--':
                dict_info_facturas[fila.id][3] = ultima_venta
            elif ultima_venta != '--':
                if dict_info_facturas[fila.id][3] < ultima_venta:
                    dict_info_facturas[fila.id][3] = ultima_venta
        else:
            dict_info_facturas[fila.id] = list_temp
        list_info_boletas.append(list_temp)

    list_info = dict_info_facturas.values()
    return list_info


class ReportesMarcaVentasView(FormView):
    template_name = "reportes_panel/marca_ventas/inicio.html"
    form_class = ReportesPanelFiltrosForm
    success_url = '.' 

    def get_form_kwargs(self):
        kwargs = super(ReportesMarcaVentasView, self).get_form_kwargs()
        kwargs['filtro_marca'] = self.request.GET.get('marca')
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_fecha_inicio'] = self.request.GET.get('fecha_inicio')
        kwargs['filtro_fecha_fin'] = self.request.GET.get('fecha_fin')
        global global_sociedad, global_fecha_inicio, global_fecha_fin, global_marca
        global_marca = self.request.GET.get('marca')
        global_sociedad = self.request.GET.get('sociedad')
        global_fecha_inicio = self.request.GET.get('fecha_inicio')
        global_fecha_fin = self.request.GET.get('fecha_fin')
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super(ReportesMarcaVentasView, self).get_context_data(**kwargs)
        contexto_filtros = []
        filtro_marca = self.request.GET.get('marca')
        filtro_sociedad = self.request.GET.get('sociedad')
        filtro_fecha_inicio = self.request.GET.get('fecha_inicio')
        filtro_fecha_fin = self.request.GET.get('fecha_fin')
        contexto_filtros.append(f"filtro_marca={filtro_marca}")
        contexto_filtros.append(f"filtro_sociedad={filtro_sociedad}")
        contexto_filtros.append(f"filtro_fecha_inicio={filtro_fecha_inicio}")
        contexto_filtros.append(f"filtro_fecha_fin={filtro_fecha_fin}")
        context["contexto_filtros"] = "&".join(contexto_filtros)
        context['contexto_marca_ventas'] = consultar_marca_ventas(filtro_sociedad, filtro_marca, filtro_fecha_inicio, filtro_fecha_fin)
        return context


class ReportesPanelTemplateView(TemplateView):
    template_name = "reportes_panel/inicio.html"
    

class ReporteExcelMarcaVentas(TemplateView):

    def get(self, request, *args,**kwargs):
        filtro_marca = self.request.GET.get('filtro_marca')
        filtro_sociedad = self.request.GET.get('filtro_sociedad')
        filtro_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        filtro_fecha_fin = self.request.GET.get('filtro_fecha_fin')

        def reporte_marca_ventas():
            wb = Workbook()
            hoja = wb.active
            hoja.title = 'MARCA VS VENTAS'

            hoja.append(('MARCA', 'DESCRIPCIÓN', 'TOTAL VENDIDO', 'FECHA ÚLTIMA VENTA', 'STOCK'))

            color_relleno = rellenoSociedad(filtro_sociedad)

            col_range = hoja.max_column
            for col in range(1, col_range + 1):
                cell_header = hoja.cell(1, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA

            info = consultar_marca_ventas(filtro_sociedad, filtro_marca, filtro_fecha_inicio, filtro_fecha_fin)
            for producto in info:
                hoja.append(producto)

            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO

            ajustarColumnasSheet(hoja)
            return wb

        if filtro_sociedad:
            query_sociedad = Sociedad.objects.filter(id = int(filtro_sociedad))[0]
            abreviatura = query_sociedad.abreviatura
        else:
            abreviatura = 'MPL-MCA'
        wb = reporte_marca_ventas()
        nombre_archivo = "Reporte Marca vs Ventas - " + abreviatura + " - " + datetime.now().strftime("%Y-%m-%d") + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


def consultar_productocliente_ventas(filtro_sociedad, filtro_producto, filtro_cliente):
    if filtro_sociedad and filtro_producto and filtro_cliente:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__sociedad_id=filtro_sociedad) & Q(factura_venta__cliente_id=filtro_cliente) & Q(content_type_id='52') & Q(id_registro=filtro_producto)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total'), id_cliente=Max('factura_venta__cliente_id')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    elif filtro_sociedad and filtro_producto:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__sociedad_id=filtro_sociedad) & Q(content_type_id='52') & Q(id_registro=filtro_producto)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total'), id_cliente=Max('factura_venta__cliente_id')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    elif filtro_sociedad and filtro_cliente:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__sociedad_id=filtro_sociedad) & Q(factura_venta__cliente_id=filtro_cliente)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')            
    elif filtro_cliente and filtro_producto:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__cliente_id=filtro_cliente) & Q(content_type_id='52') & Q(id_registro=filtro_producto)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total'), id_cliente=Max('factura_venta__cliente_id')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    elif filtro_sociedad:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__sociedad_id=filtro_sociedad)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    elif filtro_producto:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(content_type_id='52') & Q(id_registro=filtro_producto)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total'), id_cliente=Max('factura_venta__cliente_id')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    elif filtro_cliente:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4') & Q(factura_venta__cliente_id=filtro_cliente)).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')   
    else:
        info = FacturaVentaDetalle.objects.filter(Q(factura_venta__estado='4')).values_list(
                'factura_venta__cliente__razon_social',
                'factura_venta__cliente__numero_documento',
                ).annotate(ultima_fecha_venta=Max('factura_venta__fecha_emision'), nro_compras=Count('factura_venta'), total_compra=Sum('total')).order_by('-total_compra','-nro_compras','factura_venta__cliente__razon_social')
    return info


class ProductoClienteVentasView(FormView):
    template_name = "reportes_panel/producto_cliente_ventas/inicio.html"
    form_class = ReporteProductoClienteVentasFiltrosForm
    success_url = '.' 

    def get_form_kwargs(self):
        kwargs = super(ProductoClienteVentasView, self).get_form_kwargs()
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_producto'] = self.request.GET.get('producto')
        kwargs['filtro_cliente'] = self.request.GET.get('cliente')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(ProductoClienteVentasView, self).get_context_data(**kwargs)
        contexto_filtros = []
        filtro_producto = self.request.GET.get('producto')
        filtro_sociedad = self.request.GET.get('sociedad')
        filtro_cliente = self.request.GET.get('cliente')
        contexto_filtros.append(f"filtro_producto={filtro_producto}")
        contexto_filtros.append(f"filtro_sociedad={filtro_sociedad}")
        contexto_filtros.append(f"filtro_cliente={filtro_cliente}")
        context["contexto_filtros"] = "&".join(contexto_filtros)
        context['contexto_productocliente_ventas'] = consultar_productocliente_ventas(filtro_sociedad, filtro_producto, filtro_cliente)
        return context


class ProductoClienteVentasDetalle(DetailView):
    model = Cliente
    template_name = 'reportes_panel/producto_cliente_ventas/detalle_inicio.html'
    context_object_name = 'contexto_datos_cliente'

    def get_context_data(self, **kwargs):
        cliente = Cliente.objects.get(id = self.kwargs['pk'])
        context = super(ProductoClienteVentasDetalle, self).get_context_data(**kwargs)
        filtro_producto = self.request.GET.get('filtro_producto')
        filtro_sociedad = self.request.GET.get('filtro_sociedad')

        Q_filtro_estado = Q(factura_venta__estado='4')
        Q_filtro_cliente = Q(factura_venta__cliente=cliente)
        Q_filtro_producto = Q(content_type_id='52') & Q(id_registro=filtro_producto)
        Q_filtro_sociedad = Q(factura_venta__sociedad_id=filtro_sociedad)

        if filtro_sociedad:
            context['contexto_detalle_ventas_productocliente'] = FacturaVentaDetalle.objects.filter(Q_filtro_estado & Q_filtro_cliente & Q_filtro_producto & Q_filtro_sociedad)
        else:
            context['contexto_detalle_ventas_productocliente'] = FacturaVentaDetalle.objects.filter(Q_filtro_estado & Q_filtro_cliente & Q_filtro_producto)
        return context


def consultar_ubigeo_ventas(filtro_sociedad):  
    if filtro_sociedad:
        filtro_sociedad_factura = ''' AND cvf.sociedad_id='%s' ''' %(filtro_sociedad)
        filtro_sociedad_boleta = ''' AND cvb.sociedad_id='%s' ''' %(filtro_sociedad)
    else:
        filtro_sociedad_factura = ''
        filtro_sociedad_boleta = ''
    
    sql = '''SELECT 
            MAX(id) AS id, departamento_nombre, SUM(monto_total_facturado) AS monto_total_facturado 
            FROM (
                (SELECT 
                        MAX(cvf.id) as id,
                        dgdp.nombre AS departamento_nombre,
                        '--'  AS provincia_nombre,
                        '--'  AS distrito_nombre,
                        SUM(cvf.total) AS monto_total_facturado
                    FROM comprobante_venta_facturaventa cvf
                    LEFT JOIN clientes_cliente cc
                        ON cc.id=cvf.cliente_id
                    LEFT JOIN datos_globales_departamento dgdp
                        ON dgdp.codigo=SUBSTRING(cc.ubigeo,1,2)
                    WHERE cvf.estado='4' AND dgdp.nombre IS NOT NULL %s
                    GROUP BY dgdp.nombre
                    ORDER BY monto_total_facturado DESC)
                UNION
                (SELECT 
                        MAX(cvb.id) as id,
                        dgdp.nombre AS departamento_nombre,
                        '--'  AS provincia_nombre,
                        '--'  AS distrito_nombre,
                        SUM(cvb.total) AS monto_total_facturado
                    FROM comprobante_venta_boletaventa cvb
                    LEFT JOIN clientes_cliente cc
                        ON cc.id=cvb.cliente_id
                    LEFT JOIN datos_globales_departamento dgdp
                        ON dgdp.codigo=SUBSTRING(cc.ubigeo,1,2)
                    WHERE cvb.estado='4' AND dgdp.nombre IS NOT NULL %s
                    GROUP BY dgdp.nombre
                    ORDER BY monto_total_facturado DESC)) tabla_ventas
        GROUP BY departamento_nombre
        ORDER BY monto_total_facturado DESC ; ''' %(filtro_sociedad_factura, filtro_sociedad_boleta)
    query_info = FacturaVenta.objects.raw(sql)
    list_info = []
    for fila in query_info:
        list_temp = []
        list_temp.extend([
            fila.departamento_nombre,
            # fila.provincia_nombre,
            # fila.distrito_nombre,
            fila.monto_total_facturado,
            ])
        list_info.append(list_temp)
    
    return list_info
    # https://stackoverflow.com/questions/39291372/django-orm-join-without-foreign-keys-and-without-raw-queries


class UbigeoVentasView(FormView):
    template_name = "reportes_panel/ubigeo_ventas/inicio.html"
    form_class = ReporteUbigeoVentasFiltrosForm
    success_url = '.' 

    def get_form_kwargs(self):
        kwargs = super(UbigeoVentasView, self).get_form_kwargs()
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        # kwargs['filtro_producto'] = self.request.GET.get('producto')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(UbigeoVentasView, self).get_context_data(**kwargs)
        contexto_filtros = []
        filtro_sociedad = self.request.GET.get('sociedad')
        # filtro_producto = self.request.GET.get('producto')
        contexto_filtros.append(f"filtro_sociedad={filtro_sociedad}")
        # contexto_filtros.append(f"filtro_producto={filtro_producto}")
        context["contexto_filtros"] = "&".join(contexto_filtros)
        data_ubigeo_ventas = consultar_ubigeo_ventas(filtro_sociedad)
        context['contexto_ubigeo_ventas'] = data_ubigeo_ventas
        list_encabezado = [['DEPARTAMENTO', 'INGRESOS']]
        table_data = list_encabezado + data_ubigeo_ventas
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
        try:
            context['contexto_ubigeo_ventas_grafico'] = grafico_departamentos(df)
        except:
            context['contexto_ubigeo_ventas_grafico'] = None
        return context


def grafico_departamentos(df):
    # url_git = "https://github.com/juaneladio/peru-geojson/blob/master/peru_departamental_simple.geojson"
    url_datos = "https://raw.githubusercontent.com/juaneladio/peru-geojson/master/peru_departamental_simple.geojson"
    departamentos = requests.get(url_datos).json()
    # https://plotly.github.io/plotly.py-docs/generated/plotly.express.choropleth_mapbox.html
    fig = px.choropleth_mapbox(
        data_frame=df,
        geojson=departamentos,
        locations='DEPARTAMENTO',
        featureidkey='properties.NOMBDEP',
        hover_name='DEPARTAMENTO',
        color='INGRESOS',
        color_continuous_scale='solar', # Reds, magma, spectral # color_continuous_scale=px.colors.sequential.Plasma,
        width=650,height=850,
        center={"lat":-9.1900, "lon": -75.0152},
        mapbox_style='carto-positron', # 'open-street-map', 'white-bg', 'carto-positron', 'carto-darkmatter', 'stamen- terrain', 'stamen-toner', 'stamen-watercolor'
        zoom=4.5,
        )
    return plot(fig, auto_open=False, output_type='div')
    # fig = go.Figure(fig)
    # return fig.to_html(full_html=False)