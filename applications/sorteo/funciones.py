import openpyxl

from applications.sorteo.models import Ticket

def llenar_datos(excel, sorteo):
    wb = openpyxl.load_workbook(excel)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row = 0, min_col = 0):
        dato_uno = None
        dato_dos = None
        dato_tres = None
        dato_cuatro = None
        try:
            dato_uno = row[0].value
        except:
            pass
        try:
            dato_dos = row[1].value
        except:
            pass
        try:
            dato_tres = row[2].value
        except:
            pass
        try:
            dato_cuatro = row[3].value
        except:
            pass
        Ticket.objects.create(
            dato_uno = dato_uno,
            dato_dos = dato_dos,
            dato_tres = dato_tres,
            dato_cuatro = dato_cuatro,
            sorteo = sorteo,
        )
        