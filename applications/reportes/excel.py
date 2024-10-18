from django.shortcuts import render
from django.db import models
import time
from datetime import datetime, timedelta
from decimal import Decimal
from django.contrib.contenttypes.models import ContentType
from applications.comprobante_compra.models import ComprobanteCompraPI, ComprobanteCompraPIDetalle
from applications.funciones import numero_str, numeroXn
from applications.importaciones import *
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import Reference, Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from applications.material.models import Material
from applications.movimiento_almacen.managers import ordenar_movimientos
from applications.movimiento_almacen.models import MovimientosAlmacen
from applications.recepcion_compra.models import RecepcionCompra
from applications.reportes.funciones import *
from applications.datos_globales.models import CuentaBancariaSociedad, DocumentoFisico, Moneda, TipoCambioSunat
from applications.home.templatetags.funciones_propias import get_enlace_nubefact, redondear
from applications.comprobante_venta.models import BoletaVenta, BoletaVentaDetalle, FacturaVenta, FacturaVentaDetalle
from applications.nota.models import NotaCredito
from applications.crm.models import ClienteCRMDetalle
from applications.clientes.models import CorreoInterlocutorCliente, HistorialEstadoCliente, RepresentanteLegalCliente, TelefonoInterlocutorCliente
from applications.datos_globales.models import Departamento, Moneda
from applications.cotizacion.models import CotizacionVenta, PrecioListaMaterial
from applications.cobranza.models import Cuota, Deuda, Ingreso, Nota, Pago
from applications.reportes.data_resumen_ingresos_anterior import*
from applications.sede.models import Sede
from django.db.models import Subquery, OuterRef, Max, F, Sum, Case, When, Value
from django.db.models.functions import Concat, Coalesce, Cast, LPad

####################################################  FACTURACIÓN VS ASESOR COMERCIAL  ####################################################   

def ReporteFacturacionAsesorComercial(fecha_inicio, fecha_fin, asesor_comercial):
    list_comprobante = TIPO_COMPROBANTE 
    DICT_TIPO_COMPROBANTE = dict(list_comprobante)

    moneda_base = Moneda.objects.get(simbolo='$')

    if asesor_comercial:
        asesores = get_user_model().objects.filter(id=asesor_comercial)
    else:
        asesores = get_user_model().objects.filter(id__in = [cotizacion.vendedor.id for cotizacion in CotizacionVenta.objects.all()])

    wb = Workbook()
    data_resumen = []
    total_resumen = Decimal('0.00')
    count = 0
    for asesor_comercial in asesores:
        fila_resumen = []
        data = []
        data_nota = []

        notas_credito = NotaCredito.objects.filter(estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin)

        total_nota= Decimal('0.00')
        total_nota_factura = Decimal('0.00')
        total_factura = Decimal('0.00')
        for factura in FacturaVenta.objects.filter(confirmacion__cotizacion_venta__vendedor=asesor_comercial, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin).order_by("fecha_emision"):
            if factura.moneda == moneda_base:
                total_factura += factura.total
            else:
                total_factura += (factura.total / factura.tipo_cambio).quantize(Decimal('0.01'))

            fila = []
            if asesor_comercial.first_name:
                fila.append(f"{asesor_comercial.first_name} {asesor_comercial.last_name}")
            else:
                fila.append(asesor_comercial.username)
            fila.append(factura.fecha_emision.strftime('%d/%m/%Y'))
            fila.append(DICT_TIPO_COMPROBANTE[factura.tipo_comprobante])
            fila.append(f"{str(factura.serie_comprobante.serie)} - {str(factura.numero_factura).zfill(6)}")
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(factura.total))))
            data.append(fila)

        for factura_nota in FacturaVenta.objects.filter(confirmacion__cotizacion_venta__vendedor=asesor_comercial):
            fila_nota = []
            for nota in notas_credito:
                if nota.content_type_documento==DocumentoFisico.objects.get(modelo=ContentType.objects.get_for_model(FacturaVenta)) and nota.id_registro_documento==factura_nota.id:
                    if nota.moneda == moneda_base:
                        total_nota_factura += nota.total
                    else:
                        total_nota_factura += (nota.total / nota.tipo_cambio).quantize(Decimal('0.01'))
                    if asesor_comercial.first_name:
                        fila_nota.append(f"{asesor_comercial.first_name} {asesor_comercial.last_name}")
                    else:
                        fila_nota.append(asesor_comercial.username)
                    fila_nota.append(nota.fecha_emision.strftime('%d/%m/%Y'))
                    fila_nota.append(DICT_TIPO_COMPROBANTE[nota.tipo_comprobante])
                    fila_nota.append(f"{str(nota.serie_comprobante.serie)} - {str(nota.numero_nota).zfill(6)}")
                    fila_nota.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(nota.total))))
                    data_nota.append(fila_nota)

        total_nota_boleta = Decimal('0.00')
        total_boleta = Decimal('0.00')
        for boleta in BoletaVenta.objects.filter(confirmacion__cotizacion_venta__vendedor=asesor_comercial, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin).order_by("fecha_emision"):
            if boleta.moneda == moneda_base:
                total_boleta += boleta.total
            else:
                total_boleta += (boleta.total / boleta.tipo_cambio).quantize(Decimal('0.01'))

            fila = []
            if asesor_comercial.first_name:
                fila.append(f"{asesor_comercial.first_name} {asesor_comercial.last_name}")
            else:
                fila.append(asesor_comercial.username)
            fila.append(boleta.fecha_emision.strftime('%d/%m/%Y'))
            fila.append(DICT_TIPO_COMPROBANTE[boleta.tipo_comprobante])
            fila.append(f"{str(boleta.serie_comprobante.serie)} - {str(boleta.numero_boleta).zfill(6)}")
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(boleta.total))))
            data.append(fila)

        for boleta_nota in BoletaVenta.objects.filter(confirmacion__cotizacion_venta__vendedor=asesor_comercial):
            fila_nota = []
            for nota in notas_credito:
                if nota.content_type_documento==DocumentoFisico.objects.get(modelo=ContentType.objects.get_for_model(BoletaVenta)) and nota.id_registro_documento==boleta_nota.id:
                    if nota.moneda == moneda_base:
                        total_nota_boleta += nota.total
                    else:
                        total_nota_boleta += (nota.total / nota.tipo_cambio).quantize(Decimal('0.01'))

                    if asesor_comercial.first_name:
                        fila_nota.append(f"{asesor_comercial.first_name} {asesor_comercial.last_name}")
                    else:
                        fila_nota.append(asesor_comercial.username)
                    fila_nota.append(nota.fecha_emision.strftime('%d/%m/%Y'))
                    fila_nota.append(DICT_TIPO_COMPROBANTE[nota.tipo_comprobante])
                    fila_nota.append(f"{str(nota.serie_comprobante.serie)} - {str(nota.numero_nota).zfill(6)}")
                    fila_nota.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(nota.total))))
                    data_nota.append(fila_nota)

        total = total_factura + total_boleta
        total_nota = total_nota_factura + total_nota_boleta
        total_total = total - total_nota
        list_encabezado = [
            'Asesor Comercial',
            'Fecha Emisión',
            'Tipo Comprobante',
            'N° Comprobante',
            'Total',
            ]
        
        if data != []:
            if count != 0:
                hoja = wb.create_sheet(str(asesor_comercial.username))

            else:
                hoja = wb.active
                hoja.title = str(asesor_comercial.username)
                # count += 1 

            data.sort(key = lambda i: i[1], reverse=False)
            # data.append(["","","SUBTOTAL",("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))])
            
            hoja.append(tuple(list_encabezado))

            col_range = hoja.max_column  # get max columns in the worksheet
            # cabecera de la tabla
            for col in range(1, col_range + 1):
                cell_header = hoja.cell(1, col)
                cell_header.fill = RELLENO_EXCEL
                cell_header.font = NEGRITA

            for fila in data:
                hoja.append(fila)

            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if col == 1 or col==3:
                        row[col].alignment = ALINEACION_CENTRO
                    elif col == 4:
                        row[col].alignment = ALINEACION_DERECHA
            if data_nota != []:
                hoja.append(('', '', '','SUBTOTAL', ("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))))
            else:
                hoja.append(('', '', '','TOTAL', ("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))))

            nueva_fila = hoja.max_row

            for col in range(1, col_range + 1):
                if col == 4:
                    cell_header = hoja.cell(nueva_fila, col)
                    cell_header.fill = RELLENO_EXCEL
                    cell_header.font = NEGRITA
                elif col == 5:
                    cell_header = hoja.cell(nueva_fila, col)
                    cell_header.font = NEGRITA

            for i in range(hoja.max_row):
                if i >= nueva_fila-1:
                    row = list(hoja.rows)[i]
                    for col in range(hoja.max_column):
                        if col > 2:
                            row[col].border = BORDE_DELGADO
                            if col==3:
                                row[col].alignment = ALINEACION_CENTRO
                            elif col == 4:
                                row[col].alignment = ALINEACION_DERECHA

            if data_nota != []:
                data_nota.sort(key = lambda i: i[1], reverse=False)
                # data_nota.append(["","","SUBTOTAL",("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_nota))))])
                # cabecera de la tabla
                hoja.append(('',)) # Crea la fila del encabezado con los títulos
                hoja.append(tuple(list_encabezado))

                nueva_fila = hoja.max_row

                for col in range(1, col_range + 1):
                    cell_header = hoja.cell(nueva_fila, col)
                    cell_header.fill = RELLENO_EXCEL
                    cell_header.font = NEGRITA

                for fila in data_nota:
                    hoja.append(fila) # Crea la fila del encabezado con los títulos

                for i in range(hoja.max_row):
                    if i >= nueva_fila-1:
                        row = list(hoja.rows)[i]
                        for col in range(hoja.max_column):
                            row[col].border = BORDE_DELGADO
                            if col == 1 or col==3:
                                row[col].alignment = ALINEACION_CENTRO
                            elif col == 4:
                                row[col].alignment = ALINEACION_DERECHA

                hoja.append(('', '', '','SUBTOTAL', ("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_nota)))))) 

                nueva_fila = hoja.max_row

                for col in range(1, col_range + 1):
                    if col == 4:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = RELLENO_EXCEL
                        cell_header.font = NEGRITA
                    elif col == 5:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.font = NEGRITA

                for i in range(hoja.max_row):
                    if i >= nueva_fila-1:
                        row = list(hoja.rows)[i]
                        for col in range(hoja.max_column):
                            if col > 2:
                                row[col].border = BORDE_DELGADO
                                if col==3:
                                    row[col].alignment = ALINEACION_CENTRO
                                elif col == 4:
                                    row[col].alignment = ALINEACION_DERECHA

                hoja.append(('',)) # Crea la fila del encabezado con los títulos
                hoja.append(('', '', '','TOTAL', ("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_total)))))) 

                nueva_fila = hoja.max_row

                for col in range(1, col_range + 1):
                    if col == 4:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = RELLENO_EXCEL
                        cell_header.font = NEGRITA
                    elif col == 5:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.font = NEGRITA

                for i in range(hoja.max_row):
                    if i >= nueva_fila-1:
                        row = list(hoja.rows)[i]
                        for col in range(hoja.max_column):
                            if col > 2:
                                row[col].border = BORDE_DELGADO
                                if col==3:
                                    row[col].alignment = ALINEACION_CENTRO
                                elif col == 4:
                                    row[col].alignment = ALINEACION_DERECHA
            
            total_resumen += total_total

            if asesor_comercial.first_name:
                fila_resumen.append(f"{asesor_comercial.first_name} {asesor_comercial.last_name}")
            else:
                fila_resumen.append(asesor_comercial.username)
            fila_resumen.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_total))))
            fila_resumen.append(total_total)
            data_resumen.append(fila_resumen)

        ajustarColumnasSheet(hoja)
        count += 1

    list_encabezado_resumen = [
        'Asesor Comercial',
        'Total',
        'Porcentaje',
        ]
    hoja = wb.create_sheet('RESUMEN')
    
    hoja.append(tuple(list_encabezado_resumen))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA
    
    data_resumen.sort(key = lambda i: i[2], reverse=True)
    for fila in data_resumen:
        fila[2] = ("%s %s" % (intcomma(redondear((Decimal(fila[2])/total_resumen)*100)), '%'))

    for fila in data_resumen:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 1 or col == 2:
                row[col].alignment = ALINEACION_DERECHA

    ajustarColumnasSheet(hoja)
    return wb

####################################################  VENTAS VS DEPARTAMENTO  ####################################################   

def ReporteVentasDepartamento(titulo, fecha_inicio, fecha_fin, departamento_codigo):

    if departamento_codigo:
        departamentos = Departamento.objects.filter(codigo=departamento_codigo)
    else:
        departamentos = Departamento.objects.all()

    moneda_base = Moneda.objects.get(simbolo='$')

    wb = Workbook()
    data_resumen = []
    total_resumen = Decimal('0.00')
    count = 0
    for departamento in departamentos:
        fila_resumen = []
        list_encabezado = [
            'Cliente',
            'Documento',
            'Dirección',
            'Total Facturas',
            'Total Boletas',
            'Total Notas de Crédito',
            'Total',
            ]

        if count != 0:
            hoja = wb.create_sheet(str(departamento.nombre))
        else:
            hoja = wb.active
            hoja.title = str(departamento.nombre)

        hoja.append(tuple(list_encabezado))

        col_range = hoja.max_column  # get max columns in the worksheet
        # cabecera de la tabla
        for col in range(1, col_range + 1):
            cell_header = hoja.cell(1, col)
            cell_header.fill = RELLENO_EXCEL
            cell_header.font = NEGRITA
        
        total_total = Decimal('0.00')

        data = []
        
        for cliente in Cliente.objects.filter(distrito__provincia__departamento=departamento):
            total = Decimal('0.00')
            total_factura = Decimal('0.00')
            for factura in FacturaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin):
                if factura.moneda == moneda_base:
                    total_factura += factura.total
                else:
                    total_factura += (factura.total / factura.tipo_cambio).quantize(Decimal('0.01'))

            total_boleta = Decimal('0.00')
            for boleta in BoletaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin):
                if boleta.moneda == moneda_base:
                    total_boleta += boleta.total
                else:
                    total_boleta += (boleta.total / boleta.tipo_cambio).quantize(Decimal('0.01'))

            total_nota_credito = Decimal('0.00')
            for nota_credito in NotaCredito.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin):
                if nota_credito.moneda == moneda_base:
                    total_nota_credito += nota_credito.total
                else:
                    total_nota_credito += (nota_credito.total / nota_credito.tipo_cambio).quantize(Decimal('0.01'))

            total = total_factura + total_boleta - total_nota_credito
            total_total += total
            
            fila = []
            fila.append(cliente.razon_social)
            fila.append(f"{cliente.documento} - {cliente.numero_documento}")
            fila.append(cliente.direccion_fiscal)
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_factura))))
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_boleta))))
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_nota_credito))))
            fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))
            data.append(fila)

        data.sort(key = lambda i: i[6], reverse=True) #Total
        
        for fila in data:
            hoja.append(fila)

        for row in hoja.rows:
            for col in range(hoja.max_column):
                row[col].border = BORDE_DELGADO
                if 3 <= col <= 6:
                    row[col].alignment = ALINEACION_DERECHA

        hoja.append(["","","","","","Total",("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_total))))])

        nueva_fila = hoja.max_row

        for col in range(1, col_range + 1):
            if col == 6:
                cell_header = hoja.cell(nueva_fila, col)
                cell_header.fill = RELLENO_EXCEL
                cell_header.font = NEGRITA
            elif col == 7:
                cell_header = hoja.cell(nueva_fila, col)
                cell_header.font = NEGRITA

        for i in range(hoja.max_row):
            if i >= nueva_fila-1:
                row = list(hoja.rows)[i]
                for col in range(hoja.max_column):
                    if col > 4:
                        row[col].border = BORDE_DELGADO
                        if col==5:
                            row[col].alignment = ALINEACION_CENTRO
                        elif col == 6:
                            row[col].alignment = ALINEACION_DERECHA


        total_resumen += total_total


        fila_resumen.append(departamento.nombre)
        fila_resumen.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total_total))))
        fila_resumen.append(total_total)
        data_resumen.append(fila_resumen)

        hoja.freeze_panes = 'A2'
        ajustarColumnasSheet(hoja)
        count += 1

    list_encabezado_resumen = [
        'Departamento',
        'Total',
        'Porcentaje',
        ]
    hoja = wb.create_sheet('RESUMEN')
    
    hoja.append(tuple(list_encabezado_resumen))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA
    
    data_resumen.sort(key = lambda i: i[2], reverse=True)

    for fila in data_resumen:
        fila[2] = ("%s %s" % (intcomma(redondear((Decimal(fila[2])/total_resumen)*100)), '%'))

    for fila in data_resumen:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 1 or col == 2:
                row[col].alignment = ALINEACION_DERECHA

    ajustarColumnasSheet(hoja)
    return wb

####################################################  CLIENTES CRM VS MEDIOS  ####################################################   

def dataClienteCRM():

    list_estado = ESTADOS_CLIENTE
    DICT_ESTADO = dict(list_estado)
    list_medio = MEDIO
    DICT_MEDIO = dict(list_medio)

    list_encabezado = [
        'Fecha de Registro',
        'Razón Social',
        'RUC',
        'País',
        'Distrito',
        'Fecha de Última Actividad',
        'Medio',
        'Estado',
        'Teléfono',
        'Correo Electrónico',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'Clientes'
    hoja.append(tuple(list_encabezado))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA

    data = []
    
    for cliente in Cliente.objects.all():
        fila = []
        fila.append(cliente.created_at.strftime('%d/%m/%Y'))
        fila.append(cliente.razon_social)
        fila.append(cliente.numero_documento)
        fila.append(str(cliente.pais))
        fila.append(str(cliente.ubigeo_total))
        # cliente_crm_detalle = ClienteCRMDetalle.objects.filter(cliente_crm = cliente.id).order_by('-id')[0]
        # fila.append(cliente_crm_detalle.fecha.strftime('%d/%m/%Y'))
        # fila.append(DICT_MEDIO[cliente.medio])
        # fila.append(DICT_ESTADO[cliente.estado])
        fila.append("")
        fila.append("")
        fila.append("")
        try: 
            representante_legal = RepresentanteLegalCliente.objects.filter(cliente=cliente)[0]
            telefono =  TelefonoInterlocutorCliente.objects.filter(interlocutor=representante_legal.interlocutor)[0]
            fila.append(str(telefono.numero))
            correo = CorreoInterlocutorCliente.objects.filter(interlocutor=representante_legal.interlocutor)[0]
            fila.append(str(correo.correo))    

        except:
            fila.append("")
            fila.append("")
    
        data.append(fila)

    for fila in data:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 0 or col == 7:
                row[col].alignment = ALINEACION_CENTRO

    ajustarColumnasSheet(hoja)
    return wb

def ReporteClienteCRMExcel():

    wb=dataClienteCRM()
    return wb

####################################################  FACTURACIÓN GENERAL  ####################################################   

def dataFacturacionGeneral(fecha_inicio, fecha_fin):
    moneda_base = Moneda.objects.get(simbolo='$')

    list_encabezado = [
        'Razón Social',
        'Nro. Transacciones',
        'Monto Total',
        'Dias desde Registro',
        'Meses desde registro',
        'Años desde registro',
        'Correos Electrónicos',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'Facturación General'
    hoja.append(tuple(list_encabezado))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA

    data = []
    for cliente in Cliente.objects.filter(estado_sunat = 1):
        total = Decimal('0.00')
        total_factura = Decimal('0.00')
        nf = 0
        nb = 0
        nn = 0
        for factura in FacturaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin.date()):
            if factura.moneda == moneda_base:
                total_factura += factura.total
            else:
                total_factura += (factura.total / factura.tipo_cambio).quantize(Decimal('0.01'))
            
            nf += 1

        total_boleta = Decimal('0.00')
        for boleta in BoletaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin.date()):
            if boleta.moneda == moneda_base:
                total_boleta += boleta.total
            else:
                total_boleta += (boleta.total / boleta.tipo_cambio).quantize(Decimal('0.01'))
            
            nb += 1

        total_nota_credito = Decimal('0.00')
        for nota_credito in NotaCredito.objects.filter(cliente=cliente, estado=4, fecha_emision__gte=fecha_inicio, fecha_emision__lte=fecha_fin.date()):
            if nota_credito.moneda == moneda_base:
                total_nota_credito += nota_credito.total
            else:
                total_nota_credito += (nota_credito.total / nota_credito.tipo_cambio).quantize(Decimal('0.01'))
            
            nn += 1

        total = total_factura + total_boleta 
        total_transacciones = nf + nb 
        
        fila = []
        fila.append(cliente.razon_social)
        fila.append(total_transacciones)
        fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))
        fecha_creacion = cliente.created_at.strftime('%Y-%m-%d')
        fecha_registro = datetime.strptime(fecha_creacion, '%Y-%m-%d')
        dias = (fecha_fin - fecha_registro) / timedelta(days=1)
        fila.append(dias)
        meses = (fecha_fin.year - fecha_registro.year)* 12 + (fecha_fin.month - fecha_registro.month)
        fila.append(meses)
        years = (fecha_fin.year - fecha_registro.year) + (fecha_fin.month - fecha_registro.month)/12
        fila.append(years)
        fila.append(cliente.correos)
        data.append(fila)

    for fila in data:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 1:
                row[col].alignment = ALINEACION_DERECHA
            elif col == 2:
                row[col].alignment = ALINEACION_DERECHA
                row[col].number_format = FORMATO_DOLAR
            elif 3 <= col <=5:
                row[col].alignment = ALINEACION_DERECHA
                row[col].number_format = FORMATO_NUMERO


    ajustarColumnasSheet(hoja)
    return wb
    
def ReporteFacturacionGeneral(fecha_inicio, fecha_fin):

    wb=dataFacturacionGeneral(fecha_inicio, fecha_fin)
    return wb

####################################################  COMPORTAMIENTO CLIENTE  ####################################################   

def dataComportamientoCliente(cliente):
    moneda_base = Moneda.objects.get(simbolo='$')

    list_encabezado = [
        'Razón Social',
        'Nro. Transacciones',
        'Monto Total',
        'Dias desde Registro',
        'Meses desde registro',
        'Años desde registro',
        'Correos Electrónicos',
        ]

    color_relleno = rellenoSociedad('None')

    wb = Workbook()
    hoja = wb.active
    hoja.title = "sheet"
    hoja.append(tuple(list_encabezado))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = color_relleno
        cell_header.font = NEGRITA

    # data = []
    # for cliente in Cliente.objects.filter(estado_sunat = 1):
    #     total = Decimal('0.00')
    #     total_factura = Decimal('0.00')
    #     nf = 0
    #     nb = 0
    #     nn = 0
    #     for factura in FacturaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__lte=fecha_cierre.date()):
    #         if factura.moneda == moneda_base:
    #             total_factura += factura.total
    #         else:
    #             total_factura += (factura.total / factura.tipo_cambio).quantize(Decimal('0.01'))
            
    #         nf += 1

    #     total_boleta = Decimal('0.00')
    #     for boleta in BoletaVenta.objects.filter(cliente=cliente, estado=4, fecha_emision__lte=fecha_cierre.date()):
    #         if boleta.moneda == moneda_base:
    #             total_boleta += boleta.total
    #         else:
    #             total_boleta += (boleta.total / boleta.tipo_cambio).quantize(Decimal('0.01'))
            
    #         nb += 1

    #     total_nota_credito = Decimal('0.00')
    #     for nota_credito in NotaCredito.objects.filter(cliente=cliente, estado=4, fecha_emision__lte=fecha_cierre.date()):
    #         if nota_credito.moneda == moneda_base:
    #             total_nota_credito += nota_credito.total
    #         else:
    #             total_nota_credito += (nota_credito.total / nota_credito.tipo_cambio).quantize(Decimal('0.01'))
            
    #         nn += 1

    #     total = total_factura + total_boleta 
    #     total_transacciones = nf + nb 
        
    #     fila = []
    #     fila.append(cliente.razon_social)
    #     fila.append(total_transacciones)
    #     fila.append("%s %s" % (moneda_base.simbolo, intcomma(redondear(total))))
    #     fecha_creacion = cliente.created_at.strftime('%Y-%m-%d')
    #     fecha_registro = datetime.strptime(fecha_creacion, '%Y-%m-%d')
    #     dias = (fecha_cierre - fecha_registro) / timedelta(days=1)
    #     fila.append(dias)
    #     meses = (fecha_cierre.year - fecha_registro.year)* 12 + (fecha_cierre.month - fecha_registro.month)
    #     fila.append(meses)
    #     years = (fecha_cierre.year - fecha_registro.year) + (fecha_cierre.month - fecha_registro.month)/12
    #     fila.append(years)
    #     fila.append(cliente.correos)
    #     data.append(fila)

    # for fila in data:
    #     hoja.append(fila)

    # for row in hoja.rows:
    #     for col in range(hoja.max_column):
    #         row[col].border = BORDE_DELGADO
    #         if col == 1:
    #             row[col].alignment = ALINEACION_DERECHA
    #         elif col == 2:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_DOLAR
    #         elif 3 <= col <=5:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_NUMERO


    ajustarColumnasSheet(hoja)
    return wb
    
def ReporteComportamientoCliente(cliente):

    wb=dataComportamientoCliente(cliente)
    return wb

####################################################  TASA DE CONVERSION A CLIENTE FINAL ####################################################   

def dataEstadosCliente(fecha_inicio, fecha_fin):

    list_estado = ESTADOS_CLIENTE
    DICT_ESTADO = dict(list_estado)
    list_medio = MEDIO
    DICT_MEDIO = dict(list_medio)

    list_encabezado = [
        'Razón Social',
        'Fecha de Registro',
        'Estado Anterior',
        'Fecha de Registro',
        'Estado Actual',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'EstadoClientes'
    hoja.append(tuple(list_encabezado))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA

    # Subconsulta para obtener los IDs de los dos últimos registros para cada cliente
    subquery = (
        HistorialEstadoCliente.objects
        .filter(
            cliente=OuterRef('cliente'),
            created_at__range=(fecha_inicio, fecha_fin),
            estado_cliente__lte=4,
        )
        .order_by('-created_at')
        .values('id')[:2]
    )

    # Obtener los registros utilizando la subconsulta
    historiales = (
        HistorialEstadoCliente.objects
        .filter(id__in=Subquery(subquery))
        .select_related('cliente')
        .order_by('cliente', 'created_at')
    )
    
    # Crear la lista de listas de clientes
    data = []
    current_cliente = None
    cliente_info = []

    # Contadores para clientes INTERESADOS o NUEVOS y clientes FINALES
    clientes_interesados_count = 0
    clientes_finales_count = 0

    for historial in historiales:
        if historial.cliente != current_cliente:
            # Nuevo cliente, agregar la lista actual a data y reiniciar
            if cliente_info:
                data.append(cliente_info)
            current_cliente = historial.cliente
            cliente_info = [historial.cliente.razon_social]

        # Agregar información del cliente
        cliente_info.append(historial.created_at.strftime('%d/%m/%Y'))
        cliente_info.append(historial.get_estado_cliente_display())

        # Contar clientes INTERESADOS y clientes FINALES
        if historial.estado_cliente == 3:
            clientes_interesados_count += 1

        if historial.estado_cliente == 4:
            clientes_finales_count += 1

    # Añadir el último cliente a data
    if cliente_info:
        data.append(cliente_info)
    
    for fila in data:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO

    ajustarColumnasSheet(hoja)

    list_encabezado_resumen = [
        'Clientes Interesados',
        'Clientes Finales',
        'Tasa de conversión',
        ]
    hoja = wb.create_sheet('RESUMEN')
    
    hoja.append(tuple(list_encabezado_resumen))

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = RELLENO_EXCEL
        cell_header.font = NEGRITA

    tasa_conversion = (clientes_finales_count/clientes_interesados_count)*100

    list_data_resumen = [
        clientes_interesados_count,
        clientes_finales_count,
        ("%s %s" % (intcomma(redondear((Decimal(clientes_finales_count/clientes_interesados_count))*100)), '%')),
        ]

    hoja.append(tuple(list_data_resumen))

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 1 or col == 2:
                row[col].alignment = ALINEACION_DERECHA

    ajustarColumnasSheet(hoja)
    return wb

def ReporteTasaConversionCliente(fecha_inicio, fecha_fin):

    wb=dataEstadosCliente(fecha_inicio, fecha_fin)
    return wb

####################################################  REPORTE CONTADOR  ####################################################   

def dataReporteContador(sociedad, fecha_inicio, fecha_fin):
    moneda_base = Moneda.objects.get(simbolo='$')

    list_encabezado = [
        'FECHA',
        'TIPO DE COMP.',
        'N° COMPROB.',
        'RAZON SOCIAL',
        'RUC',
        'PRODUCTOS',
        'CANT.',
        'PRECIO UNIT. (US$) CON IGV',
        'MONTO (US$)',
        'IGV (US$)',
        'DESCUENTO GLOBAL',
        'TOTAL (US$)',
        'TIPO DE CAMBIO',
        'MONTO SOLES (S/)',
        'OBSERVACIONES',
        'LINK',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'Reporte'
    hoja.append(tuple(list_encabezado))

    color_relleno = rellenoSociedadCorregido(sociedad)

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = color_relleno
        cell_header.font = NEGRITA

    info = []
    facturas = FacturaVenta.objects.filter(
        fecha_emision__gte=fecha_inicio,
        fecha_emision__lte=fecha_fin,
        sociedad=sociedad,
    )
    boletas = BoletaVenta.objects.filter(
        fecha_emision__gte=fecha_inicio,
        fecha_emision__lte=fecha_fin,
        sociedad=sociedad,
    )

    tipo_cambio_sunat = TipoCambioSunat.objects.filter(
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    )

    for factura in facturas:
        tipo_cambio_sunat_documento = tipo_cambio_sunat.get(fecha=factura.fecha_emision)
        if tipo_cambio_sunat_documento.tipo_cambio_venta:
            tipo_cambio_venta = tipo_cambio_sunat_documento.tipo_cambio_venta
        else:
            tipo_cambio_venta = Decimal('0.00')
        fila = []
        fila.append(factura.fecha_emision)  #0
        fila.append(factura.get_tipo_comprobante_display()) #1
        fila.append(f"{factura.serie_comprobante.serie}-{numeroXn(factura.numero_factura,6)}")    #2
        fila.append(factura.cliente.razon_social)   #3
        fila.append(factura.cliente.numero_documento)    #4
        contador = factura.contador
        fila.append(contador[0]) #5
        fila.append(contador[1])  #6
        fila.append(contador[2])  #7
        fila.append(factura.total_gravada)  #8
        fila.append(factura.total_igv)    #9
        fila.append(factura.descuento_global)   #10
        fila.append(factura.total)  #11
        fila.append(tipo_cambio_venta.quantize(Decimal('0.01')))    #12
        fila.append(tipo_cambio_venta * factura.total)    #13
        fila.append(factura.observaciones)  #14
        if factura.url_nubefact:
            fila.append(get_enlace_nubefact(factura.url_nubefact))   #15
        else:
            fila.append("")   #15
        fila.append(factura.estado) #16
        info.append(fila)

    for boleta in boletas:
        tipo_cambio_sunat_documento = tipo_cambio_sunat.get(fecha=boleta.fecha_emision)
        fila = []
        fila.append(boleta.fecha_emision)  #0
        fila.append(boleta.get_tipo_comprobante_display()) #1
        fila.append(f"{boleta.serie_comprobante.serie}-{numeroXn(boleta.numero_boleta,6)}")    #2
        fila.append(boleta.cliente.razon_social)   #3
        fila.append(boleta.cliente.numero_documento)    #4
        contador = boleta.contador
        fila.append(contador[0]) #5
        fila.append(contador[1])  #6
        fila.append(contador[2])  #7
        fila.append(boleta.total_gravada)  #8
        fila.append(boleta.total_igv)    #9
        fila.append(boleta.descuento_global)   #10
        fila.append(boleta.total)    #11
        fila.append(tipo_cambio_venta.quantize(Decimal('0.01')))    #12
        fila.append(tipo_cambio_venta * boleta.total)    #13
        fila.append(boleta.observaciones)  #14
        if boleta.url_nubefact:
            fila.append(get_enlace_nubefact(boleta.url_nubefact))   #15
        else:
            fila.append("")   #15
        fila.append(boleta.estado) #16
        info.append(fila)

    info.sort(key = lambda i:i[2])
    info.sort(key = lambda i:i[0])

    for fila in info:
        if fila[16] != 3:
            try:
                fila[8] = float(fila[8])
                fila[9] = float(fila[9])
                fila[10] =float(fila[10])
                fila[11] =float(fila[11])
                fila[12] = round(float(fila[12]),2)
                fila[13] = float(fila[13])
            except:
                pass
        else:
            for i in range(16):
                if i == 3:
                    fila[i] = 'ANULADO'
                elif i > 3:
                    fila[i] = ''
        fila.pop(-1)

    for fila in info:
        hoja.append(fila)

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if 8 <= col <=11:
                row[col].alignment = ALINEACION_DERECHA
                row[col].number_format = FORMATO_DOLAR
            elif col == 13:
                row[col].alignment = ALINEACION_DERECHA
                row[col].number_format = FORMATO_SOLES
            elif col == 15:
                if row[col].value != 'LINK':
                    row[col].hyperlink =  row[col].value
                    row[col].font =  COLOR_AZUL

    info2 = []
    notas = NotaCredito.objects.filter(
        fecha_emision__gte=fecha_inicio,
        fecha_emision__lte=fecha_fin,
        sociedad=sociedad,
    )

    for nota in notas:
        fila = []
        fila.append(nota.fecha_emision) #0
        fila.append(nota.get_tipo_comprobante_display())    #1
        fila.append(f"{nota.serie_comprobante.serie}-{numeroXn(nota.numero_nota,6)}")    #2
        fila.append(nota.cliente.razon_social)  #3
        fila.append(nota.cliente.numero_documento)   #4
        fila.append(nota.documento.documento)  #5
        fila.append(nota.observaciones)   #6
        contador = nota.contador
        fila.append(contador[1])    #7
        fila.append(contador[0])    #8
        fila.append(contador[2])    #9
        fila.append(nota.descuento_global)  #10
        fila.append(nota.total_gravada)   #11
        fila.append(nota.total_igv)   #12
        fila.append(nota.total)   #13
        fila.append(nota.get_tipo_nota_credito_display())  #14
        if nota.url_nubefact:
            fila.append(get_enlace_nubefact(nota.url_nubefact))   #15
        else:
            fila.append("")   #15
        info2.append(fila)  

    for fila in info2:
        fila[10] = float(fila[10])
        fila[11] = float(fila[11])
        fila[12] = float(fila[12])
        fila[13] = float(fila[13])

    if info2 != []:
        # cabecera de la tabla
        hoja.append(('',)) # Crea la fila del encabezado con los títulos
        hoja.append(('FECHA', 'TIPO DE COMP.', 'N° COMPROB.', 'RAZON SOCIAL', 'RUC', 'COMPROBANTE QUE SE MODIFICA', '', 'CANT.', 'DESCRIPCION', 'PRECIO UNIT. (US$) SIN IGV', 'DESCUENTO GLOBAL', 'VALOR DE VENTA (US$)', 'IGV (US$)', 'TOTAL (US$)', 'MOTIVO DE LA NOTA', 'LINK')) # Crea la fila del encabezado con los títulos
        nueva_fila = hoja.max_row

        for col in range(1, col_range + 1):
            cell_header = hoja.cell(nueva_fila, col)
            cell_header.fill = color_relleno
            cell_header.font = NEGRITA

        for fila in info2:
            hoja.append(fila) # Crea la fila del encabezado con los títulos

        for i in range(hoja.max_row):
            if i >= nueva_fila-1:
                row = list(hoja.rows)[i]
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if 10 <= col <=13:
                        row[col].alignment = ALINEACION_DERECHA
                        row[col].number_format = FORMATO_DOLAR
                    elif col == 15:
                        if row[col].value != 'LINK':
                            row[col].hyperlink =  row[col].value
                            row[col].font =  COLOR_AZUL

    ajustarColumnasSheet(hoja)
    return wb
    
def ReporteContadorCorregido(sociedad, fecha_inicio, fecha_fin):
    
    wb=dataReporteContador(sociedad, fecha_inicio, fecha_fin)
    return wb

####################################################  REPORTE ROTACIÓN  ####################################################   

def dataReporteRotacion2(sociedad=None):
    moneda_base = Moneda.objects.get(simbolo='$')

    list_encabezado = [
        'ID MAT.',
        'FAMILIA',
        'NOMBRE',
        'DESCRIPCIÓN',
        'PRECIO',
        'STOCK',
        'VENTA TOTAL',
        'VENTAS MENSUALES DEL TOTAL',
        'VENTA DESDE ÚLTIMOS 6 MESES (total/promedio)',
        'VENTA DESDE ULTIMO INGRESO (total/promedio)',
        'TIEMPO DURACION (APROX.)',
        'PEDIDO PARA 5 MESES',
        'SUGERENCIA',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'Reporte'
    hoja.append(tuple(list_encabezado))

    color_relleno = rellenoSociedadCorregido(sociedad)

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = color_relleno
        cell_header.font = NEGRITA

    lista_tipo_stock = [2,3,4,5,16,17,22,]
    # Recibido 2, Disponible 3, bloqueo sin serie 4, bloqueo sin qa 5, reservado 16, confirmado 17, prestado 22,
    # Recepcion compra 101, confirmacion por venta 120
    movimientos = MovimientosAlmacen.objects.filter(
        tipo_stock__codigo__in=lista_tipo_stock,
    )
    if sociedad:
        movimientos = movimientos.filter(
            sociedad=sociedad,
        )

    resumen = {}
    # id, familia, nombre, descripcion, precio, stock,
    # venta_total, fecha_inicial, fecha_final, venta_desde_6_meses, fecha_inicio_6_meses, venta_desde_ultimo_ingreso, fecha_ultimo_ingreso
    # tiempo_total = fecha_final - fecha_inicial
    # ventas_mensuales_del_total = 30 * venta_total / tiempo_total
    print("******************************************")
    print(len(movimientos), datetime.now())
    print("******************************************")
    for movimiento in movimientos:
        producto = movimiento.producto
        if producto == 'Error': continue
        if not producto in resumen:
            resumen[producto] = {
                'ID':producto.id,
                'FAMILIA':producto.subfamilia.familia.nombre,
                'NOMBRE':producto.descripcion_corta,
                'DESCRIPCION':producto.descripcion_venta,
                'PRECIO':producto.precio_lista,
                'FECHA INICIAL':None,
                'FECHA FINAL':None,
                'VENTA TOTAL':0,
                'FECHA INICIO 6 MESES':None,
                'VENTA 6 MESES':0,
                'FECHA ULTIMO INGRESO':None,
                'VENTA ULTIMO INGRESO':0,
                'INGRESOS':{},
                'VENTAS':{},
                'RECIBIDO':0,
                'DISPONIBLE':0,
                'BLOQUEO SIN SERIE':0,
                'BLOQUEO SIN QA':0,
                'RESERVADO':0,
                'CONFIRMADO':0,
                'PRESTADO':0,
            }
        if movimiento.tipo_movimiento.codigo == 101 and movimiento.tipo_stock.codigo == 2: #Recepción de compra : RECIBIDO
            if not resumen[producto]['FECHA ULTIMO INGRESO']:
                resumen[producto]['FECHA ULTIMO INGRESO'] = movimiento.fecha
            elif resumen[producto]['FECHA ULTIMO INGRESO'] < movimiento.fecha:
                resumen[producto]['FECHA ULTIMO INGRESO'] = movimiento.fecha
            if not movimiento.fecha in resumen[producto]['INGRESOS']: resumen[producto]['INGRESOS'][movimiento.fecha] = 0
            resumen[producto]['INGRESOS'][movimiento.fecha] = resumen[producto]['INGRESOS'][movimiento.fecha] + movimiento.cantidad
        elif movimiento.tipo_movimiento.codigo == 120 and movimiento.tipo_stock.codigo == 17: #Confirmado por venta : CONFIRMADO
            if not resumen[producto]['FECHA INICIAL']:
                resumen[producto]['FECHA INICIAL'] = movimiento.fecha
            elif resumen[producto]['FECHA INICIAL'] > movimiento.fecha:
                resumen[producto]['FECHA INICIAL'] = movimiento.fecha
            resumen[producto]['FECHA FINAL'] = movimiento.fecha
            if movimiento.fecha > date.today() - timedelta(180):
                resumen[producto]['VENTA 6 MESES'] = resumen[producto]['VENTA 6 MESES'] + movimiento.cantidad
                if not resumen[producto]['FECHA INICIO 6 MESES']:
                    resumen[producto]['FECHA INICIO 6 MESES'] = movimiento.fecha
                elif resumen[producto]['FECHA INICIO 6 MESES'] < movimiento.fecha:
                    resumen[producto]['FECHA INICIO 6 MESES'] = movimiento.fecha
            resumen[producto]['VENTA TOTAL'] = resumen[producto]['VENTA TOTAL'] + movimiento.cantidad
            if not movimiento.fecha in resumen[producto]['VENTAS']: resumen[producto]['VENTAS'][movimiento.fecha] = 0
            resumen[producto]['VENTAS'][movimiento.fecha] = resumen[producto]['VENTAS'][movimiento.fecha] + movimiento.cantidad

        if movimiento.tipo_stock.codigo == 2:
            resumen[producto]['RECIBIDO'] = resumen[producto]['RECIBIDO'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 3:
            resumen[producto]['DISPONIBLE'] = resumen[producto]['DISPONIBLE'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 4:
            resumen[producto]['BLOQUEO SIN SERIE'] = resumen[producto]['BLOQUEO SIN SERIE'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 5:
            resumen[producto]['BLOQUEO SIN QA'] = resumen[producto]['BLOQUEO SIN QA'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 16:
            resumen[producto]['RESERVADO'] = resumen[producto]['RESERVADO'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 17:
            resumen[producto]['CONFIRMADO'] = resumen[producto]['CONFIRMADO'] + movimiento.cantidad
        elif movimiento.tipo_stock.codigo == 22:
            resumen[producto]['PRESTADO'] = resumen[producto]['PRESTADO'] + movimiento.cantidad
            
    print("******************************************")
    print(len(resumen), datetime.now())
    print("******************************************")
    
    for producto, valores in resumen.items():
        fila = []
        precio_compra = Decimal('0.00')
        if valores['PRECIO']:
            precio_compra = valores['PRECIO'].precio_compra
        stock = valores['DISPONIBLE'] + valores['BLOQUEO SIN SERIE'] + valores['BLOQUEO SIN QA'] - valores['RESERVADO'] - valores['CONFIRMADO'] - valores['PRESTADO']
        if not valores['FECHA INICIAL'] and not valores['FECHA FINAL']:
            tiempo_total = None
        else:
            tiempo_total = (valores['FECHA FINAL'] - valores['FECHA INICIAL']).days
        if tiempo_total:
            ventas_mensuales = 30 * valores['VENTA TOTAL'] / tiempo_total
        else:
            ventas_mensuales = 30 * valores['VENTA TOTAL']

        if valores['FECHA INICIO 6 MESES']:
            tiempo_6_meses = (date.today() - valores['FECHA INICIO 6 MESES']).days
        else:
            tiempo_6_meses = 180
        if tiempo_6_meses:
            promedio_6_meses = 30 * valores['VENTA 6 MESES'] / tiempo_6_meses
        else:
            promedio_6_meses = 30 * valores['VENTA 6 MESES']

        for fecha, cantidad in valores['VENTAS'].items():
            if not valores['FECHA ULTIMO INGRESO']: valores['FECHA ULTIMO INGRESO'] = valores['FECHA INICIAL']
            if fecha >= valores['FECHA ULTIMO INGRESO']:
                valores['VENTA ULTIMO INGRESO'] = valores['VENTA ULTIMO INGRESO'] + cantidad
        
        tiempo_ultimo_ingreso = None
        if valores['FECHA ULTIMO INGRESO']:
            tiempo_ultimo_ingreso = (date.today() - valores['FECHA ULTIMO INGRESO']).days
        if tiempo_ultimo_ingreso:
            promedio_ultimo_ingreso = 30 * valores['VENTA ULTIMO INGRESO'] / tiempo_ultimo_ingreso
        else:
            promedio_ultimo_ingreso = 30 * valores['VENTA ULTIMO INGRESO']

        venta_promedio_mensual = Decimal(max(promedio_6_meses, promedio_ultimo_ingreso))

        if venta_promedio_mensual:
            tiempo_duracion = stock / venta_promedio_mensual
        else:
            tiempo_duracion = "-"

        pedido_5_meses = venta_promedio_mensual * 5

        if pedido_5_meses == 0:
            sugerencia = 'NO SE VENDIÓ'
        elif pedido_5_meses >= stock:
            sugerencia = 'EVALUAR'
        elif pedido_5_meses < stock:
            sugerencia = 'NO TRAER'

        fila.append(valores['ID'])
        fila.append(valores['FAMILIA'])
        fila.append(valores['NOMBRE'])
        fila.append(valores['DESCRIPCION'])
        fila.append(precio_compra)
        fila.append(stock)
        fila.append(valores['VENTA TOTAL'])
        fila.append(ventas_mensuales)
        fila.append(f"{valores['VENTA 6 MESES']} / {promedio_6_meses}")
        fila.append(f"{valores['VENTA ULTIMO INGRESO']} / {promedio_ultimo_ingreso}")
        fila.append(promedio_ultimo_ingreso)
        fila.append(tiempo_duracion)
        fila.append(pedido_5_meses)
        fila.append(sugerencia)
        
        hoja.append(fila)

    print("******************************************")
    print(datetime.now())
    print("******************************************")

    # for row in hoja.rows:
    #     for col in range(hoja.max_column):
    #         row[col].border = BORDE_DELGADO
    #         if 8 <= col <=11:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_DOLAR
    #         elif col == 13:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_SOLES
    #         elif col == 15:
    #             if row[col].value != 'LINK':
    #                 row[col].hyperlink =  row[col].value
    #                 row[col].font =  COLOR_AZUL

    ajustarColumnasSheet(hoja)
    return wb


def dataReporteRotacion(sociedad=None):
    moneda_base = Moneda.objects.get(simbolo='$')

    list_encabezado = [
        'ID MAT.',
        'FAMILIA',
        'NOMBRE',
        'DESCRIPCIÓN',
        'PRECIO',
        'STOCK',
        'VENTA TOTAL',
        'VENTAS MENSUALES DEL TOTAL',
        'VENTA DESDE ÚLTIMOS 6 MESES (total/promedio)',
        'VENTA DESDE ULTIMO INGRESO (total/promedio)',
        'TIEMPO DURACION (APROX.)',
        'PEDIDO PARA 5 MESES',
        'SUGERENCIA',
        ]

    wb = Workbook()
    hoja = wb.active
    hoja.title = 'Reporte'
    hoja.append(tuple(list_encabezado))

    color_relleno = rellenoSociedadCorregido(sociedad)

    col_range = hoja.max_column  # get max columns in the worksheet
    # cabecera de la tabla
    for col in range(1, col_range + 1):
        cell_header = hoja.cell(1, col)
        cell_header.fill = color_relleno
        cell_header.font = NEGRITA

    lista_tipo_stock = [2,3,4,5,15,16,17,22,]
    # Recibido 2, Disponible 3, bloqueo sin serie 4, bloqueo sin qa 5, despachado 15, reservado 16, confirmado 17, prestado 22,
    # Remover LABORATORIO
    # Recepcion compra 101, confirmacion por venta 120
    movimientos = MovimientosAlmacen.objects.filter(
        tipo_stock__codigo__in=lista_tipo_stock,
        content_type_producto=ContentType.objects.get_for_model(Material),
    )
    if sociedad:
        movimientos = movimientos.filter(
            sociedad=sociedad,
        )

    # Obtén el último precio de lista para cada material
    ultimos_precios = PrecioListaMaterial.objects.filter(
        content_type_producto=ContentType.objects.get_for_model(Material),
        id_registro_producto=OuterRef('id')
    ).order_by('-created_at').values('precio_lista')[:1]


    # Consulta para obtener todos los materiales con subfamilia y último precio de lista
    materiales = Material.objects.annotate(
        precio=Subquery(ultimos_precios)
    )
    
    # FAMILIA.append(producto.subfamilia.familia.nombre)
    # NOMBRE.append(producto.descripcion_corta)
    # DESCRIPCION.append(producto.descripcion_venta)
    # PRECIO.append(producto.precio_lista)

    movimientos_valores = movimientos.values_list(
        'id_registro_producto',
        'fecha_documento',
        'tipo_movimiento__codigo',
        'tipo_stock__codigo',
        'cantidad',
        'signo_factor_multiplicador',
        )

    materiales_valores = materiales.values_list(
        'id',
        'subfamilia__familia__nombre',
        'descripcion_corta',
        'descripcion_venta',
        'precio',
        )
    
    # Convierte el diccionario en un DataFrame de Pandas
    dfMateriales = pd.DataFrame(materiales_valores, columns=['ID', 'FAMILIA', 'NOMBRE', 'DESCRIPCION', 'PRECIO'])
    df = pd.DataFrame(movimientos_valores, columns=['ID', 'FECHA', 'TIPO_MOVIMIENTO', 'TIPO_STOCK', 'CANTIDAD', 'FACTOR'])

    # Ahora df contiene los datos de la consulta ORM en un DataFrame de Pandas
    df['FECHA'] = pd.to_datetime(df['FECHA'])
    
    # Suponiendo que tu DataFrame se llama df
    # Calcula la fecha actual
    fecha_actual = datetime.now()

    # Función personalizada para calcular STOCK
    def calcular_stock(group):
        filtro_stock = group['TIPO_STOCK'].isin([3, 4, 5]) & ~group['TIPO_STOCK'].isin([17, 22]) # Disponible, bloqueo sin serie, bloqueo sin qa - Confirmado por venta, prestado
        return (group.loc[filtro_stock, 'CANTIDAD'] * group.loc[filtro_stock, 'FACTOR']).sum()

    # Función personalizada para calcular FECHA_VENTA_TOTAL
    def calcular_fecha_venta_total(group):
        filtro_venta_total = ((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)) # Confirmado por venta y Confirmado - Devolución por Nota de Crédito y DESPACHADO
        return (fecha_actual - group.loc[filtro_venta_total, 'FECHA'].min()).days

    # Función personalizada para calcular FECHA_ULTIMO_INGRESO
    def calcular_fecha_ultimo_ingreso(group):
        filtro_ultimo_ingreso = (group['TIPO_MOVIMIENTO'] == 101) & (group['TIPO_STOCK'] == 2) # Recepcion compra y Recibido
        return (fecha_actual - group.loc[filtro_ultimo_ingreso, 'FECHA'].max()).days

    # Calcular la columna VENTA_ULTIMO_INGRESO para cada grupo
    def calcular_venta_ultimo_ingreso(group):
        ultimas_compras = group[(group['TIPO_MOVIMIENTO'] == 101) & (group['TIPO_STOCK'] == 2)]
        ultima_fecha_compra = ultimas_compras['FECHA'].max()
        group['VENTA_ULTIMO_INGRESO'] = group.apply(lambda row: row['CANTIDAD'] * row['FACTOR']
                                                    if row['TIPO_MOVIMIENTO'] == 120
                                                    and row['TIPO_STOCK'] == 17
                                                    and ultima_fecha_compra <= row['FECHA'] <= fecha_actual
                                                    else 0, axis=1)
        return group['VENTA_ULTIMO_INGRESO'].sum()


    # Aplica las funciones personalizadas y agrupa por ID
    print(
        df.groupby('ID').apply(lambda group: pd.Series({
            'VENTA_TOTAL': (group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)), 'CANTIDAD'] * group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)), 'FACTOR']),
        }))
    )
    time.sleep(10)
    resultados = df.groupby('ID').apply(lambda group: pd.Series({
        'STOCK': calcular_stock(group),
        'FECHA_VENTA_TOTAL': calcular_fecha_venta_total(group),
        'VENTA_TOTAL': (group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)), 'CANTIDAD'] * group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)), 'FACTOR']).sum(),
        'VENTA_ULTIMOS_6_MESES': (group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)) & (group['FECHA'] >= (fecha_actual - pd.Timedelta(days=180))), 'CANTIDAD'] * group.loc[((group['TIPO_MOVIMIENTO'] == 120) & (group['TIPO_STOCK'] == 17)) | ((group['TIPO_MOVIMIENTO'] == 159) & (group['TIPO_STOCK'] == 15)) & (group['FECHA'] >= (fecha_actual - pd.Timedelta(days=180))), 'FACTOR']).sum(),
        'FECHA_ULTIMO_INGRESO': calcular_fecha_ultimo_ingreso(group),
        'VENTA_ULTIMO_INGRESO': calcular_venta_ultimo_ingreso(group)
    }))
    resultados['STOCK'] = resultados['STOCK'].astype(float)
    resultados['FECHA_VENTA_TOTAL'] = resultados['FECHA_VENTA_TOTAL'].astype(float)
    resultados['VENTA_TOTAL'] = resultados['VENTA_TOTAL'].astype(float)
    resultados['VENTA_ULTIMOS_6_MESES'] = resultados['VENTA_ULTIMOS_6_MESES'].astype(float)
    resultados['FECHA_ULTIMO_INGRESO'] = resultados['FECHA_ULTIMO_INGRESO'].astype(float)

    # Calcula las demás columnas
    resultados['VENTA_MENSUALES_TOTAL'] = resultados['VENTA_TOTAL'] / resultados['FECHA_VENTA_TOTAL']
    resultados['VENTA_ULTIMOS_6_MESES_PROMEDIO'] = resultados['VENTA_ULTIMOS_6_MESES'] / 6
    resultados['VENTA_ULTIMO_INGRESO_PROMEDIO'] = (resultados['VENTA_TOTAL'] * 30) / resultados['FECHA_ULTIMO_INGRESO']
    resultados['TIEMPO_DURACION'] = resultados['STOCK'] / resultados[['VENTA_MENSUALES_TOTAL', 'VENTA_ULTIMOS_6_MESES_PROMEDIO', 'VENTA_ULTIMO_INGRESO_PROMEDIO']].max(axis=1)
    resultados['PEDIDO_5_MESES'] = 5 * resultados[['VENTA_MENSUALES_TOTAL', 'VENTA_ULTIMOS_6_MESES_PROMEDIO', 'VENTA_ULTIMO_INGRESO_PROMEDIO']].max(axis=1)
    resultados['SUGERENCIA'] = 'NO SE VENDIÓ'
    resultados.loc[resultados['PEDIDO_5_MESES'] >= resultados['STOCK'], 'SUGERENCIA'] = 'EVALUAR'
    resultados.loc[resultados['PEDIDO_5_MESES'] < resultados['STOCK'], 'SUGERENCIA'] = 'EVALUAR'

    # Ver los resultados
    resultados.reset_index(inplace=True)
    print(resultados)

    resultado_combinado = pd.merge(dfMateriales, resultados, on='ID', how='right')

    
    for r_idx, row in enumerate(dataframe_to_rows(resultado_combinado, index=False, header=True), 1):
        if r_idx == 1: continue
        hoja.cell(row=r_idx, column=1, value=row[0]) #ID
        hoja.cell(row=r_idx, column=2, value=row[1]) #FAMILIA
        hoja.cell(row=r_idx, column=3, value=row[2]) #NOMBRE
        hoja.cell(row=r_idx, column=4, value=row[3]) #DESCRIPCION
        hoja.cell(row=r_idx, column=5, value=row[4]) #PRECIO
        hoja.cell(row=r_idx, column=6, value=row[5]) #STOCK

        hoja.cell(row=r_idx, column=7, value=round(row[7], 2)) #VENTA_TOTAL
        hoja.cell(row=r_idx, column=8, value=round(row[11], 2)) #VENTA_MENSUALES_TOTAL
        hoja.cell(row=r_idx, column=9, value=f"{round(row[8], 2)} / {round(row[12], 2)}") #VENTA_ULTIMOS_6_MESES / VENTA_ULTIMOS_6_MESES_PROMEDIO
        hoja.cell(row=r_idx, column=10, value=f"{round(row[10], 2)} / {round(row[13], 2)}") #VENTA_ULTIMO_INGRESO / VENTA_ULTIMO_INGRESO_PROMEDIO
        hoja.cell(row=r_idx, column=11, value=round(row[14], 2)) #TIEMPO_DURACION
        hoja.cell(row=r_idx, column=12, value=round(row[15], 2)) #PEDIDO_5_MESES
        hoja.cell(row=r_idx, column=13, value=row[16]) #SUGERENCIA

        # hoja.cell(row=r_idx, column=7, value=row[6]) #FECHA_VENTA_TOTAL
        # hoja.cell(row=r_idx, column=8, value=row[7]) #VENTA_TOTAL
        # hoja.cell(row=r_idx, column=9, value=row[8]) #VENTA_ULTIMOS_6_MESES
        # hoja.cell(row=r_idx, column=10, value=row[9]) #FECHA_ULTIMO_INGRESO
        # hoja.cell(row=r_idx, column=11, value=row[10]) #VENTA_ULTIMO_INGRESO
        # hoja.cell(row=r_idx, column=12, value=row[11]) #VENTA_MENSUALES_TOTAL
        # hoja.cell(row=r_idx, column=13, value=row[12]) #VENTA_ULTIMOS_6_MESES_PROMEDIO
        # hoja.cell(row=r_idx, column=14, value=row[13]) #VENTA_ULTIMO_INGRESO_PROMEDIO
        # hoja.cell(row=r_idx, column=15, value=row[14]) #TIEMPO_DURACION
        # hoja.cell(row=r_idx, column=16, value=row[15]) #PEDIDO_5_MESES
        # hoja.cell(row=r_idx, column=17, value=row[16]) #SUGERENCIA
        
        
        # for c_idx, value in enumerate(row, 1):
        #     hoja.cell(row=r_idx, column=c_idx, value=value)

    # # Aplicar formato a las columnas específicas
    # venta_total_column = hoja['C']
    # venta_total_column.font = Font(bold=True)

    # for row in hoja.rows:
    #     for col in range(hoja.max_column):
    #         row[col].border = BORDE_DELGADO
    #         if 8 <= col <=11:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_DOLAR
    #         elif col == 13:
    #             row[col].alignment = ALINEACION_DERECHA
    #             row[col].number_format = FORMATO_SOLES
    #         elif col == 15:
    #             if row[col].value != 'LINK':
    #                 row[col].hyperlink =  row[col].value
    #                 row[col].font =  COLOR_AZUL

    ajustarColumnasSheet(hoja)
    return wb
    
def ReporteRotacionCorregido(sociedad):
    
    wb=dataReporteRotacion(sociedad)
    return wb

####################################################  REPORTE DEPÓSITOS CUENTAS BANCARIAS  ####################################################   

def dataReporteDepositoCuentasBancarias(sociedad, fecha_inicio, fecha_fin):
    moneda_dolar = Moneda.objects.get(simbolo='$')
    moneda_sol = Moneda.objects.get(simbolo='S/')

    wb = Workbook()
    hoja = wb.active
    color_relleno = rellenoSociedadCorregido(sociedad)

    cuentas = CuentaBancariaSociedad.objects.filter(
        sociedad = sociedad, 
        estado = 1, 
        efectivo = False
        )
    
    info_cuentas = []
    for cuenta in cuentas:
        data = []
        if cuenta.banco:
            data.append(cuenta.banco.razon_social)
            data.append(cuenta.sociedad.razon_social)
            data.append(cuenta.numero_cuenta)
            data.append(cuenta.numero_cuenta_interbancaria)
            data.append(cuenta.moneda.nombre)
            info_cuentas.append(data)

    tipo_cambio_sunat = TipoCambioSunat.objects.filter(
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
        )

    list_temp_hojas = []
    dict_totales_cuentas = {}
    count_cuenta = 0
    list_nro_cuentas = []
    for fila in info_cuentas:
        nro_cuenta = fila[2]
        list_nro_cuentas.append("%s %s" %(fila[0],fila[4]))
        info_depositos = []
        # numero_operacion = []

        ingresos = Ingreso.objects.filter(
            cuenta_bancaria__numero_cuenta = nro_cuenta, 
            fecha__gte=fecha_inicio, 
            fecha__lte=fecha_fin,).order_by('fecha')

        for ingreso in ingresos:
            if ingreso.pagos:
                data = []
                tipo_cambio_sunat_documento = tipo_cambio_sunat.get(fecha=ingreso.fecha)
                data.append(ingreso.fecha.strftime('%d/%m/%Y'))
                data.append(ingreso.numero_operacion)
                if ingreso.cuenta_bancaria.moneda.abreviatura == "USD":
                    data.append(ingreso.monto)
                    data.append(ingreso.monto*tipo_cambio_sunat_documento.tipo_cambio_venta)
                else:
                    data.append(ingreso.monto/tipo_cambio_sunat_documento.tipo_cambio_venta)
                    data.append(ingreso.monto)
                for pago in ingreso.pagos:
                    if pago == ingreso.pagos[0]:
                        data.append(pago.deuda.documento_objeto[3].razon_social)
                        data.append("%s: %s" % (pago.deuda.documento_objeto[0], pago.deuda.documento_objeto[1]))
                        data.append(pago.deuda.documento_objeto[2].strftime('%d/%m/%Y'))
                        if ingreso.cuenta_bancaria.moneda.abreviatura == "USD":
                            data.append("%s %s" %(moneda_dolar.simbolo, intcomma(redondear(pago.monto))))
                            data.append("%s %s" %(moneda_sol.simbolo, intcomma(redondear(pago.monto*tipo_cambio_sunat_documento.tipo_cambio_venta))))
                        else:
                            data.append("%s %s" %(moneda_dolar.simbolo, intcomma(redondear(pago.monto/tipo_cambio_sunat_documento.tipo_cambio_venta))))
                            data.append("%s %s" %(moneda_sol.simbolo, intcomma(redondear(pago.monto))))
                    else:
                        data[4] = "%s\n%s" %(data[4], pago.deuda.documento_objeto[3].razon_social)
                        data[5] = "%s\n%s" %(data[5], "%s: %s" % (pago.deuda.documento_objeto[0], pago.deuda.documento_objeto[1]))
                        data[6] = "%s\n%s" %(data[6], pago.deuda.documento_objeto[2].strftime('%d/%m/%Y'))
                        if ingreso.cuenta_bancaria.moneda.abreviatura == "USD":
                            data[7] = "%s\n%s" %(data[7], "%s %s" %(moneda_dolar.simbolo, intcomma(redondear(pago.monto))))
                            data[8] = "%s\n%s" %(data[8], "%s %s" %(moneda_sol.simbolo, intcomma(redondear(pago.monto*tipo_cambio_sunat_documento.tipo_cambio_venta))))
                        else:
                            data[7] = "%s\n%s" %(data[7], "%s %s" %(moneda_dolar.simbolo, intcomma(redondear(pago.monto/tipo_cambio_sunat_documento.tipo_cambio_venta))))
                            data[8] = "%s\n%s" %(data[8], "%s %s" %(moneda_sol.simbolo, intcomma(redondear(pago.monto)))) 
                info_depositos.append(data)

            else:
                data = []
                tipo_cambio_sunat_documento = tipo_cambio_sunat.get(fecha=ingreso.fecha)
                data.append(ingreso.fecha.strftime('%d/%m/%Y'))
                if ingreso.comentario:
                    data.append("%s | %s" % (ingreso.numero_operacion, ingreso.comentario))
                else:
                    data.append("%s | %s" % (ingreso.numero_operacion, 'Sin Comentario'))
                if ingreso.cuenta_bancaria.moneda.abreviatura == "USD":
                    data.append(ingreso.monto)
                    data.append(ingreso.monto*tipo_cambio_sunat_documento.tipo_cambio_venta)
                else:
                    data.append(ingreso.monto/tipo_cambio_sunat_documento.tipo_cambio_venta)
                    data.append(ingreso.monto)
                
                data.append('')
                data.append('')
                data.append('')
                data.append('')
                data.append('')
                info_depositos.append(data)

        count_cuenta += 1
        # print('**********************************')
        # print(info_depositos)
        # print('**********************************')
        count_mes = 0
        list_general = []
        list_mes = []
        for deposito in info_depositos:
            if count_mes != 0:
                año_mes_actual = deposito[0][6:] + deposito[0][3:5]
                año_mes_anterior = info_depositos[count_mes-1][0][6:] + info_depositos[count_mes-1][0][3:5]
                if año_mes_actual != año_mes_anterior:
                    list_general.append(list_mes)
                    list_mes = []
            list_mes.append(deposito)
            try:
                deposito[2] = float(deposito[2])
                deposito[3] = float(deposito[3])
            except:
                ""
            count_mes += 1
        if list_mes != []:
            list_general.append(list_mes)
        # print('**********************************')
        # print(list_general)
        # print('**********************************')

        count = 0
        for list_mes_deposito in list_general:
            mes = list_mes_deposito[0][0][3:5]
            año = list_mes_deposito[0][0][6:]
            name_sheet = DICT_MESES[str(mes)] + ' - ' + str(año)
            # print('*************************************')
            # print(str(mes), name_sheet, list_temp_hojas)
            # print('*************************************')
            if count != 0:
                if name_sheet not in list_temp_hojas:
                    hoja = wb.create_sheet(name_sheet)
                    list_temp_hojas.append(name_sheet)
                else:
                    hoja = wb[name_sheet]
            else:
                hoja = wb.active
                hoja.title = name_sheet
                if name_sheet not in list_temp_hojas:
                    list_temp_hojas.append(name_sheet)
                count += 1
                # count_cuenta += 1

            hoja.append(('', ''))
            hoja.append(('', ''))
            hoja.append(('BANCO:', fila[0]))
            hoja.append(('EMPRESA:', fila[1]))
            hoja.append(('CUENTA:', fila[2]))
            hoja.append(('CCI:', fila[3]))
            hoja.append(('MONEDA:', fila[4]))
            # wb.active = hoja
            hoja.append(('FECHA', 'REFERENCIA', 'MONTO (US$)', 'MONTO (S/)', 'EMPRESAS', 'DOCUMENTOS', 'FECHA DOCUMENTOS', 'PAGOS (US$)', 'PAGOS (S/)'))
            col_range = hoja.max_column
            nueva_fila = hoja.max_row

            for col in range(1, col_range + 1):
                cell_header = hoja.cell(nueva_fila, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA
                for count_fila in range(1,6):
                    cell_header = hoja.cell(nueva_fila-count_fila, col)
                    cell_header.fill = color_relleno

            total_mes_cuenta_dolares = 0
            total_mes_cuenta_soles = 0
            for fila_deposito in list_mes_deposito:
                if fila_deposito[4] != "" and fila_deposito[4] != None:
                    if 'Nota Credito:' not in fila_deposito[1]:
                        total_mes_cuenta_dolares += float(fila_deposito[2])
                        try:
                            total_mes_cuenta_soles += float(fila_deposito[3])
                        except:
                            pass
                hoja.append(fila_deposito)
            cuenta_banco_ingreso = fila[0] + ' ' + fila[4]
            dict_totales_cuentas[cuenta_banco_ingreso + '|' + name_sheet] = str(round(total_mes_cuenta_soles,2)) + '|' + str(round(total_mes_cuenta_dolares,2))

            for i in range(hoja.max_row):
                if i >= nueva_fila-1:
                    row = list(hoja.rows)[i]
                    for col in range(hoja.max_column):
                        row[col].border = BORDE_DELGADO
                        if col == 2:
                            row[col].alignment = ALINEACION_DERECHA
                            row[col].number_format = FORMATO_DOLAR
                        if col == 3:
                            row[col].alignment = ALINEACION_DERECHA
                            row[col].number_format = FORMATO_SOLES
                        if 4 <= col <= 6:
                            row[col].alignment = AJUSTAR_TEXTO
                        if col == 7 or col == 8:
                            row[col].alignment = AJUSTAR_TEXTO_DERECHA

            ajustarColumnasSheet(hoja)

    hoja = wb.create_sheet('Resumen Ingresos')
    # hoja.append(('', ''))
    hoja.append(('ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
    hoja.append(('', ''))
    num_cuentas = len(list_nro_cuentas)
    hoja.cell(row= 3, column=1).value = "SOLES"
    hoja.cell(row= 3, column= num_cuentas+4).value = "DÓLARES"
    hoja.merge_cells(start_row = 3, start_column = 1, end_row = 3, end_column = num_cuentas+2)
    hoja.merge_cells(start_row = 3, start_column = num_cuentas+4, end_row = 3, end_column = 2*num_cuentas+5)
    celda_multiple = hoja.cell(row = 3, column = 1)
    celda_multiple.alignment = ALINEACION_CENTRO
    celda_multiple = hoja.cell(row = 3, column = num_cuentas+4)
    celda_multiple.alignment = ALINEACION_CENTRO
    list_encabezado = [''] + list_nro_cuentas + ['TOTAL','',''] + list_nro_cuentas + ['TOTAL']
    hoja.append(tuple(list_encabezado))

    col_range = 2*num_cuentas+5
    nueva_fila = hoja.max_row
    for col in range(1, col_range+1):
        if col != num_cuentas+3:
            cell_header = hoja.cell(nueva_fila, col)
            cell_header.fill = color_relleno
            cell_header.font = NEGRITA
            for count_fila in range(1,2):
                cell_header = hoja.cell(nueva_fila-count_fila, col)
                cell_header.fill = color_relleno

    def insertarResumenDataAnterior(hoja, data):
        for fila in data:
            hoja.append(fila)
        # return hoja
        
    if str(fecha_inicio) <= '2022-01-01':
        if sociedad.id == 2:
            insertarResumenDataAnterior(hoja, list_resumen_ingresos_sis_anterior_mpl)
        if sociedad.id == 1:
            insertarResumenDataAnterior(hoja, list_resumen_ingresos_sis_anterior_mca)

    for mes in list_temp_hojas:
        list_temp_fila = []
        # list_temp.append(mes)
        list_temp_soles = []
        list_temp_dolares = []
        total_soles = 0
        total_dolares = 0
        for k,value in dict_totales_cuentas.items():
            if mes in k:
                # print(k)
                monto_soles = float(value[:value.find("|")])
                monto_dolares = float(value[value.find("|")+1:])
                total_soles += monto_soles
                total_dolares += monto_dolares
                list_temp_soles.append(monto_soles)
                list_temp_dolares.append(monto_dolares)
                
        if len(list_temp_soles) < num_cuentas:
            while len(list_temp_soles) == num_cuentas:
                list_temp_soles.insert(len(list_temp_soles), " ")
                list_temp_dolares.insert(len(list_temp_dolares), " ")

        list_temp_soles.append(total_soles)
        list_temp_dolares.append(total_dolares)
        list_temp_fila = [mes] + list_temp_soles + ['', mes] + list_temp_dolares
        hoja.append(tuple(list_temp_fila))

    i = 0
    for row in hoja.rows:
        if i >= 2:
            for col in range(col_range):
                if col != num_cuentas+2:
                    row[col].border = BORDE_DELGADO
                if 1 <= col <= num_cuentas + 1:
                    row[col].number_format = FORMATO_SOLES
                if col >= num_cuentas + 4:
                    row[col].number_format = FORMATO_DOLAR
        i += 1
    ajustarColumnasSheet(hoja)

    def extraer_resumen_bloc_de_notas(hoja):

        # nro_col = col_range
        file_mpl = open(f'resumen_ingresos_{sociedad.abreviatura}_{fecha_inicio}_{fecha_fin}.txt', "w")
        file = file_mpl

        list_temp = []
        for i in range(5, hoja.max_row + 1): # Nro de filas del excel
            celda = hoja.cell(row = i, column = col_range)
            if celda.value == None:
                valor_celda = ''
            else:
                valor_celda = celda.value
            list_temp.append(valor_celda)
        # print(list_temp)

        for pos in range(len(list_temp)):
            list_temp[pos] = numero_str(list_temp[pos],2)

        info = '\n'.join(list_temp)
        file.write(info)
        file.close()

    extraer_resumen_bloc_de_notas(hoja)

    def grafico_resumen_ingresos(ws):
        max_fila = ws.max_row

        chart = LineChart()
        # print(help(chart))
        chart.height = 15 # default is 7.5
        chart.width = 30 # default is 15
        chart.y_axis.title = 'INGRESOS'
        chart.x_axis.title = 'MESES'
        chart.legend.position = 'b' #bottom
        # chart.style = 12

        count = 1
        fila_base = 5
        year_base = 2018
        chart.title = f'RESUMEN DE INGRESOS - {sociedad.abreviatura}'
        data_col = col_range # montos en dolares

        while max_fila >= 12:
            values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
            series = Series(values, title = "Ingresos del " + str(year_base + count))
            chart.append(series)
            max_fila -= 12
            count += 1
        if 1 <= max_fila <= 12:
            values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
            series = Series(values, title = "Ingresos del " + str(year_base + count))
            chart.append(series)

        meses = Reference(ws, min_col = 1, min_row = 1, max_col = 12, max_row = 1)
        chart.set_categories(meses)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.dLblPos = 't' # top

        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        ws.add_chart(chart)

    grafico_resumen_ingresos(hoja)

    return wb

def dataReporteDepositoCuentasBancariasResumen(sociedades,fecha_inicio, fecha_fin):
    nombres_soc = []
    listas = []
    for sociedad in sociedades:
        try:
            file = open(f'resumen_ingresos_{sociedad.abreviatura}_{fecha_inicio}_{fecha_fin}.txt', "r")
            contenido = file.read()
            file.close()
            lista = contenido.split("\n")
            for pos in range(len(lista)):
                if lista[pos] != "":
                    lista[pos]=float(lista[pos])
                else:
                    lista[pos]=float('0.00')
            listas.append(lista)
            nombres_soc.append(sociedad.nombre_comercial)
        except Exception as e:
            print(e)
            

    max_lon = max(list(map(len, listas)))
    rango_listas = range(len(listas))
    rango_max = range(max_lon)
    
    listas_todas_empresas = []

    for b in rango_max:
        total_lista = float('0.00')
        for a in rango_listas:
            if b in range(len(listas[a])):
                total_lista+=listas[a][b]
            else:
                total_lista+=float('0.00')

        listas_todas_empresas.append(total_lista)

    wb = Workbook()
    hoja = wb.active
    hoja.append(('','ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
    # for dato in list_ambas_empresas:
    for dato in listas_todas_empresas:
        hoja.append((dato,''))

    for row in hoja.rows:
        for col in range(hoja.max_column):
            row[col].border = BORDE_DELGADO
            if col == 0:
                row[col].number_format = FORMATO_DOLAR

    def grafico_resumen_ingresos(ws):
        max_fila = ws.max_row

        chart = LineChart()
        # print(help(chart))
        chart.height = 15 # default is 7.5
        chart.width = 30 # default is 15
        chart.y_axis.title = 'INGRESOS'
        chart.x_axis.title = 'MESES'
        chart.legend.position = 'b' #bottom
        chart.title = f'RESUMEN DE INGRESOS - {" + ".join(nombres_soc)}'
        # chart.style = 12

        count = 1
        fila_base = 2
        year_base = 2018
        data_col = 1 # montos en dolares
        while max_fila >= 12:
            values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
            series = Series(values, title = "Ingresos del " + str(year_base + count))
            chart.append(series)
            max_fila -= 12
            count += 1
        if 1 <= max_fila <= 12:
            values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
            series = Series(values, title = "Ingresos del " + str(year_base + count))
            chart.append(series)

        meses = Reference(ws, min_col = 2, min_row = 1, max_col = 13, max_row = 1)
        chart.set_categories(meses)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.dLblPos = 't' # top
        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        ws.add_chart(chart)

    grafico_resumen_ingresos(hoja)

    return wb
    
def ReporteDepositoCuentasBancariasCorregido(sociedad, fecha_inicio, fecha_fin):

    if sociedad!=None:
        wb=dataReporteDepositoCuentasBancarias(sociedad, fecha_inicio, fecha_fin)
        return wb
    else:
        sociedades = Sociedad.objects.only('id','abreviatura','nombre_comercial')
        wb=dataReporteDepositoCuentasBancariasResumen(sociedades, fecha_inicio, fecha_fin)
        return wb

####################################################  REPORTE RESUMEN STOCK PRODUCTOS  ####################################################   

def ReporteResumenStockProductosCorregido(sede):

    if sede:
        sedes = Sede.objects.filter(id=sede.id)
    else:
        sedes = Sede.objects.all()

    material = Material.objects.only('id','descripcion_corta').order_by('descripcion_corta')

    material_valores = material.values_list(
    'id',
    'descripcion_corta',
    )
    
    df_material = pd.DataFrame(material_valores, columns=['ID','DESCRIPCION CORTA'])

    df_material = df_material.assign(ALMACEN_1=None, ALMACEN_2=None, ALMACEN_3=None, ALMACEN_4=None, ALMACEN_5=None, SUMA_CONTEO=None, DIFERENCIA=None)

    wb = Workbook()
    count = 0
    for sede in sedes:

        list_encabezado = [
            'SEDE',
            'DESCRIPCIÓN',
            'ALMACÉN #1',
            'ALMACÉN #2',
            'ALMACÉN #3',
            'ALMACÉN #4',
            'ALMACÉN #5',
            'SUMA CONTEO',
            'DISPONIBLE',
            'BLOQ. SIN QA',
            'BLOQ. POR QA',
            'BLOQ. DESG.',
            'SUMATORIA',
            'DIFERENCIA',
            ]

        if count != 0:
            hoja = wb.create_sheet(str(sede.nombre))
        else:
            hoja = wb.active
            hoja.title = str(sede.nombre)

        hoja.append(tuple(list_encabezado))

        col_range = hoja.max_column  # get max columns in the worksheet
        # cabecera de la tabla
        for col in range(1, col_range + 1):
            if col == 8 or col == 13:
                # color_celda_cabecera = PatternFill(start_color='8C4966', end_color='8C4966', fill_type='solid')
                color_celda_cabecera = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            elif col == 14:
                color_celda_cabecera = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            else:
                color_celda_cabecera = RELLENO_EXCEL
            cell_header = hoja.cell(1, col)
            cell_header.fill = color_celda_cabecera
            cell_header.font = NEGRITA

        data = MovimientosAlmacen.objects.filter(
                content_type_producto=ContentType.objects.get_for_model(Material), 
                id_registro_producto__in = Material.objects.values_list('id', flat=True),
                almacen__sede__id = sede.id,
            ).values(
                'id_registro_producto', 'almacen__sede__nombre'
            ).exclude(
                tipo_stock__codigo__in = [1, 2, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 32, 33, 34, 35, 36, 37, 38],
            ).annotate(
                stock_disponible=Sum(Case(When(tipo_stock__codigo='3', then=F('cantidad') * F('signo_factor_multiplicador')), default=0.0, output_field=models.DecimalField())),
                stock_sin_qa=Sum(Case(When(tipo_stock__codigo__in=['4', '5'], then=F('cantidad') * F('signo_factor_multiplicador')), default=0.0, output_field=models.DecimalField())),
                stock_por_qa=Sum(Case(When(tipo_stock__codigo__in=['6', '26'], then=F('cantidad') * F('signo_factor_multiplicador')), default=0.0, output_field=models.DecimalField())),
                stock_otros=Sum(Case(When(tipo_stock__codigo__in=['3', '4', '5', '6', '26'], then=0.0), default=F('cantidad') * F('signo_factor_multiplicador'), output_field=models.DecimalField())),
                total_stock=Sum(F('cantidad') * F('signo_factor_multiplicador'), output_field=models.DecimalField())
            ).order_by('id_registro_producto', 'almacen__sede__id')

        info = [[item[key] for key in item] for item in data]

        # Ejecutar la consulta
        df_info = pd.DataFrame(info, columns=['ID', 'SEDE', 'STOCK DISPONIBLE', 'STOCK SIN QA', 'STOCK POR QA', 'STOCK OTROS', 'TOTAL STOCK'])

        # Combinar las consultas
        df_combinado = pd.merge(df_info, df_material, on='ID', how='inner')

        # Reordenar las columnas
        columnas_ordenadas = ['SEDE', 'DESCRIPCION CORTA', 'ALMACEN_1', 'ALMACEN_2', 'ALMACEN_3', 'ALMACEN_4', 'ALMACEN_5', 'SUMA_CONTEO', 'STOCK DISPONIBLE', 'STOCK SIN QA', 'STOCK POR QA', 'STOCK OTROS', 'TOTAL STOCK', 'DIFERENCIA']

        df_combinado_final = df_combinado[columnas_ordenadas].sort_values(by='DESCRIPCION CORTA')

        # info = df_combinado_final.values.tolist()

        for r_idx, row in enumerate(dataframe_to_rows(df_combinado_final, index=False, header=True), 1):
            if r_idx == 1: continue
            for c_idx, value in enumerate(row, 1):
                hoja.cell(row=r_idx, column=c_idx, value=value)
                if c_idx >=8:
                    hoja.cell(row=r_idx, column=c_idx, value=value)

        for row in hoja.rows:
            for col in range(hoja.max_column):
                row[col].border = BORDE_DELGADO
                if col >= 8:
                    row[col].number_format = FORMATO_NUMERO

        hoja.freeze_panes = 'C2'
        ajustarColumnasSheet(hoja)
        count += 1
    return wb

####################################################  REPORTE VENTAS FACTURADAS  ####################################################

def ReporteVentasFacturadasCorregido(sociedad, fecha_inicio, fecha_fin):

    query_notas = NotaCredito.objects.filter(
        sociedad_id=sociedad.id,
        fecha_emision__gte=fecha_inicio,
        fecha_emision__lte=fecha_fin
    ).annotate(
        fecha_emision_nota=Max('fecha_emision', format='%d/%m/%Y'),
        comprobante=Case(
            When(tipo_comprobante='3', then=Value('NOTA DE CRÉDITO')),
            default=Value('-'),
            output_field=models.CharField(),
        ),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_nota'), output_field=models.CharField()), 6, Value('0'))
        ),
        cliente_denominacion=Max('cliente__razon_social'),
        facturado=Sum(F('NotaCreditoDetalle_nota_credito__total') * Value(-1)),
        amortizado=Sum(F('NotaCreditoDetalle_nota_credito__total') * Value(-1)),
        pendiente=Value('0.00'),
        estado_nota=Value('PENDIENTE'),
        fecha_vencimiento_nota=Max('fecha_emision', format='%d/%m/%Y'),
        credito=Value('0.00'),
        dias=Value(''),
        guia=Value(''),
        obs=Value(''),
        letras=Value(''),
        pagos=Value('')
    ).values(
        'fecha_emision_nota', 
        'comprobante', 
        'nro_comprobante', 
        'cliente_denominacion',
        'facturado', 
        'amortizado', 
        'pendiente', 
        'estado_nota', 
        'fecha_vencimiento_nota', 
        'credito',
        'dias', 
        'guia', 
        'obs', 
        'letras', 
        'pagos'
    )

    # Ejecutar la consulta ORM y convertir el resultado en un DataFrame de Pandas
    data = list(query_notas)

    # Procesar los datos en Python
    for item in data:
        item['fecha_emision_nota'] = item['fecha_emision_nota'].strftime('%d/%m/%Y')
        item['fecha_vencimiento_nota'] = item['fecha_vencimiento_nota'].strftime('%d/%m/%Y')
        item['facturado'] = float(item['facturado'])  # Convertir 'facturado' a float
        item['amortizado'] = float(item['amortizado'])  # Convertir 'amortizado' a float
        item['credito'] = float(item['credito'])  # Convertir 'credito' a float

    # Crear el DataFrame de Pandas
    info_notas = pd.DataFrame(list(data))

##################################################################################################################################

    factura_anuladas = FacturaVenta.objects.filter(
        estado=3,
        sociedad_id=sociedad.id
    ).annotate(
        nro_comprobante=Case(
            When(serie_comprobante__tipo_comprobante_id=ContentType.objects.get_for_model(FacturaVenta),
                then=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_factura'), output_field=models.CharField()), 6, Value('0'))))
        ),
        id_comprobante=Max('id')
    ).values('id_comprobante', 'nro_comprobante')

    # Consulta ORM para boletas anuladas
    boleta_anuladas = BoletaVenta.objects.filter(
        estado=3,
        sociedad_id=sociedad.id
    ).annotate(
        nro_comprobante=Case(
            When(serie_comprobante__tipo_comprobante_id=ContentType.objects.get_for_model(BoletaVenta),
                then=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_boleta'), output_field=models.CharField()), 6, Value('0'))))
        ),
        id_comprobante=Max('id')
    ).values('id_comprobante', 'nro_comprobante')

    # Combinar los resultados de las consultas ORM en un DataFrame de Pandas
    info_anulados = pd.DataFrame.from_records(
        list(factura_anuladas) + list(boleta_anuladas),
        columns=['id_comprobante', 'nro_comprobante']
    )

##################################################################################################################################

    # Consulta para Facturas
    factura_queryset = FacturaVenta.objects.filter(
        sociedad_id=sociedad.id,
        estado=4
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('FACTURA'),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_factura'), output_field=models.CharField()), 6, Value('0'))),
        cliente_denominacion=F('cliente__razon_social')
    ).values(
        'id', 
        'fecha_emision_comprobante', 
        'comprobante', 'nro_comprobante', 
        'cliente_denominacion'
    ).order_by('cliente_denominacion')

    # Consulta para Boletas
    boleta_queryset = BoletaVenta.objects.filter(
        sociedad_id=sociedad.id,
        estado=4
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('BOLETA'),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_boleta'), output_field=models.CharField()), 6, Value('0'))),
        cliente_denominacion=F('cliente__razon_social')
    ).values(
        'id', 
        'fecha_emision_comprobante', 
        'comprobante', 
        'nro_comprobante', 
        'cliente_denominacion'
    ).order_by('cliente_denominacion')

    # Unir las consultas de Facturas y Boletas
    queryset = factura_queryset.union(boleta_queryset, all=True).order_by('cliente_denominacion')

    # Obtener los pagos asociados a cada comprobante
    for row in queryset:
        deudas = Deuda.objects.filter(
            content_type_id__in=[ContentType.objects.get_for_model(FacturaVenta), ContentType.objects.get_for_model(BoletaVenta)],
            id_registro=row['id']
        )
        ingresos = Ingreso.objects.values(
            'id', 'fecha', 'cuenta_bancaria__moneda__abreviatura', 'cuenta_bancaria__moneda__simbolo', 
        )

        pagos = Pago.objects.filter(
            deuda__in=deudas,
            content_type_id=ContentType.objects.get_for_model(Ingreso),
            id_registro__in = Ingreso.objects.values_list('id', flat=True),
        ).annotate(
            pago_monto=F('monto'),
            pago_tipo_cambio=F('tipo_cambio'),
            pago_fecha=Subquery(ingresos.filter(id=OuterRef('id_registro')).values('fecha')[:1]),
            moneda_simbolo=Subquery(ingresos.filter(id=OuterRef('id_registro')).values('cuenta_bancaria__moneda__simbolo')[:1]),
            moneda_abreviatura=Subquery(ingresos.filter(id=OuterRef('id_registro')).values('cuenta_bancaria__moneda__abreviatura')[:1]),
        ).values('id_registro', 'pago_monto', 'pago_tipo_cambio', 'pago_fecha', 'moneda_simbolo', 'moneda_abreviatura')

        # Crear la cadena de pagos
        pagos_str = '\n'.join([
            f"{p['pago_fecha'].strftime('%d/%m/%Y')} {p['moneda_simbolo']} {p['pago_monto']:.2f}"
            + (' ($ {:.2f})'.format(p['pago_monto'] / p['pago_tipo_cambio']) if p['moneda_abreviatura'] == 'PEN' else '')
            for p in pagos
            ])
        row['pagos'] = pagos_str

    # Convertir a DataFrame de Pandas
    df_pagos = pd.DataFrame(list(queryset))

##################################################################################################################################

    # Consulta para FacturaVenta
    facturas = FacturaVenta.objects.filter(
        tipo_venta='2',
        sociedad_id=sociedad.id,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        estado='4'
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('FACTURA', output_field=CharField()),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), 
            Value('-'), 
            LPad(Cast(Max('numero_factura'), output_field=models.CharField()), 6, Value('0'))),
    ).values(
        'id',
        'fecha_emision_comprobante',
        'comprobante',
        'nro_comprobante'
    )

    # Consulta para BoletaVenta
    boletas = BoletaVenta.objects.filter(
        tipo_venta='2',
        sociedad_id=sociedad.id,
        fecha_emision__range=[fecha_inicio, fecha_fin],
        estado='4'
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('BOLETA', output_field=CharField()),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), 
            Value('-'), 
            LPad(Cast(Max('numero_boleta'), output_field=models.CharField()), 6, Value('0'))),
    ).values(
        'id',
        'fecha_emision_comprobante',
        'comprobante',
        'nro_comprobante'
    )

   # Unir las consultas de Facturas y Boletas
    queryset = facturas.union(boletas, all=True).order_by('fecha_emision_comprobante', 'nro_comprobante')
    for row in queryset:
        cuotas = Cuota.objects.filter(
            deuda__content_type_id__in=[ContentType.objects.get_for_model(FacturaVenta), ContentType.objects.get_for_model(BoletaVenta)],
            deuda__id_registro=row['id']
        ).values('fecha', 'monto')

        letras_str = '\n'.join([
            f"{cuota['fecha'].strftime('%d/%m/%Y')} $ {cuota['monto']:.2f}" 
            for cuota in cuotas
            ])
        row['letras'] = letras_str
    
    df_letras = pd.DataFrame(list(queryset))

    # Formatear la fecha de emisión
    df_letras['fecha_emision_comprobante'] = pd.to_datetime(df_letras['fecha_emision_comprobante'])
    df_letras['fecha_emision_comprobante'] = df_letras['fecha_emision_comprobante'].dt.strftime('%d/%m/%Y')


##################################################################################################################################

    # Consulta para Facturas
    factura_queryset = FacturaVenta.objects.filter(
        sociedad_id=sociedad.id,
        estado=4
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('FACTURA'),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_factura'), output_field=models.CharField()), 6, Value('0'))),
        cliente_denominacion=F('cliente__razon_social')
    ).values(
        'id', 
        'comprobante', 
        'nro_comprobante', 
        'cliente_denominacion'
    ).order_by('cliente_denominacion')

    # Consulta para Boletas
    boleta_queryset = BoletaVenta.objects.filter(
        sociedad_id=sociedad.id,
        estado=4
    ).annotate(
        fecha_emision_comprobante=F('fecha_emision'),
        comprobante=Value('BOLETA'),
        nro_comprobante=Concat(
            Max('serie_comprobante__serie'), Value('-'), LPad(Cast(Max('numero_boleta'), output_field=models.CharField()), 6, Value('0'))),
        cliente_denominacion=F('cliente__razon_social')
    ).values(
        'id', 
        'comprobante', 
        'nro_comprobante', 
        'cliente_denominacion'
    ).order_by('cliente_denominacion')

    # Unir las consultas de Facturas y Boletas
    queryset = factura_queryset.union(boleta_queryset, all=True).order_by('cliente_denominacion')

    # Obtener los pagos asociados a cada comprobante
    for row in queryset:
        deudas = Deuda.objects.filter(
            content_type_id__in=[ContentType.objects.get_for_model(FacturaVenta), ContentType.objects.get_for_model(BoletaVenta)],
            id_registro=row['id']
        )
        notas = Nota.objects.values(
            'id', 'monto', 
        )

        pagos = Pago.objects.filter(
            deuda__in=deudas,
            content_type_id=ContentType.objects.get_for_model(Nota),
            id_registro__in = Nota.objects.values_list('id', flat=True),
        ).annotate(
            monto_total=Sum(Subquery(notas.filter(id=OuterRef('id_registro')).values('monto')[:1])),
        ).values('id_registro', 'monto_total')

        # Crear la cadena de pagos

        letras_str = '\n'.join([
            f"{p['monto_total']:.2f}"
            for p in pagos
            ])
        row['monto'] = letras_str

    # Convertir a DataFrame de Pandas
    df = pd.DataFrame(list(queryset))

    df_reducido = df.dropna(subset=['monto'])
    df_reducido = df_reducido[df_reducido['monto'].astype(str).str.strip() != '']

    print("#######################################################################")
    print(df_reducido, df_reducido.shape[0])
    print("#######################################################################")


    sql_cobranza_nota = ''' (SELECT
        MAX(cn.id) AS id,
        CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
        SUM(cn.monto) as monto_nota_credito,
        MAX(cc.razon_social) AS cliente_denominacion,
        'FACTURA' AS tipo_comprobante
        FROM cobranza_nota cn
        LEFT JOIN cobranza_pago cp
            ON cp.id_registro=cn.id AND cp.content_type_id='%s'
        LEFT JOIN cobranza_deuda cd
            ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
        LEFT JOIN comprobante_venta_facturaventa cvf
            ON cd.id_registro=cvf.id
        LEFT JOIN datos_globales_seriescomprobante dgsc
            ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
        LEFT JOIN clientes_cliente cc
            ON cc.id=cvf.cliente_id
        LEFT JOIN datos_globales_moneda dgm
            ON dgm.id=cn.moneda_id
        WHERE cvf.sociedad_id='%s'
        GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
        ORDER BY 4)
        UNION
        (SELECT
        MAX(cn.id) AS id,
        CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
        SUM(cn.monto) as monto_nota_credito,
        MAX(cc.razon_social) AS cliente_denominacion,
        'BOLETA' AS tipo_comprobante
        FROM cobranza_nota cn
        LEFT JOIN cobranza_pago cp
            ON cp.id_registro=cn.id AND cp.content_type_id='%s'
        LEFT JOIN cobranza_deuda cd
            ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
        LEFT JOIN comprobante_venta_boletaventa cvb
            ON cd.id_registro=cvb.id
        LEFT JOIN datos_globales_seriescomprobante dgsc
            ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
        LEFT JOIN clientes_cliente cc
            ON cc.id=cvb.cliente_id
        LEFT JOIN datos_globales_moneda dgm
            ON dgm.id=cn.moneda_id
        WHERE cvb.sociedad_id='%s'
        GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
        ORDER BY 4) ; ''' % (DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], sociedad.id, DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], sociedad.id)
    query_info = Nota.objects.raw(sql_cobranza_nota)

    info_cobranza_nota = []
    for fila in query_info:
        lista_datos = []
        lista_datos.append(fila.nro_comprobante)
        lista_datos.append(fila.monto_nota_credito)
        lista_datos.append(fila.cliente_denominacion)
        lista_datos.append(fila.tipo_comprobante)
        info_cobranza_nota.append(lista_datos)


    lista_ordenada = sorted(info_cobranza_nota, key=lambda x: x[2])


    print("#######################################################################")
    for sublista in lista_ordenada:
        print(sublista)
    print("#######################################################################") 

    dict_cobranza_nota = {}
    for fila in info_cobranza_nota:
        dict_cobranza_nota[fila[0]+'|'+fila[3]] = fila[1]

    def reporte_ventas():

        # verificar esto..
        sql_guias = '''(SELECT
            MAX(cvf.id) AS id,
            to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'FACTURA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
            STRING_AGG(
                DISTINCT(CONCAT(dgsc2.serie, '-', lpad(CAST(cdg.numero_guia AS TEXT),6,'0'), ' ', to_char(cdg.fecha_emision, 'DD/MM/YYYY'))), '\n') AS documento_guias
            FROM comprobante_venta_facturaventa cvf
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvf.cliente_id
            LEFT JOIN logistica_notasalidadocumento lnsd
                ON lnsd.content_type_id='%s' AND lnsd.id_registro=cvf.confirmacion_id
            LEFT JOIN logistica_notasalida lns
                ON lns.id=lnsd.nota_salida_id
            LEFT JOIN logistica_despacho ld
                ON ld.nota_salida_id=lns.id
            LEFT JOIN comprobante_despacho_guia cdg
                ON cdg.despacho_id=ld.id
            LEFT JOIN datos_globales_seriescomprobante dgsc2
                ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cdg.serie_comprobante_id
            WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
            ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC)
        UNION
            (SELECT 
            MAX(cvb.id) AS id,
            to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'BOLETA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
            STRING_AGG(
                DISTINCT(CONCAT(dgsc2.serie, '-', lpad(CAST(cdg.numero_guia AS TEXT),6,'0'), ' ', to_char(cdg.fecha_emision, 'DD/MM/YYYY'))), '\n') AS documento_guias
            FROM comprobante_venta_boletaventa cvb
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvb.cliente_id
            LEFT JOIN logistica_notasalidadocumento lnsd
                ON lnsd.content_type_id='%s' AND lnsd.id_registro=cvb.confirmacion_id
            LEFT JOIN logistica_notasalida lns
                ON lns.id=lnsd.nota_salida_id
            LEFT JOIN logistica_despacho ld
                ON ld.nota_salida_id=lns.id
            LEFT JOIN comprobante_despacho_guia cdg
                ON cdg.despacho_id=ld.id
            LEFT JOIN datos_globales_seriescomprobante dgsc2
                ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cdg.serie_comprobante_id
            WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
            ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC) ; ''' %(
                DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                DICT_CONTENT_TYPE['cotizacion | confirmacionventa'], 
                DICT_CONTENT_TYPE['comprobante_despacho | guia'], 
                sociedad, 
                fecha_inicio, 
                fecha_fin, 
                DICT_CONTENT_TYPE['comprobante_venta | boletaventa'],
                DICT_CONTENT_TYPE['cotizacion | confirmacionventa'],  
                DICT_CONTENT_TYPE['comprobante_despacho | guia'], 
                sociedad, 
                fecha_inicio, 
                fecha_fin
                )
        query_info_guias = FacturaVenta.objects.raw(sql_guias)

        info_guias = []
        for fila in query_info_guias:
            lista_datos = []
            lista_datos.append(fila.fecha_emision_comprobante)
            lista_datos.append(fila.tipo_comprobante)
            lista_datos.append(fila.nro_comprobante)
            lista_datos.append(fila.documento_guias)
            info_guias.append(lista_datos)

        dict_guias = {}
        for fila in info_guias:
            dict_guias[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[3]

        sql_facturas = ''' (SELECT 
            MAX(cvf.id) AS id,
            to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'FACTURA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            MAX(cvf.total) AS monto_facturado,
            SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_amortizado,
            MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_pendiente,
            (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) AS estado_cobranza,
            to_char(MAX(cvf.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
            MAX(cvf.fecha_vencimiento) - MAX(cvf.fecha_emision) AS dias_credito,
            MAX(cvf.fecha_vencimiento) AS dias_vencimiento,
            '' AS documento_guias,
            '' AS letras,
            '' AS pagos
            FROM comprobante_venta_facturaventa cvf
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvf.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
            LEFT JOIN cobranza_pago cp
                ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_ingreso ci
                ON ci.id=cp.id_registro
            LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                ON dgcb.id=ci.cuenta_bancaria_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=dgcb.moneda_id
            LEFT JOIN cobranza_redondeo cr
                ON cr.deuda_id=cd.id
            WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
            ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC)
            UNION
            (SELECT 
            MAX(cvb.id) AS id,
            to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'BOLETA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            MAX(cvb.total) AS monto_facturado,
            SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_amortizado,
            MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_pendiente,
            (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) AS estado_cobranza,
            to_char(MAX(cvb.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
            MAX(cvb.fecha_vencimiento) - MAX(cvb.fecha_emision) AS dias_credito,
            MAX(cvb.fecha_vencimiento) AS dias_vencimiento,
            '' AS documento_guias,
            '' AS letras,
            '' AS pagos
            FROM comprobante_venta_boletaventa cvb
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvb.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
            LEFT JOIN cobranza_pago cp
                ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_ingreso ci
                ON ci.id=cp.id_registro
            LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                ON dgcb.id=ci.cuenta_bancaria_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=dgcb.moneda_id
            LEFT JOIN cobranza_redondeo cr
                ON cr.deuda_id=cd.id
            WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
            ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC) ; ''' %(
                DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                DICT_CONTENT_TYPE['cobranza | ingreso'], 
                sociedad, 
                fecha_inicio, 
                fecha_fin, 
                DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], 
                DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], 
                DICT_CONTENT_TYPE['cobranza | ingreso'], 
                sociedad, 
                fecha_inicio, 
                fecha_fin
                )
        query_info_facturas = FacturaVenta.objects.raw(sql_facturas)
        
        info_facturas = []
        for fila in query_info_facturas:
            lista_datos = []
            lista_datos.append(fila.fecha_emision_comprobante)
            lista_datos.append(fila.tipo_comprobante)
            lista_datos.append(fila.nro_comprobante)
            lista_datos.append(fila.cliente_denominacion)
            lista_datos.append(fila.monto_facturado)
            lista_datos.append(fila.monto_amortizado)
            lista_datos.append(fila.monto_pendiente)
            lista_datos.append(fila.estado_cobranza)
            lista_datos.append(fila.fecha_vencimiento_comprobante)
            lista_datos.append(fila.dias_credito)
            lista_datos.append(str(fila.dias_vencimiento))
            lista_datos.append(fila.documento_guias)
            lista_datos.append(fila.letras)
            lista_datos.append(fila.pagos)
            info_facturas.append(lista_datos)

        # dict_comprobante = {
        #     'Factura': '1',
        #     'Boleta': '2',
        #     }
        for fila in info_facturas:
            if fila[2] not in list_anulados:
                try:
                    fila[4] = float(fila[4])
                    if fila[5] == None:
                        fila[5] = '0.00'
                    if fila[6] == None:
                        fila[6] = fila[4]
                    fila[5] = float(fila[5])
                    fila[6] = float(fila[6])
                    fila[9] = float(fila[9])
                except:
                    ''
                if fila[10] != '':
                    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
                    fecha1 = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                    fecha2 = datetime.strptime(fila[10], '%Y-%m-%d')
                    dias = (fecha1 - fecha2) / timedelta(days=1)
                    if float(fila[9]) == float(0):
                        fila[10] = ''
                    elif fila[7] != 'CANCELADO':
                        fila[10] = float(dias)
                    else:
                        fila[10] = ''
                if float(dias) > float(0) and fila[7]=='PENDIENTE':
                    fila[7] = 'VENCIDO'
                if fila[0]+'|'+fila[1]+'|'+fila[2] in dict_pagos:
                    fila[13] = dict_pagos[fila[0]+'|'+fila[1]+'|'+fila[2]]
                if fila[2] + '|' + fila[1] in dict_cobranza_nota:
                    fila[5] += float(dict_cobranza_nota[fila[2] + '|' + fila[1]])
                    fila[6] = fila[4] - fila[5]
                    if fila[6] <= float(0):
                        fila[7] = 'CANCELADO'
                if fila[0]+'|'+fila[1]+'|'+fila[2] in dict_guias:
                    fila[11] = dict_guias[fila[0]+'|'+fila[1]+'|'+fila[2]]
            else:
                for i in range(13):
                    if i == 3:
                        fila[i] = 'ANULADO'
                    elif i > 3:
                        fila[i] = ''

        info = info_facturas + info_notas
        info_ordenado = sorted(info, key=lambda x:formatearFecha3(x[0]))

        wb = Workbook()

        count = 0
        list_general = []
        list_temp = []
        for fila in info_ordenado:
            if count != 0:
                if fila[0][3:5] != info_ordenado[count-1][0][3:5]:
                    list_general.append(list_temp)
                    list_temp = []
            list_temp.append(fila)
            count += 1
        list_general.append(list_temp)
        # if list_general == [] and list_temp != []:
            # list_general.append(list_temp)
        # print('Listas encontradas:',len(list_general))

        count = 0
        total_facturado = 0
        list_montos_totales = []
        for info in list_general:
            name_sheet = DICT_MESES[str(info[0][0][3:5])] + ' - ' + str(info[0][0][6:])
            if count != 0:
                hoja = wb.create_sheet(name_sheet)
                # wb.active = hoja
            else:
                hoja = wb.active
                hoja.title = name_sheet

            hoja.append(('FECHA', 'TIPO DE COMP.', 'N° COMPROB.', 'RAZON SOCIAL', 'FACTURADO (US$)', 'AMORITZADO (US$)', 'PENDIENTE(US$)', 'ESTADO', 'FECHA DE VENC.', 'CRÉDITO', 'DIAS DE VENC.', 'GUIAS', 'LETRAS', 'PAGOS')) # Crea la fila del encabezado con los títulos

            color_relleno = rellenoSociedad(sociedad)

            col_range = hoja.max_column  # get max columns in the worksheet
            # cabecera de la tabla
            for col in range(1, col_range + 1):
                cell_header = hoja.cell(1, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA

            total_facturado_mes = 0
            for producto in info:
                hoja.append(producto)
                try:
                    total_facturado_mes += producto[4]
                except Exception as e:
                    print(e)

            hoja.append(('','','','TOTAL:',total_facturado_mes))

            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if 4 <= col <= 6:
                        row[col].alignment = ALINEACION_DERECHA
                        row[col].number_format = FORMATO_DOLAR
                    if col == 9 or col == 10:
                        row[col].alignment = ALINEACION_DERECHA
                    if 11 <= col <= 13:
                        row[col].alignment = AJUSTAR_TEXTO

            ajustarColumnasSheet(hoja)
            list = [name_sheet, total_facturado_mes] # Agregar mes en 0.00
            list_montos_totales.append(list)
            total_facturado += total_facturado_mes
            count += 1

        list_meses = [
            'ENERO',
            'FEBRERO',
            'MARZO',
            'ABRIL',
            'MAYO',
            'JUNIO',
            'JULIO',
            'AGOSTO',
            'SETIEMBRE',
            'OCTUBRE',
            'NOVIEMBRE',
            'DICIEMBRE'
            ]

        dict_meses_valor = {
            'ENERO' : 1,
            'FEBRERO': 2,
            'MARZO': 3,
            'ABRIL': 4,
            'MAYO': 5,
            'JUNIO': 6,
            'JULIO': 7,
            'AGOSTO': 8,
            'SETIEMBRE': 9,
            'OCTUBRE': 10,
            'NOVIEMBRE': 11,
            'DICIEMBRE' : 12,
        }

        # print()
        # print(list_montos_totales)
        # print()

        count = 0
        list_general = []
        list_temp = []
        for fila in list_montos_totales:
            if count != 0:
                mes_resumen = fila[0].split(" - ")[0]
                mes_anterior = list_montos_totales[count-1][0].split(" - ")[0]
                if list_meses.index(mes_resumen) <= list_meses.index(mes_anterior):
                    list_general.append(list_temp)
                    list_temp = []
            list_temp.append(fila)
            count += 1
        list_general.append(list_temp)

        # MES/AÑO MONTO FACTURADO CERO
        i = 0
        for list_resumen_anual in list_general:
            if i != 0:
                now_year = list_resumen_anual[-1][0].split(' - ')[1]
                past_year = list_general[i-1][-1][0].split(' - ')[1]
                diff_years = int(now_year) - int(past_year)
                while diff_years > 1:
                    name_year_faltante = 'ENERO - ' + str(int(past_year) + diff_years - 1)
                    list_general.insert(i, [[name_year_faltante, 0.0]])
                    diff_years -= 1
            i += 1

        list_resumen_anuales = []
        for list_anual in list_general:
            list_temp_anual = []
            for fila in list_anual:
                list_temp_anual.append(fila[0].split(" - ")[0])
            for mes_calendario in list_meses:
                if mes_calendario not in list_temp_anual:
                    list_anual.insert(dict_meses_valor[mes_calendario]-1, [mes_calendario + " - " + fila[0].split(" - ")[1] , 0])

            list_resumen_anuales += list_anual

        color_relleno = rellenoSociedad(sociedad)

        hoja = wb.create_sheet('Resumen')
        hoja.append(('ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
        hoja.append(('', ''))
        hoja.append(('Periodo','Monto'))
        # for col in range(1, hoja.max_column + 1):
        for col in range(1, 2 + 1):
            cell_header = hoja.cell(3, col)
            cell_header.fill = color_relleno
            cell_header.font = NEGRITA

        for fila in list_resumen_anuales: # aqui! ANTES: "list_montos_totales"
            hoja.append(fila)
        # hoja.append(('TOTAL:',total_facturado)) # Es la sumatoria de todos los montos

        for row in hoja.rows:
            # for col in range(hoja.max_column):
            for col in range(2):
                row[col].border = BORDE_DELGADO
                if col == 1:
                    row[col].number_format = FORMATO_DOLAR
        ajustarColumnasSheet(hoja)


        # Insertar Gráfico

        def grafico_resumen_ingresos(ws):
            max_fila = ws.max_row - 3 # 3 filas de espacio antes de los meses

            chart = LineChart()
            # print(help(chart))
            chart.height = 15 # default is 7.5
            chart.width = 30 # default is 15
            chart.y_axis.title = 'VENTAS FACTURDAS'
            chart.x_axis.title = 'MESES'
            chart.legend.position = 'b' #bottom
            # chart.style = 12

            count = 1
            fila_base = 4
            year_base = int(ws['A' + str(fila_base)].value.split(' - ')[1]) - count # 2018
            data_col = 2 # montos en dolares
            if sociedad == '2':
                chart.title = 'RESUMEN VENTAS FACTURADAS - MULTIPLAY'
            if sociedad == '1':
                chart.title = 'RESUMEN VENTAS FACTURADAS - MULTICABLE'
            while max_fila >= 12:
                values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                # series = Series(values, title = "Ventas del " + str(year_base + count))
                year_referencia = int(ws['A' + str(fila_base + 12*(count-1))].value.split(' - ')[1]) # 2018
                series = Series(values, title = "Ventas del " + str(year_referencia))
                chart.append(series)
                max_fila -= 12
                count += 1
            if 1 <= max_fila <= 12:
                values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                # series = Series(values, title = "Ventas del " + str(year_base + count))
                year_referencia = int(ws['A' + str(fila_base + 12*(count-1))].value.split(' - ')[1]) # 2018
                series = Series(values, title = "Ventas del " + str(year_referencia))
                chart.append(series)

            meses = Reference(ws, min_col = 1, min_row = 1, max_col = 12, max_row = 1)
            chart.set_categories(meses)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.dLblPos = 't' # top

            chart.plot_area.dTable = DataTable()
            chart.plot_area.dTable.showHorzBorder = True
            chart.plot_area.dTable.showVertBorder = True
            chart.plot_area.dTable.showOutline = True
            chart.plot_area.dTable.showKeys = True

            ws.add_chart(chart)

        grafico_resumen_ingresos(hoja)


    return wb


