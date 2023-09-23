from datetime import date, datetime
from django.contrib.contenttypes.models import ContentType
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import*
from openpyxl.styles.borders import Border, Side

from applications.clientes.models import Cliente
from applications.sociedad.models import Sociedad


DICT_CONTENT_TYPE = {}
query_content_type = ContentType.objects.all()
for fila in query_content_type:
    c_type = str(fila.app_label)+' | '+str(fila.model)
    DICT_CONTENT_TYPE[c_type] = fila.id

DICT_SOCIEDAD = {}
query_sociedad = Sociedad.objects.all()
for dato in query_sociedad:
    c_id = str(dato.id)
    DICT_SOCIEDAD[c_id] = dato

DICT_TIPO_NOTA_CREDITO = {
    '1' :"ANULACIÓN DE LA OPERACIÓN",
    '2' :"ANULACIÓN POR ERROR EN EL RUC",
    '3' :"CORRECCIÓN POR ERROR EN LA DESCRIPCIÓN",
    '4' :"DESCUENTO GLOBAL",
    '5' :"DESCUENTO POR ÍTEM",
    '6' :"DEVOLUCIÓN TOTAL",
    '7' :"DEVOLUCIÓN POR ÍTEM",
    '8' :"BONIFICACIÓN",
    '9' :"DISMINUCIÓN EN EL VALOR",
    '10':"OTROS CONCEPTOS",
    '11':"AJUSTES AFECTOS AL IVAP",
    '12':"AJUSTES DE OPERACIONES DE EXPORTACIÓN",
    '13':"AJUSTES - MONTOS Y/O FECHAS DE PAGO"
    }

DICT_MESES = {
    '01': 'ENERO',
    '02': 'FEBRERO',
    '03': 'MARZO',
    '04': 'ABRIL',
    '05': 'MAYO',
    '06': 'JUNIO',
    '07': 'JULIO',
    '08': 'AGOSTO',
    '09': 'SETIEMBRE',
    '10': 'OCTUBRE',
    '11': 'NOVIEMBRE',
    '12': 'DICIEMBRE'
    }


# cod_soc = input('Ingrese el Codigo de Sociedad: ')
# cargarEstilosSheet()
RELLENO_MC = PatternFill(start_color='FFCA0B', end_color='FFCA0B', fill_type='solid')
RELLENO_MP = PatternFill(start_color='8FC6EA', end_color='8FC6EA', fill_type='solid')
# relleno_mp_mc = PatternFill(start_color='00BB2D', end_color='00BB2D', fill_type='solid')
RELLENO_MP_MC = PatternFill(start_color='77DD77', end_color='77DD77', fill_type='solid')
RELLENO_EXCEL = PatternFill(start_color='77DD77', end_color='77DD77', fill_type='solid')

def rellenoSociedad(id_sociedad):
    if id_sociedad != 'None':
        if id_sociedad:
            color_sociedad = str(DICT_SOCIEDAD[id_sociedad].color)[1:]
            return PatternFill(start_color=color_sociedad, end_color=color_sociedad, fill_type='solid')
    else:
        return PatternFill(start_color='77DD77', end_color='77DD77', fill_type='solid')

def rellenoSociedadCorregido(sociedad):
    if sociedad:
        color_sociedad = str(sociedad.color)[1:]
        return PatternFill(start_color=color_sociedad, end_color=color_sociedad, fill_type='solid')
    else:
        return PatternFill(start_color='77DD77', end_color='77DD77', fill_type='solid')

ALINEACION_DERECHA = Alignment(horizontal='right')
ALINEACION_CENTRO = Alignment(horizontal='center')
AJUSTAR_TEXTO = Alignment(wrap_text=True)
BORDE_DELGADO = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
NEGRITA = Font(bold=True)
COLOR_ROJO = Font(color='FFFF0000')
COLOR_AZUL = Font(color=colors.BLUE)
FORMATO_NUMERO = '#,##0.00'
FORMATO_DOLAR = '$ #,##0.00'
FORMATO_SOLES = '"S/" #,##0.00'

def ajustarColumnasSheet(hoja):
    dims = {}
    for row in hoja.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), round((len(str(cell.value))+2)*1.2,0) ))
    for col, value in dims.items():
        hoja.column_dimensions[col].width = value

FECHA_HOY = datetime.now().strftime("%Y-%m-%d")
fecha_doc = datetime.now().strftime("%d-%m-%Y")

def formatearFecha3(fecha):
    # if fecha=="":
    if not fecha:
        return ""
    fecha=fecha.split("/")
    fecha.reverse()
    return "/".join(fecha)

def StrToDate(FechaString):
    if FechaString == "None" or FechaString == None or FechaString == "":
        return date(2000,1,1)
    else:
        return date(int(FechaString.split("-")[0]), int(FechaString.split("-")[1]),int(FechaString.split("-")[2]))

def formatoFechaTexto(fecha):
    meses = ("Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre")
    dia = fecha.day
    mes = meses[fecha.month - 1]
    año = fecha.year
    fecha_texto = "{} de {} del {}".format(dia, mes, año)
    return fecha_texto

def formatearDecimal(str, nro):
    try:
        decimal = float(str)
        decimalRound = round(decimal,int(nro))
        cantDecimales = "{:,." + nro + "f}"
        decimalStr = cantDecimales.format(decimalRound)
        return decimalStr
    except:
        ""