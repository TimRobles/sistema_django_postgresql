from django.shortcuts import render
import time
from datetime import datetime, timedelta
from applications.importaciones import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import *
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import Reference, Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable

from applications.datos_globales.models import CuentaBancariaSociedad, Departamento
from applications.sociedad.models import Sociedad
from applications.cobranza.models import Nota, Pago
from applications.nota.models import NotaCredito
from applications.material.models import Material
from applications.comprobante_venta.models import FacturaVenta, FacturaVentaDetalle
from django.contrib.contenttypes.models import ContentType

from applications.reportes.forms import (
    ReporteComportamientoClienteExcelForm,
    ReporteDepositoCuentasBancariasForm,
    ReporteTasaConversionClienteForm,
    ReporteFacturacionAsesorComercialExcelForm,
    ReporteFacturacionGeneralExcelForm,
    ReporteResumenStockProductosForm, 
    ReporteStockSociedadPdfForm, 
    ReporteVentasDepartamentoPdfForm,
    ReportesContadorForm, 
    ReportesFiltrosForm, 
    ReporteVentasDepartamentoExcelForm,
    ReportesRotacionForm
    )
from applications.reportes.funciones import *
from applications.reportes.data_resumen_ingresos_anterior import*
from applications.pdf import*
from applications.reportes.pdf import (
    generar_reporte_cobranza, 
    generarReporteCobranza, 
    generarReporteDeudas, 
    generarReporteResumenStockProductos, 
    generarReporteStockMalogradoSociedad, 
    generarReporteStockSociedad, 
    generarReporteVentasDepartamento, 
    reporte_cobranza
    )
from applications.movimiento_almacen.models import MovimientosAlmacen
from applications.comprobante_compra.models import ComprobanteCompraPIDetalle
from applications.variables import ESTADOS_CLIENTE, MEDIO, ESTADOS_EVENTO_CRM
from applications.reportes.excel import (
    ReporteClienteCRMExcel,
    ReporteComportamientoCliente,
    ReporteContadorCorregido,
    ReporteDepositoCuentasBancariasCorregido,
    ReporteFacturacionAsesorComercial, 
    ReporteFacturacionGeneral,
    ReporteResumenStockProductosCorregido,
    ReporteRotacionCorregido,
    ReporteTasaConversionCliente, 
    ReporteVentasDepartamento
    )
from applications.sede.models import Sede
    
class ReportesView(FormView):
    template_name = "reportes/inicio.html"
    form_class = ReportesFiltrosForm
    success_url = '.'

    def get_form_kwargs(self):
        kwargs = super(ReportesView, self).get_form_kwargs()
        kwargs['filtro_sociedad'] = self.request.GET.get('sociedad')
        kwargs['filtro_fecha_inicio'] = self.request.GET.get('fecha_inicio')
        kwargs['filtro_fecha_fin'] = self.request.GET.get('fecha_fin')
        kwargs['filtro_cliente'] = self.request.GET.get('cliente')
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contexto_filtros = []
        filtro_sociedad = self.request.GET.get('sociedad')
        filtro_fecha_inicio = self.request.GET.get('fecha_inicio')
        filtro_fecha_fin = self.request.GET.get('fecha_fin')
        filtro_cliente = self.request.GET.get('cliente')
        contexto_filtros.append(f"filtro_sociedad={filtro_sociedad}")
        contexto_filtros.append(f"filtro_fecha_inicio={filtro_fecha_inicio}")
        contexto_filtros.append(f"filtro_fecha_fin={filtro_fecha_fin}")
        contexto_filtros.append(f"filtro_cliente={filtro_cliente}")
        context["contexto_filtros"] = "&".join(contexto_filtros)
        return context
    

class ReporteContador(TemplateView):
    def get(self,request, *args,**kwargs):

        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')
        
        def consultaNotasContador():

            sql = ''' (SELECT
                MAX(nnc.id) AS id,
                to_char(MAX(nnc.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_nota,
                (CASE WHEN nnc.tipo_comprobante='3' THEN 'NOTA DE CRÉDITO' ELSE '-' END) as comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(nnc.numero_nota AS TEXT), 6, '0')) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cc.numero_documento) AS ruc,
                CONCAT(MAX(dgsc2.serie), '-', MAX(lpad(CAST(cvf.numero_factura AS TEXT), 6, '0'))) AS comprobante_modifica,
                '' AS obs,
                STRING_AGG(CAST(ROUND(nncd.cantidad, 2) AS TEXT), ' | ') AS cantidad,
                STRING_AGG(mm.descripcion_corta, ' | ') AS productos,
                STRING_AGG(CAST(ROUND(nncd.precio_unitario_sin_igv, 2) AS TEXT), ' | ') AS precios,
                MAX(nnc.descuento_global) AS dscto_global,
                ROUND(SUM(nncd.sub_total), 2) AS valor_venta,
                SUM(nncd.igv)AS igv,
                ROUND(SUM(nncd.total), 2) AS total_venta,
                MAX(nnc.tipo_nota_credito) AS motivo_nota,
                MAX(nnc.nubefact) AS url_nota
                FROM nota_notacredito nnc
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=nnc.serie_comprobante_id AND dgsc.id='10'
                LEFT JOIN clientes_cliente cc
                    ON cc.id=nnc.cliente_id
                LEFT JOIN datos_globales_documentofisico dgdf
                    ON dgdf.id=nnc.content_type_documento_id AND dgdf.modelo_id='%s'
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=nnc.id_registro_documento
                LEFT JOIN datos_globales_seriescomprobante dgsc2
                    ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cvf.serie_comprobante_id
                LEFT JOIN nota_notacreditodetalle nncd
                    ON nnc.id=nncd.nota_credito_id
                LEFT JOIN material_material mm
                    ON nncd.content_type_id='%s' AND mm.id=nncd.id_registro
                WHERE dgsc.serie!='' AND nnc.sociedad_id='%s' AND '%s' <= nnc.fecha_emision AND nnc.fecha_emision <= '%s'
                GROUP BY nnc.sociedad_id, nnc.tipo_comprobante, nnc.serie_comprobante_id, nnc.numero_nota) 
                UNION
                (SELECT
                MAX(nnc.id) AS id,
                to_char(MAX(nnc.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_nota,
                (CASE WHEN nnc.tipo_comprobante='3' THEN 'NOTA DE CRÉDITO' ELSE '-' END) as comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(nnc.numero_nota AS TEXT), 6, '0')) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cc.numero_documento) AS ruc,
                CONCAT(MAX(dgsc2.serie), '-', MAX(lpad(CAST(cvb.numero_boleta AS TEXT), 6, '0'))) AS comprobante_modifica,
                '' AS obs,
                STRING_AGG(CAST(ROUND(nncd.cantidad, 2) AS TEXT), ' | ') AS cantidad,
                STRING_AGG(mm.descripcion_corta, ' | ') AS productos,
                STRING_AGG(CAST(ROUND(nncd.precio_unitario_sin_igv, 2) AS TEXT), ' | ') AS precios,
                MAX(nnc.descuento_global) AS dscto_global,
                ROUND(SUM(nncd.sub_total), 2) AS valor_venta,
                SUM(nncd.igv)AS igv,
                ROUND(SUM(nncd.total), 2) AS total_venta,
                MAX(nnc.tipo_nota_credito) AS motivo_nota,
                MAX(nnc.nubefact) AS url_nota
                FROM nota_notacredito nnc
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=nnc.serie_comprobante_id AND dgsc.id='11'
                LEFT JOIN clientes_cliente cc
                    ON cc.id=nnc.cliente_id
                LEFT JOIN datos_globales_documentofisico dgdf
                    ON dgdf.id=nnc.content_type_documento_id AND dgdf.modelo_id='%s'
                LEFT JOIN comprobante_venta_boletaventa cvb
                    ON cvb.id=nnc.id_registro_documento
                LEFT JOIN datos_globales_seriescomprobante dgsc2
                    ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cvb.serie_comprobante_id
                LEFT JOIN nota_notacreditodetalle nncd
                    ON nnc.id=nncd.nota_credito_id
                LEFT JOIN material_material mm
                    ON nncd.content_type_id='%s' AND mm.id=nncd.id_registro
                WHERE dgsc.serie!='' AND nnc.sociedad_id='%s' AND '%s' <= nnc.fecha_emision AND nnc.fecha_emision <= '%s'
                GROUP BY nnc.sociedad_id, nnc.tipo_comprobante, nnc.serie_comprobante_id, nnc.numero_nota) 
                ORDER BY fecha_emision_nota, nro_comprobante ;''' % (DICT_CONTENT_TYPE['nota | notacredito'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['material | material'], global_sociedad,global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['nota | notacredito'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['material | material'], global_sociedad,global_fecha_inicio, global_fecha_fin)
            query_info = NotaCredito.objects.raw(sql)
            
            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_nota)
                lista_datos.append(fila.comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.ruc)
                lista_datos.append(fila.comprobante_modifica)
                lista_datos.append(fila.obs)
                lista_datos.append(fila.cantidad)
                lista_datos.append(fila.productos)
                lista_datos.append(fila.precios)
                lista_datos.append(fila.dscto_global)
                lista_datos.append(fila.valor_venta)
                lista_datos.append(fila.igv)
                lista_datos.append(fila.total_venta)
                lista_datos.append(str(fila.motivo_nota))
                lista_datos.append(fila.url_nota)
                info.append(lista_datos)

            for fila in info:
                fila[10] =float(fila[10])
                fila[11] =float(fila[11])
                fila[12] = float(fila[12])
                fila[13] = float(fila[13])
                fila[14] = DICT_TIPO_NOTA_CREDITO[fila[14]]
            return info


        def consultaFacturasContador():
            # global_fecha_inicio = '2022-01-01'
            # global_fecha_fin = '2022-01-31'

            sql_anuladas = ''' (SELECT
                MAX(cvf.id) as id,
                CONCAT(MAX(dgsc.serie), '-', MAX(lpad(CAST(cvf.numero_factura AS TEXT), 6, '0'))) as nro_comprobante
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                WHERE cvf.estado='3' and cvf.sociedad_id='%s'
                GROUP BY cvf.sociedad_id, cvf.serie_comprobante_id, cvf.numero_factura)
                UNION
                (SELECT
                MAX(cvb.id) as id,
                CONCAT(MAX(dgsc.serie), '-', MAX(lpad(CAST(cvb.numero_boleta AS TEXT), 6, '0'))) as nro_comprobante
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                WHERE cvb.estado='3' and cvb.sociedad_id='%s'
                GROUP BY cvb.sociedad_id, cvb.serie_comprobante_id, cvb.numero_boleta) ; ''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
            query_info_anulados = FacturaVenta.objects.raw(sql_anuladas)

            info_anulados = []
            for fila in query_info_anulados:
                list_temp = []
                list_temp.append(fila.nro_comprobante)
                info_anulados.append(list_temp)  

            # print('****************************')
            # print(info_anulados)
            # print('****************************')

            list_anulados = []
            for fact in info_anulados:
                list_anulados.append(fact[0])

            sql = ''' (SELECT
                MAX(cvf.id) AS id,
                MAX(cvf.fecha_emision) AS fecha_orden,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', MAX(lpad(CAST(cvf.numero_factura AS TEXT), 6, '0'))) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cc.numero_documento) AS ruc,
                STRING_AGG(mm.descripcion_corta, ' | ') AS productos,
                STRING_AGG(CAST(ROUND(cvfd.cantidad, 2) AS TEXT), ' | ') AS cantidad,
                STRING_AGG(CAST(ROUND(cvfd.precio_final_con_igv, 2) AS TEXT), ' | ') AS precio_final,
                ROUND((SUM(cvfd.total)-ROUND((MAX(cvf.descuento_global)*1.18),2))/1.18,2) AS monto,
                MAX(cvf.total_igv) AS igv,
                MAX(cvf.descuento_global) AS dscto_global,
                SUM(cvfd.total)-ROUND((MAX(cvf.descuento_global)*1.18),2) AS total_dolares,
                ROUND(MAX(dgtc.tipo_cambio_venta), 2) AS tipo_cambio_fact,
                ROUND((SUM(cvfd.total)-ROUND((MAX(cvf.descuento_global)*1.18),2))*MAX(dgtc.tipo_cambio_venta),2) AS total_soles,
                '' AS observaciones,
                (CASE WHEN MAX(cvf.nubefact) IS NOT NULL THEN MAX(cvf.nubefact) ELSE MAX(dgnr.respuesta->>'enlace') END) AS link_nubefact
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN comprobante_venta_facturaventadetalle cvfd
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN material_material mm
                    ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
                LEFT JOIN datos_globales_tipocambiosunat dgtc
                    ON dgtc.fecha=cvf.fecha_emision
                LEFT JOIN datos_globales_nubefactrespuesta dgnr
                    ON dgnr.content_type_id='%s' AND dgnr.id_registro=cvf.id AND dgnr.error=False AND dgnr.id=(select max(id) from datos_globales_nubefactrespuesta  where content_type_id='%s' AND id_registro=dgnr.id_registro AND dgnr.error=False)
                WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura)
                UNION
                (SELECT
                MAX(cvb.id) as id,
                MAX(cvb.fecha_emision) AS fecha_orden,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', MAX(lpad(CAST(cvb.numero_boleta AS TEXT), 6, '0'))) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cc.numero_documento) AS ruc,
                STRING_AGG(mm.descripcion_corta, ' | ') AS productos,
                STRING_AGG(CAST(ROUND(cvbd.cantidad, 2) AS TEXT), ' | ') AS cantidad,
                STRING_AGG(CAST(ROUND(cvbd.precio_final_con_igv, 2) AS TEXT), ' | ') AS precio_final,
                ROUND((SUM(cvbd.total)-ROUND((MAX(cvb.descuento_global)*1.18),2))/1.18,2) AS monto,
                MAX(cvb.total_igv) AS igv,
                MAX(cvb.descuento_global) AS dscto_global,
                SUM(cvbd.total)-ROUND((MAX(cvb.descuento_global)*1.18),2) AS total_dolares,
                ROUND(MAX(dgtc.tipo_cambio_venta), 2) AS tipo_cambio_fact,
                ROUND((SUM(cvbd.total)-ROUND((MAX(cvb.descuento_global)*1.18),2))*MAX(dgtc.tipo_cambio_venta),2) as total_soles,
                '' AS observaciones,
                (CASE WHEN MAX(cvb.nubefact) IS NOT NULL THEN MAX(cvb.nubefact) ELSE MAX(dgnr.respuesta->>'enlace') END) AS link_nubefact
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN comprobante_venta_boletaventadetalle cvbd
                    ON cvb.id=cvbd.boleta_venta_id
                LEFT JOIN material_material mm
                    ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
                LEFT JOIN datos_globales_tipocambiosunat dgtc
                    ON dgtc.fecha=cvb.fecha_emision
                LEFT JOIN datos_globales_nubefactrespuesta dgnr
                    ON dgnr.content_type_id='%s' AND dgnr.id_registro=cvb.id AND dgnr.error=False AND dgnr.id=(select max(id) from datos_globales_nubefactrespuesta  where content_type_id='%s' AND id_registro=dgnr.id_registro AND dgnr.error=False)
                WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta)
                ORDER BY fecha_orden, nro_comprobante  ;''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['material | material'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['material | material'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad, global_fecha_inicio, global_fecha_fin)
            query_info = FacturaVenta.objects.raw(sql)

            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.ruc)
                lista_datos.append(fila.productos)
                lista_datos.append(fila.cantidad)
                lista_datos.append(fila.precio_final)
                lista_datos.append(fila.monto)
                lista_datos.append(fila.igv)
                lista_datos.append(fila.dscto_global)
                lista_datos.append(fila.total_dolares)
                lista_datos.append(fila.tipo_cambio_fact)
                lista_datos.append(fila.total_soles)
                lista_datos.append(fila.observaciones)
                lista_datos.append(fila.link_nubefact)
                info.append(lista_datos)

            for fila in info:
                if fila[2] not in list_anulados:
                    try:
                        fila[8] = float(fila[8])
                        fila[9] = float(fila[9])
                        fila[10] =float(fila[10])
                        fila[11] =float(fila[11])
                        fila[12] = round(float(fila[12]),2)
                        fila[13] = float(fila[13])
                    except:
                        ''
                else:
                    for i in range(16):
                        if i == 3:
                            fila[i] = 'ANULADO'
                        elif i > 3:
                            fila[i] = ''
            return info


        def reporte_facturas_contador():

            wb = Workbook()
            #Para la primera hoja
            hoja = wb.active
            hoja.title = 'sheet'

            # Encabezado
            hoja.append(('FECHA', 'TIPO DE COMP.', 'N° COMPROB.', 'RAZON SOCIAL', 'RUC', 'PRODUCTOS', 'CANT.', 'PRECIO UNIT. (US$) CON IGV', 'MONTO (US$)', 'IGV (US$)', 'DESCUENTO GLOBAL', 'TOTAL (US$)', 'TIPO DE CAMBIO', 'MONTO SOLES (S/)', 'OBSERVACIONES', 'LINK')) # Crea la fila del encabezado con los títulos

            # self.relleno_mc = PatternFill(start_color='FFCA0B', end_color='FFCA0B', fill_type='solid')
            # wb['A1'].fill = redFill

            # for cell in hoja['A1:N1']:
            #     for _ in cell:
            #         _.fill = redFill
            
            color_relleno = rellenoSociedad(global_sociedad)

            col_range = hoja.max_column  # get max columns in the worksheet
            # cabecera de la tabla
            for col in range(1, col_range + 1):
                cell_header = hoja.cell(1, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA

            info = consultaFacturasContador()
            for producto in info:
                hoja.append(producto) # Crea la fila del encabezado con los títulos

            # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11, M=12, N=13
            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if 8 <= col <=11:
                        row[col].alignment = ALINEACION_DERECHA
                        row[col].number_format = FORMATO_DOLAR
                    elif col == 13:
                        row[col].alignment = ALINEACION_DERECHA
                        row[col].number_format = FORMATO_SOLES
                    elif col == 15:
                        if row[col].value != 'LINK':
                            row[col].hyperlink =  row[col].value
                            row[col].font =  COLOR_AZUL
                        # print(row[col].value)
                        # row[col].number_format = self.formato_soles
                # # sheet.cell(fila,3).number_format = '$ #,##0.00'
                # cell_N.number_format = self.formato_dolar

            info2 = consultaNotasContador()
            if info2 != []:
                # cabecera de la tabla
                hoja.append(('',)) # Crea la fila del encabezado con los títulos
                hoja.append(('FECHA', 'TIPO DE COMP.', 'N° COMPROB.', 'RAZON SOCIAL', 'RUC', 'COMPROBANTE QUE SE MODIFICA', '', 'CANT.', 'DESCRIPCION', 'PRECIO UNIT. (US$) SIN IGV', 'DESCUENTO GLOBAL', 'VALOR DE VENTA (US$)', 'IGV (US$)', 'TOTAL (US$)', 'MOTIVO DE LA NOTA', 'LINK')) # Crea la fila del encabezado con los títulos
                nueva_fila = hoja.max_row

                for col in range(1, col_range + 1):
                    cell_header = hoja.cell(nueva_fila, col)
                    cell_header.fill = color_relleno
                    cell_header.font = NEGRITA

                for fila in info2:
                    hoja.append(fila) # Crea la fila del encabezado con los títulos

                for i in range(hoja.max_row):
                    if i >= nueva_fila-1:
                        row = list(hoja.rows)[i]
                        for col in range(hoja.max_column):
                            row[col].border = BORDE_DELGADO
                            if 10 <= col <=13:
                                row[col].alignment = ALINEACION_DERECHA
                                row[col].number_format = FORMATO_DOLAR
                            elif col == 15:
                                if row[col].value != 'LINK':
                                    row[col].hyperlink =  row[col].value
                                    row[col].font =  COLOR_AZUL
            ajustarColumnasSheet(hoja)
            return wb

        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        wb=reporte_facturas_contador()
        nombre_archivo = "Reporte Contador - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteVentasFacturadas(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        print("*******************************")
        print(global_sociedad)
        print(type(global_sociedad))
        print("*******************************")
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def consulta_ventas_notas():
            sql = ''' SELECT
                MAX(nnc.id) AS id,
                to_char(MAX(nnc.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_nota,
                (CASE WHEN nnc.tipo_comprobante='3' THEN 'NOTA DE CRÉDITO' ELSE '-' END) as comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(nnc.numero_nota) AS TEXT), 6, '0')) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                (SUM(nncd.total)*(-1)) AS facturado,
                (SUM(nncd.total)*(-1)) AS amortizado,
                '0.00' AS pendiente,
                'PENDIENTE' AS estado,
                to_char(MAX(nnc.fecha_emision), 'DD/MM/YYYY') AS fecha_vencimiento_nota,
                '0.00' AS credito,
                '' AS dias,
                '' AS guia,
                '' AS obs,
                '' AS letras,
                '' AS pagos
                FROM nota_notacredito nnc
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=nnc.serie_comprobante_id AND dgsc.id='10'
                LEFT JOIN nota_notacreditodetalle nncd
                    ON nnc.id=nncd.nota_credito_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=nnc.cliente_id
                WHERE nnc.sociedad_id='%s' AND '%s' <= nnc.fecha_emision AND nnc.fecha_emision <= '%s'
                GROUP BY nnc.sociedad_id, nnc.tipo_comprobante, nnc.serie_comprobante_id, nnc.numero_nota ; ''' %(DICT_CONTENT_TYPE['nota | notacredito'], global_sociedad, global_fecha_inicio, global_fecha_fin)
            query_info = NotaCredito.objects.raw(sql)
            
            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_nota)
                lista_datos.append(fila.comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.facturado)
                lista_datos.append(fila.amortizado)
                lista_datos.append(fila.pendiente)
                lista_datos.append(fila.estado)
                lista_datos.append(fila.fecha_vencimiento_nota)
                lista_datos.append(fila.credito)
                lista_datos.append(fila.dias)
                lista_datos.append(fila.guia)
                lista_datos.append(fila.obs)
                lista_datos.append(fila.letras)
                lista_datos.append(fila.pagos)
                info.append(lista_datos)

            for fila in info:
                try:
                    fila[4] = float(fila[4])
                    fila[5] = float(fila[5])
                    fila[6] = float(fila[6])
                    fila[9] = float(fila[9])
                except:
                    ''
            return info

        def reporte_ventas():
            info_notas = consulta_ventas_notas()
            # FECHA_INICIO = '2021-12-20'
            # FECHA_FIN = '2022-01-31'
            # sql = ''' SELECT cf.Fecha_Emision, IF(cf.Tipo_Comprobante='1','Factura','Boleta') AS comprobante, CONCAT(cf.Serie, '-', cf.Nro_Facturacion) AS nro_comprobante,
            #     mc.Razon_Social, SUM(df.Sub_Total)-ROUND((cf.Descuento_Global*1.18),2) as facturado, SUM(cob.Monto_Deposito) as total, SUM(cob.Monto_Deposito)-SUM(cob.Monto_Saldo) AS amoritzado, SUM(cob.Monto_Saldo) AS pendiente, IF(SUM(cob.Monto_Saldo)=0.00,'CANCELADO','PENDIENTE') as estado, cf.Fecha_Vencimiento, CONCAT(cg.Serie,'-',cg.Nro_Guia) AS guia, cg.Fecha_Emision, GROUP_CONCAT(CONCAT(dob.Fecha_Operacion,'  $ ',dob.Monto_Deposito) SEPARATOR ' | ')
            # 	FROM `TAB_VENTA_009_Cabecera_Facturacion` cf
            #     LEFT JOIN TAB_VENTA_010_Detalle_Facturacion df
            #         ON cf.Cod_Soc=df.Cod_Soc AND cf.Año=df.Año AND cf.Tipo_Comprobante=df.Tipo_Comprobante AND cf.Serie=df.Serie AND cf.Nro_Facturacion=df.Nro_Facturacion
            #     LEFT JOIN `TAB_COM_001_Maestro Clientes` mc ON cf.Cod_Cliente=mc.Cod_Cliente
            #     LEFT JOIN TAB_VENTA_005_Cabecera_Operaciones_Bancarias_Por_Cotizacion cob ON cob.Cod_Soc=cf.Cod_Soc AND LEFT(cf.Nro_Cotización,4)=cob.Año_Cot_Client AND MID(cf.Nro_Cotización,6,10)=cob.Nro_Cot_Client AND cf.Cod_Cliente=cob.Cod_Cliente
            #     LEFT JOIN TAB_VENTA_006_Detalle_Operaciones_Bancarias_Por_Cotizacion dob ON cob.Cod_Soc=dob.Cod_Emp AND dob.Año_Cot_Client=cob.Año_Cot_Client AND dob.Nro_Cot_Client=cob.Nro_Cot_Client AND dob.Cod_Cliente=cob.Cod_Cliente
            #     LEFT JOIN TAB_VENTA_013_Cabecera_Guia_Remision cg ON CONCAT(cf.Serie,'-',cf.Nro_Facturacion)=cg.Doc_Referencia
            #     WHERE cf.Fecha_Emision != '0000-00-00' AND cf.Cod_Soc='2000'
            #     GROUP BY cf.Cod_Soc, cf.Año, cf.Tipo_Comprobante, cf.Serie, cf.Nro_Facturacion, cob.Cod_Soc, cob.Año_Cot_Client, cob.Nro_Cot_Client, dob.Cod_Emp ;'''

            sql_anuladas = ''' (SELECT
                MAX(cvf.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT), 6, '0')) as nro_comprobante
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                WHERE cvf.estado='3' and cvf.sociedad_id='%s'
                GROUP BY cvf.sociedad_id, cvf.serie_comprobante_id, cvf.numero_factura)
                UNION
                (SELECT
                MAX(cvb.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT), 6, '0')) as nro_comprobante
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                WHERE cvb.estado='3' and cvb.sociedad_id='%s'
                GROUP BY cvb.sociedad_id, cvb.serie_comprobante_id, cvb.numero_boleta) ; '''%(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
            query_info_anulados = FacturaVenta.objects.raw(sql_anuladas)

            info_anulados = []
            for fila in query_info_anulados:
                list_temp = []
                list_temp.append(fila.nro_comprobante)
                info_anulados.append(list_temp)  

            list_anulados = []
            for fact in info_anulados:
                list_anulados.append(fact[0])
            
            
            print(20*'*')
            print(info_anulados)
            print(20*'*')
            print(20*'-')
            print(list_anulados)
            print(20*'-')


            sql_pagos = ''' (SELECT
                MAX(cvf.id) AS id, 
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT), 6, '0')) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(CONCAT(
                            to_char(ci.fecha, 'DD/MM/YYYY'),
                            CONCAT(' ', dgm.simbolo, CAST(ROUND(cp.monto,2) AS TEXT)),
                        (CASE WHEN dgm.abreviatura='PEN'
                        THEN (
                            CONCAT(' ($ ', CAST(ROUND(cp.monto/cp.tipo_cambio,2) AS TEXT), ')')
                        ) ELSE (
                            ''
                        ) END)
                        ), '\n') AS pagos
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                WHERE cvf.sociedad_id='%s' AND cvf.estado='4'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY cliente_denominacion ASC, pagos DESC )
                UNION
                (SELECT 
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT), 6, '0')) as nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(CONCAT(
                            to_char(ci.fecha, 'DD/MM/YYYY'),
                            CONCAT(' ', dgm.simbolo, CAST(ROUND(cp.monto,2) AS TEXT)),
                        (CASE WHEN dgm.abreviatura='PEN'
                        THEN (
                            CONCAT(' ($ ', CAST(ROUND(cp.monto/cp.tipo_cambio,2) AS TEXT), ')')
                        ) ELSE (
                            ''
                        ) END)
                        ), '\n') AS pagos
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                WHERE cvb.sociedad_id='%s' AND cvb.estado='4'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY cliente_denominacion ASC, pagos DESC)
                ORDER BY cliente_denominacion ASC, pagos DESC; '''%(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad,  DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad)
            query_info_pagos = FacturaVenta.objects.raw(sql_pagos)

            info = []
            for fila in query_info_pagos:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.pagos)
                info.append(lista_datos)

            dict_pagos = {}
            for fila in info:
                dict_pagos[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[4]


            sql_letras = ''' (SELECT 
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                (CASE WHEN MAX(cvf.tipo_venta)='2'
                THEN (
                    STRING_AGG(CONCAT(
                        to_char(cobc.fecha, 'DD/MM/YYYY'),
                        ' $ ',
                        CAST(ROUND(cobc.monto,2) AS TEXT)
                        ), '\n')
                ) ELSE (
                    ''
                ) END) AS letras
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
                LEFT JOIN cobranza_cuota cobc
                    ON cobc.deuda_id=cd.id
                WHERE cvf.tipo_venta='2' AND cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s' AND cvf.estado='4'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC)
                UNION
                (SELECT 
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                (CASE WHEN MAX(cvb.tipo_venta)='2'
                THEN (
                    STRING_AGG(CONCAT(
                        to_char(cobc.fecha, 'DD/MM/YYYY'),
                        ' $ ',
                        CAST(ROUND(cobc.monto,2) AS TEXT)
                        ), '\n')
                ) ELSE (
                    ''
                ) END) AS letras
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
                LEFT JOIN cobranza_cuota cobc
                    ON cobc.deuda_id=cd.id
                WHERE cvb.tipo_venta='2' AND cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s' AND cvb.estado='4'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC) ; ''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad, global_fecha_inicio, global_fecha_fin)
            query_info_letras = FacturaVenta.objects.raw(sql_letras)

            info = []
            for fila in query_info_letras:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.letras)
                info.append(lista_datos)

            dict_letras = {}
            for fila in info:
                dict_letras[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[3]

            sql_cobranza_nota = ''' (SELECT
                MAX(cn.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                SUM(cn.monto) as monto_nota_credito,
                MAX(cc.razon_social) AS cliente_denominacion,
                'FACTURA' AS tipo_comprobante
                FROM cobranza_nota cn
                LEFT JOIN cobranza_pago cp
                    ON cp.id_registro=cn.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_deuda cd
                    ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cd.id_registro=cvf.id
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=cn.moneda_id
                WHERE cvf.sociedad_id='%s'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY 4)
                UNION
                (SELECT
                MAX(cn.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                SUM(cn.monto) as monto_nota_credito,
                MAX(cc.razon_social) AS cliente_denominacion,
                'BOLETA' AS tipo_comprobante
                FROM cobranza_nota cn
                LEFT JOIN cobranza_pago cp
                    ON cp.id_registro=cn.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_deuda cd
                    ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
                LEFT JOIN comprobante_venta_boletaventa cvb
                    ON cd.id_registro=cvb.id
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=cn.moneda_id
                WHERE cvb.sociedad_id='%s'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY 4) ; ''' % (DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
            query_info = Nota.objects.raw(sql_cobranza_nota)

            info_cobranza_nota = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.monto_nota_credito)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.tipo_comprobante)
                info_cobranza_nota.append(lista_datos)

            dict_cobranza_nota = {}
            for fila in info_cobranza_nota:
                dict_cobranza_nota[fila[0]+'|'+fila[3]] = fila[1]

            # verificar esto..
            sql_guias = '''(SELECT
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                STRING_AGG(
                    DISTINCT(CONCAT(dgsc2.serie, '-', lpad(CAST(cdg.numero_guia AS TEXT),6,'0'), ' ', to_char(cdg.fecha_emision, 'DD/MM/YYYY'))), '\n') AS documento_guias
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN logistica_notasalidadocumento lnsd
                    ON lnsd.content_type_id='%s' AND lnsd.id_registro=cvf.confirmacion_id
                LEFT JOIN logistica_notasalida lns
                    ON lns.id=lnsd.nota_salida_id
                LEFT JOIN logistica_despacho ld
                    ON ld.nota_salida_id=lns.id
                LEFT JOIN comprobante_despacho_guia cdg
                    ON cdg.despacho_id=ld.id
                LEFT JOIN datos_globales_seriescomprobante dgsc2
                    ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cdg.serie_comprobante_id
                WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC)
            UNION
                (SELECT 
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                STRING_AGG(
                    DISTINCT(CONCAT(dgsc2.serie, '-', lpad(CAST(cdg.numero_guia AS TEXT),6,'0'), ' ', to_char(cdg.fecha_emision, 'DD/MM/YYYY'))), '\n') AS documento_guias
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN logistica_notasalidadocumento lnsd
                    ON lnsd.content_type_id='%s' AND lnsd.id_registro=cvb.confirmacion_id
                LEFT JOIN logistica_notasalida lns
                    ON lns.id=lnsd.nota_salida_id
                LEFT JOIN logistica_despacho ld
                    ON ld.nota_salida_id=lns.id
                LEFT JOIN comprobante_despacho_guia cdg
                    ON cdg.despacho_id=ld.id
                LEFT JOIN datos_globales_seriescomprobante dgsc2
                    ON dgsc2.tipo_comprobante_id='%s' AND dgsc2.id=cdg.serie_comprobante_id
                WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC) ; ''' %(
                    DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                    DICT_CONTENT_TYPE['cotizacion | confirmacionventa'], 
                    DICT_CONTENT_TYPE['comprobante_despacho | guia'], 
                    global_sociedad, 
                    global_fecha_inicio, 
                    global_fecha_fin, 
                    DICT_CONTENT_TYPE['comprobante_venta | boletaventa'],
                    DICT_CONTENT_TYPE['cotizacion | confirmacionventa'],  
                    DICT_CONTENT_TYPE['comprobante_despacho | guia'], 
                    global_sociedad, 
                    global_fecha_inicio, 
                    global_fecha_fin
                    )
            query_info_guias = FacturaVenta.objects.raw(sql_guias)

            info_guias = []
            for fila in query_info_guias:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.documento_guias)
                info_guias.append(lista_datos)

            dict_guias = {}
            for fila in info_guias:
                dict_guias[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[3]

            sql_facturas = ''' (SELECT 
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cvf.total) AS monto_facturado,
                SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_amortizado,
                MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_pendiente,
                (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) AS estado_cobranza,
                to_char(MAX(cvf.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
                MAX(cvf.fecha_vencimiento) - MAX(cvf.fecha_emision) AS dias_credito,
                MAX(cvf.fecha_vencimiento) AS dias_vencimiento,
                '' AS documento_guias,
                '' AS letras,
                '' AS pagos
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                LEFT JOIN cobranza_redondeo cr
                    ON cr.deuda_id=cd.id
                WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC)
                UNION
                (SELECT 
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cvb.total) AS monto_facturado,
                SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_amortizado,
                MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) AS monto_pendiente,
                (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) AS estado_cobranza,
                to_char(MAX(cvb.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
                MAX(cvb.fecha_vencimiento) - MAX(cvb.fecha_emision) AS dias_credito,
                MAX(cvb.fecha_vencimiento) AS dias_vencimiento,
                '' AS documento_guias,
                '' AS letras,
                '' AS pagos
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                LEFT JOIN cobranza_redondeo cr
                    ON cr.deuda_id=cd.id
                WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY fecha_emision_comprobante ASC, nro_comprobante ASC) ; ''' %(
                    DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                    DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], 
                    DICT_CONTENT_TYPE['cobranza | ingreso'], 
                    global_sociedad, 
                    global_fecha_inicio, 
                    global_fecha_fin, 
                    DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], 
                    DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], 
                    DICT_CONTENT_TYPE['cobranza | ingreso'], 
                    global_sociedad, 
                    global_fecha_inicio, 
                    global_fecha_fin
                    )
            query_info_facturas = FacturaVenta.objects.raw(sql_facturas)
            
            info_facturas = []
            for fila in query_info_facturas:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.monto_facturado)
                lista_datos.append(fila.monto_amortizado)
                lista_datos.append(fila.monto_pendiente)
                lista_datos.append(fila.estado_cobranza)
                lista_datos.append(fila.fecha_vencimiento_comprobante)
                lista_datos.append(fila.dias_credito)
                lista_datos.append(str(fila.dias_vencimiento))
                lista_datos.append(fila.documento_guias)
                lista_datos.append(fila.letras)
                lista_datos.append(fila.pagos)
                info_facturas.append(lista_datos)

            # dict_comprobante = {
            #     'Factura': '1',
            #     'Boleta': '2',
            #     }
            for fila in info_facturas:
                if fila[2] not in list_anulados:
                    try:
                        fila[4] = float(fila[4])
                        if fila[5] == None:
                            fila[5] = '0.00'
                        if fila[6] == None:
                            fila[6] = fila[4]
                        fila[5] = float(fila[5])
                        fila[6] = float(fila[6])
                        fila[9] = float(fila[9])
                    except:
                        ''
                    if fila[10] != '':
                        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
                        fecha1 = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                        fecha2 = datetime.strptime(fila[10], '%Y-%m-%d')
                        dias = (fecha1 - fecha2) / timedelta(days=1)
                        if float(fila[9]) == float(0):
                            fila[10] = ''
                        elif fila[7] != 'CANCELADO':
                            fila[10] = float(dias)
                        else:
                            fila[10] = ''
                    if float(dias) > float(0) and fila[7]=='PENDIENTE':
                        fila[7] = 'VENCIDO'
                    if fila[0]+'|'+fila[1]+'|'+fila[2] in dict_pagos:
                        fila[13] = dict_pagos[fila[0]+'|'+fila[1]+'|'+fila[2]]
                    if fila[2] + '|' + fila[1] in dict_cobranza_nota:
                        fila[5] += float(dict_cobranza_nota[fila[2] + '|' + fila[1]])
                        fila[6] = fila[4] - fila[5]
                        if fila[6] <= float(0):
                            fila[7] = 'CANCELADO'
                    if fila[0]+'|'+fila[1]+'|'+fila[2] in dict_guias:
                        fila[11] = dict_guias[fila[0]+'|'+fila[1]+'|'+fila[2]]
                else:
                    for i in range(13):
                        if i == 3:
                            fila[i] = 'ANULADO'
                        elif i > 3:
                            fila[i] = ''

            info = info_facturas + info_notas
            info_ordenado = sorted(info, key=lambda x:formatearFecha3(x[0]))

            wb = Workbook()

            count = 0
            list_general = []
            list_temp = []
            for fila in info_ordenado:
                if count != 0:
                    if fila[0][3:5] != info_ordenado[count-1][0][3:5]:
                        list_general.append(list_temp)
                        list_temp = []
                list_temp.append(fila)
                count += 1
            list_general.append(list_temp)
            # if list_general == [] and list_temp != []:
                # list_general.append(list_temp)
            # print('Listas encontradas:',len(list_general))

            count = 0
            total_facturado = 0
            list_montos_totales = []
            for info in list_general:
                name_sheet = DICT_MESES[str(info[0][0][3:5])] + ' - ' + str(info[0][0][6:])
                if count != 0:
                    hoja = wb.create_sheet(name_sheet)
                    # wb.active = hoja
                else:
                    hoja = wb.active
                    hoja.title = name_sheet

                hoja.append(('FECHA', 'TIPO DE COMP.', 'N° COMPROB.', 'RAZON SOCIAL', 'FACTURADO (US$)', 'AMORITZADO (US$)', 'PENDIENTE(US$)', 'ESTADO', 'FECHA DE VENC.', 'CRÉDITO', 'DIAS DE VENC.', 'GUIAS', 'LETRAS', 'PAGOS')) # Crea la fila del encabezado con los títulos

                color_relleno = rellenoSociedad(global_sociedad)

                col_range = hoja.max_column  # get max columns in the worksheet
                # cabecera de la tabla
                for col in range(1, col_range + 1):
                    cell_header = hoja.cell(1, col)
                    cell_header.fill = color_relleno
                    cell_header.font = NEGRITA

                total_facturado_mes = 0
                for producto in info:
                    hoja.append(producto)
                    try:
                        total_facturado_mes += producto[4]
                    except Exception as e:
                        print(e)

                hoja.append(('','','','TOTAL:',total_facturado_mes))

                for row in hoja.rows:
                    for col in range(hoja.max_column):
                        row[col].border = BORDE_DELGADO
                        if 4 <= col <= 6:
                            row[col].alignment = ALINEACION_DERECHA
                            row[col].number_format = FORMATO_DOLAR
                        if col == 9 or col == 10:
                            row[col].alignment = ALINEACION_DERECHA
                        if 11 <= col <= 13:
                            row[col].alignment = AJUSTAR_TEXTO

                ajustarColumnasSheet(hoja)
                list = [name_sheet, total_facturado_mes] # Agregar mes en 0.00
                list_montos_totales.append(list)
                total_facturado += total_facturado_mes
                count += 1

            list_meses = [
                'ENERO',
                'FEBRERO',
                'MARZO',
                'ABRIL',
                'MAYO',
                'JUNIO',
                'JULIO',
                'AGOSTO',
                'SETIEMBRE',
                'OCTUBRE',
                'NOVIEMBRE',
                'DICIEMBRE'
                ]

            dict_meses_valor = {
                'ENERO' : 1,
                'FEBRERO': 2,
                'MARZO': 3,
                'ABRIL': 4,
                'MAYO': 5,
                'JUNIO': 6,
                'JULIO': 7,
                'AGOSTO': 8,
                'SETIEMBRE': 9,
                'OCTUBRE': 10,
                'NOVIEMBRE': 11,
                'DICIEMBRE' : 12,
            }

            # print()
            # print(list_montos_totales)
            # print()

            count = 0
            list_general = []
            list_temp = []
            for fila in list_montos_totales:
                if count != 0:
                    mes_resumen = fila[0].split(" - ")[0]
                    mes_anterior = list_montos_totales[count-1][0].split(" - ")[0]
                    if list_meses.index(mes_resumen) <= list_meses.index(mes_anterior):
                        list_general.append(list_temp)
                        list_temp = []
                list_temp.append(fila)
                count += 1
            list_general.append(list_temp)

            # MES/AÑO MONTO FACTURADO CERO
            i = 0
            for list_resumen_anual in list_general:
                if i != 0:
                    now_year = list_resumen_anual[-1][0].split(' - ')[1]
                    past_year = list_general[i-1][-1][0].split(' - ')[1]
                    diff_years = int(now_year) - int(past_year)
                    while diff_years > 1:
                        name_year_faltante = 'ENERO - ' + str(int(past_year) + diff_years - 1)
                        list_general.insert(i, [[name_year_faltante, 0.0]])
                        diff_years -= 1
                i += 1

            list_resumen_anuales = []
            for list_anual in list_general:
                list_temp_anual = []
                for fila in list_anual:
                    list_temp_anual.append(fila[0].split(" - ")[0])
                for mes_calendario in list_meses:
                    if mes_calendario not in list_temp_anual:
                        list_anual.insert(dict_meses_valor[mes_calendario]-1, [mes_calendario + " - " + fila[0].split(" - ")[1] , 0])

                list_resumen_anuales += list_anual

            color_relleno = rellenoSociedad(global_sociedad)

            hoja = wb.create_sheet('Resumen')
            hoja.append(('ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
            hoja.append(('', ''))
            hoja.append(('Periodo','Monto'))
            # for col in range(1, hoja.max_column + 1):
            for col in range(1, 2 + 1):
                cell_header = hoja.cell(3, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA

            for fila in list_resumen_anuales: # aqui! ANTES: "list_montos_totales"
                hoja.append(fila)
            # hoja.append(('TOTAL:',total_facturado)) # Es la sumatoria de todos los montos

            for row in hoja.rows:
                # for col in range(hoja.max_column):
                for col in range(2):
                    row[col].border = BORDE_DELGADO
                    if col == 1:
                        row[col].number_format = FORMATO_DOLAR
            ajustarColumnasSheet(hoja)


            # Insertar Gráfico

            def grafico_resumen_ingresos(ws):
                max_fila = ws.max_row - 3 # 3 filas de espacio antes de los meses

                chart = LineChart()
                # print(help(chart))
                chart.height = 15 # default is 7.5
                chart.width = 30 # default is 15
                chart.y_axis.title = 'VENTAS FACTURDAS'
                chart.x_axis.title = 'MESES'
                chart.legend.position = 'b' #bottom
                # chart.style = 12

                count = 1
                fila_base = 4
                year_base = int(ws['A' + str(fila_base)].value.split(' - ')[1]) - count # 2018
                data_col = 2 # montos en dolares
                if global_sociedad == '2':
                    chart.title = 'RESUMEN VENTAS FACTURADAS - MULTIPLAY'
                if global_sociedad == '1':
                    chart.title = 'RESUMEN VENTAS FACTURADAS - MULTICABLE'
                while max_fila >= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    # series = Series(values, title = "Ventas del " + str(year_base + count))
                    year_referencia = int(ws['A' + str(fila_base + 12*(count-1))].value.split(' - ')[1]) # 2018
                    series = Series(values, title = "Ventas del " + str(year_referencia))
                    chart.append(series)
                    max_fila -= 12
                    count += 1
                if 1 <= max_fila <= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    # series = Series(values, title = "Ventas del " + str(year_base + count))
                    year_referencia = int(ws['A' + str(fila_base + 12*(count-1))].value.split(' - ')[1]) # 2018
                    series = Series(values, title = "Ventas del " + str(year_referencia))
                    chart.append(series)

                meses = Reference(ws, min_col = 1, min_row = 1, max_col = 12, max_row = 1)
                chart.set_categories(meses)
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                chart.dataLabels.dLblPos = 't' # top

                chart.plot_area.dTable = DataTable()
                chart.plot_area.dTable.showHorzBorder = True
                chart.plot_area.dTable.showVertBorder = True
                chart.plot_area.dTable.showOutline = True
                chart.plot_area.dTable.showKeys = True

                ws.add_chart(chart)

            grafico_resumen_ingresos(hoja)


            return wb
            # wb.save('reporte_ventas_prueba.xlsx')

        # reporte_ventas('1000')

        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        wb=reporte_ventas()
        nombre_archivo = "Reporte Ventas Facturadas - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteFacturasPendientes(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def reporte_facturas_pendientes():
            wb = Workbook()
            hoja = wb.active

            color_relleno = rellenoSociedad(global_sociedad)

            sql_productos = ''' (SELECT
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                STRING_AGG(mm.descripcion_corta, ' | ') as productos
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN comprobante_venta_facturaventadetalle cvfd
                    ON cvfd.factura_venta_id=cvf.id
                LEFT JOIN material_material mm
                    ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
                WHERE cvf.estado='4'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura)
                UNION
                (SELECT
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                STRING_AGG(mm.descripcion_corta, ' | ') as productos
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN comprobante_venta_boletaventadetalle cvbd
                    ON cvbd.boleta_venta_id=cvb.id
                LEFT JOIN material_material mm
                    ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
                WHERE cvb.estado='4'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta) ;'''%(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'],DICT_CONTENT_TYPE['material | material'],DICT_CONTENT_TYPE['comprobante_venta | boletaventa'],DICT_CONTENT_TYPE['material | material'],)
            query_info = FacturaVenta.objects.raw(sql_productos)

            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.productos)
                info.append(lista_datos)

            dict_productos = {}
            for fila in info:
                dict_productos[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[3]

            # Letras según fechas pactadas al cotizar
            sql_letras = ''' (SELECT
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                (CASE WHEN MAX(cvf.tipo_venta)='2'
                THEN (
                    STRING_AGG(CONCAT(
                        to_char(cobc.fecha, 'DD/MM/YYYY'),
                        ' $ ',
                        CAST(ROUND(cobc.monto,2) AS TEXT)
                        ), '\n' ORDER BY cobc.fecha)
                ) ELSE (
                    ''
                ) END) AS letras
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
                LEFT JOIN cobranza_cuota cobc
                    ON cobc.deuda_id=cd.id
                WHERE cvf.tipo_venta='2' AND cvf.sociedad_id='%s' AND cd.id IS NOT NULL AND cvf.estado='4'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY cliente_denominacion ASC, letras ASC)
                UNION
                (SELECT
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                (CASE WHEN MAX(cvb.tipo_venta)='2'
                THEN (
                    STRING_AGG(CONCAT(
                        to_char(cobc.fecha, 'DD/MM/YYYY'),
                        ' $ ',
                        CAST(ROUND(cobc.monto,2) AS TEXT)
                        ), '\n' ORDER BY cobc.fecha)
                ) ELSE (
                    ''
                ) END) AS letras
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
                LEFT JOIN cobranza_cuota cobc
                    ON cobc.deuda_id=cd.id
                WHERE cvb.tipo_venta='2' AND cvb.sociedad_id='%s' AND cd.id IS NOT NULL AND cvb.estado='4'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY cliente_denominacion ASC, letras ASC)
                ORDER BY cliente_denominacion ASC, letras ASC ; ''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'],DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'],DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
            query_info = FacturaVenta.objects.raw(sql_letras)

            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.letras)
                info.append(lista_datos)

            dict_letras = {}
            for fila in info:
                dict_letras[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[4]

            sql_cobranza_nota = ''' (SELECT
                MAX(cn.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                SUM(cn.monto) as monto_nota_credito,
                MAX(cc.razon_social) AS cliente_denominacion,
                'FACTURA' AS tipo_comprobante
                FROM cobranza_nota cn
                LEFT JOIN cobranza_pago cp
                    ON cp.id_registro=cn.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_deuda cd
                    ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cd.id_registro=cvf.id
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=cn.moneda_id
                WHERE cvf.sociedad_id='%s'
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                ORDER BY 4)
                UNION
                (SELECT
                MAX(cn.id) AS id,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                SUM(cn.monto) as monto_nota_credito,
                MAX(cc.razon_social) AS cliente_denominacion,
                'BOLETA' AS tipo_comprobante
                FROM cobranza_nota cn
                LEFT JOIN cobranza_pago cp
                    ON cp.id_registro=cn.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_deuda cd
                    ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
                LEFT JOIN comprobante_venta_boletaventa cvb
                    ON cd.id_registro=cvb.id
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=cn.moneda_id
                WHERE cvb.sociedad_id='%s'
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                ORDER BY 4) ; ''' % (DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
            query_info = Nota.objects.raw(sql_cobranza_nota)

            info_cobranza_nota = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.monto_nota_credito)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.tipo_comprobante)
                info_cobranza_nota.append(lista_datos)

            dict_cobranza_nota = {}
            for fila in info_cobranza_nota:
                dict_cobranza_nota[fila[0]+'|'+fila[3]] = fila[1]

            # facturas por jarrones -- mejorar esto..
            DICT_FACT_INVALIDAS = {}

            list_fact_invalidas_mpl = [
                'F001-000231',
                'F001-000233',
                'F001-000245',
                'F001-000298',
                ]
            list_fact_invalidas_mca = [
                'F001-002098',
                'F001-002099',
                ]
            DICT_FACT_INVALIDAS['2'] = list_fact_invalidas_mpl
            DICT_FACT_INVALIDAS['1'] = list_fact_invalidas_mca

            sql = ''' (SELECT
                MAX(cvf.id) AS id,
                to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'FACTURA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cvf.total) AS monto_facturado,
                SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_amortizado,
                MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_pendiente,
                (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) AS estado_cobranza,
                to_char(MAX(cvf.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
                MAX(cvf.fecha_vencimiento) - MAX(cvf.fecha_emision) AS dias_credito,
                MAX(cvf.fecha_vencimiento) AS dias_vencimiento,
                '' AS observaciones,
                '' AS letras,
                '' AS productos
                FROM comprobante_venta_facturaventa cvf
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                LEFT JOIN cobranza_redondeo cr
                    ON cr.deuda_id=cd.id
                WHERE cvf.sociedad_id='%s' AND cvf.estado='4' AND cd.id IS NOT NULL
                GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
                HAVING (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) = 'PENDIENTE' AND CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) NOT IN %s
                ORDER BY cliente_denominacion ASC, nro_comprobante ASC)
                UNION
                (SELECT
                MAX(cvb.id) AS id,
                to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
                'BOLETA' AS tipo_comprobante,
                CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
                MAX(cc.razon_social) AS cliente_denominacion,
                MAX(cvb.total) AS monto_facturado,
                SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_amortizado,
                MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_pendiente,
                (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) AS estado_cobranza,
                to_char(MAX(cvb.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
                MAX(cvb.fecha_vencimiento) - MAX(cvb.fecha_emision) AS dias_credito,
                MAX(cvb.fecha_vencimiento) AS dias_vencimiento,
                '' AS observaciones,
                '' AS letras,
                '' AS productos
                FROM comprobante_venta_boletaventa cvb
                LEFT JOIN datos_globales_seriescomprobante dgsc
                    ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN cobranza_deuda cd
                    ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
                LEFT JOIN cobranza_pago cp
                    ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
                LEFT JOIN cobranza_ingreso ci
                    ON ci.id=cp.id_registro
                LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                    ON dgcb.id=ci.cuenta_bancaria_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                LEFT JOIN cobranza_redondeo cr
                    ON cr.deuda_id=cd.id
                WHERE cvb.sociedad_id='%s' AND cvb.estado='4' AND cd.id IS NOT NULL
                GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
                HAVING (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                    THEN (
                        'CANCELADO'
                    ) ELSE (
                        'PENDIENTE'
                    ) END) = 'PENDIENTE'
                ORDER BY cliente_denominacion ASC, nro_comprobante ASC)
                ORDER BY cliente_denominacion ASC, nro_comprobante ASC ''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad, tuple(DICT_FACT_INVALIDAS[global_sociedad]), DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad)
            query_info = FacturaVenta.objects.raw(sql)

            info_general = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.fecha_emision_comprobante)
                lista_datos.append(fila.tipo_comprobante)
                lista_datos.append(fila.nro_comprobante)
                lista_datos.append(fila.cliente_denominacion)
                lista_datos.append(fila.monto_facturado)
                lista_datos.append(fila.monto_amortizado)
                lista_datos.append(fila.monto_pendiente)
                lista_datos.append(fila.estado_cobranza)
                lista_datos.append(fila.fecha_vencimiento_comprobante)
                lista_datos.append(fila.dias_credito)
                lista_datos.append(str(fila.dias_vencimiento))
                lista_datos.append(fila.observaciones)
                lista_datos.append(fila.letras)
                lista_datos.append(fila.productos)
                info_general.append(lista_datos)

            dias = 0
            for fila in info_general:
                fila[4] = float(fila[4])
                if fila[5] == None:
                    fila[5] = '0.00'
                if fila[6] == None:
                    fila[6] = fila[4]
                fila[5] = float(fila[5])
                fila[6] = float(fila[6])
                fila[9] = float(fila[9])
                if fila[2] + '|' + fila[1] in dict_cobranza_nota:
                    fila[5] += float(dict_cobranza_nota[fila[2] + '|' + fila[1]])
                    fila[6] = fila[4] - fila[5]

                if fila[10] != '':
                    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
                    fecha1 = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                    fecha2 = datetime.strptime(fila[10], '%Y-%m-%d')
                    dias = (fecha1 - fecha2) / timedelta(days=1)
                    fila[10] = str(dias)
                if float(dias) > float(0):
                    fila[7] = 'VENCIDO'

                if fila[0]+'|'+fila[1]+'|'+fila[2] in dict_letras:
                    fila[12] = dict_letras[fila[0]+'|'+fila[1]+'|'+fila[2]]
                    div = fila[12].split('\n')
                    rest = float(fila[5])
                    list_resumen_letra = []
                    try:
                        for sub_div in div:
                            list_fecha_monto = sub_div.split(' $ ')
                            fecha_letra = list_fecha_monto[0]
                            monto_letra = float(list_fecha_monto[1])
                            rest = rest - monto_letra
                            if rest >= float(0):
                                estado_letra = "CANCELADO"
                            else:
                                fecha_base = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                                fecha_letra_dt = datetime.strptime(fecha_letra, '%d/%m/%Y')
                                dias = (fecha_base - fecha_letra_dt) / timedelta(days=1)
                                if float(dias) > float(0):
                                    estado_letra = 'VENCIDO'
                                else:
                                    estado_letra = "PENDIENTE"
                            fila_letra = fecha_letra + ' $ ' + str(monto_letra) + ' ' + estado_letra
                            list_resumen_letra.append(fila_letra)
                        fila[12] = '\n'.join(list_resumen_letra)
                    except Exception as e:
                        print(e)
                else:
                    'ERROR AL EXTRAER LAS CUOTAS'
                fila[13] = dict_productos[fila[0]+'|'+fila[1]+'|'+fila[2]]


            info = [] # nueva lista de listas filtrada con pagos por notas de credito:
            for fila_general in info_general:
                if round(fila_general[6],3) > float(0):
                    info.append(fila_general)

            num_fila = 0
            total_deuda_cliente = 0
            for fila in info:
                if num_fila != 0:
                    if fila[3] != info[num_fila-1][3]:
                        hoja.append(('','','','Total:',float(total_deuda_cliente)))
                        # cell_total = hoja.cell(hoja.max_row, 4)
                        # cell_total.alignment = self.alineacion_derecha
                        # cell_total.number_format = self.formato_dolar
                        hoja.append(('',))
                        hoja.append((fila[3],))
                        hoja.append(('FECHA','N° COMPROBANTE','FACTURADO (US$)','AMORTIZADO','PENDIENTE','ESTADO','FECHA DE VENC.','CRÉDITO (DÍAS)','DÍAS DE VENC.','OBSERVACIONES','LETRAS','PRODUCTOS'))
                        total_deuda_cliente = 0
                        col_range = hoja.max_column
                        nueva_fila = hoja.max_row
                        for col in range(1, col_range + 1):
                            cell_header = hoja.cell(nueva_fila, col)
                            cell_header.fill = color_relleno
                            cell_header.font = NEGRITA
                else:
                    hoja.append((fila[3],))
                    hoja.append(('FECHA','N° COMPROBANTE','FACTURADO (US$)','AMORTIZADO','PENDIENTE','ESTADO','FECHA DE VENC.','CRÉDITO (DÍAS)','DÍAS DE VENC.','OBSERVACIONES','LETRAS','PRODUCTOS'))
                    col_range = hoja.max_column
                    nueva_fila = hoja.max_row
                    for col in range(1, col_range + 1):
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = color_relleno
                        cell_header.font = NEGRITA

                c = 0
                list_temp = []
                for dato in fila:
                    if c != 1 and c != 3: # tipo_comprobante, razon_social (datos omitidos en la tabla)
                        list_temp.append(dato)
                    c+=1

                hoja.append(list_temp)
                total_deuda_cliente += list_temp[4]
                num_fila += 1

                for i in range(hoja.max_row):
                    if i >= nueva_fila-1:
                        row = list(hoja.rows)[i]
                        for col in range(hoja.max_column):
                            row[col].border = BORDE_DELGADO
                            if 2 <= col <= 4:
                                row[col].alignment = ALINEACION_DERECHA
                                row[col].number_format = FORMATO_DOLAR
                            if col == 7 or col == 8:
                                row[col].alignment = ALINEACION_DERECHA
                            if 10 <= col <= 11:
                                row[col].alignment = AJUSTAR_TEXTO

                if fila == info[-1]:
                    hoja.append(('','','','Total:',float(total_deuda_cliente)))
                    # cell_total = hoja.cell(hoja.max_row, 4)
                    # cell_total.alignment = self.alineacion_derecha
                    # cell_total.number_format = self.formato_dolar

            ajustarColumnasSheet(hoja)
            return wb
            # wb.save('reporte4_prueba.xlsx')


        # reporte_facturas_pendientes('2000')
        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        wb=reporte_facturas_pendientes()
        nombre_archivo = "Reporte Facturas Pendientes - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteDepositosCuentasBancarias(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def reporte_depositos_cuentas():
            # self.fecha_inicio = '2021-12-20'
            # self.fecha_fin = '2022-01-31'

            wb = Workbook()
            hoja = wb.active

            color_relleno = rellenoSociedad(global_sociedad)

            sql_cuentas = ''' SELECT
                dgcb.id,
                dgb.razon_social AS nombre_banco,
                ss.razon_social AS nombre_titular,
                dgcb.numero_cuenta AS cuenta_banco,
                dgcb.numero_cuenta_interbancaria AS cuenta_cci_banco,
                dgm.nombre AS moneda_descripcion
                FROM datos_globales_cuentabancariasociedad dgcb
                LEFT JOIN datos_globales_banco dgb
                    ON dgb.id=dgcb.banco_id
                LEFT JOIN datos_globales_moneda dgm
                    ON dgm.id=dgcb.moneda_id
                LEFT JOIN sociedad_sociedad ss
                    ON ss.id=dgcb.sociedad_id
                WHERE dgcb.sociedad_id='%s' AND dgcb.estado='1' AND dgcb.efectivo=False ; ''' %(global_sociedad)
            query_info = CuentaBancariaSociedad.objects.raw(sql_cuentas)

            info_cuentas = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.nombre_banco)
                lista_datos.append(fila.nombre_titular)
                lista_datos.append(fila.cuenta_banco)
                lista_datos.append(fila.cuenta_cci_banco)
                lista_datos.append(fila.moneda_descripcion)
                info_cuentas.append(lista_datos)

            list_temp_hojas = []
            dict_totales_cuentas = {}
            count_cuenta = 0
            list_nro_cuentas = []
            dict_nro_cuentas = {}
            for fila in info_cuentas:
                nro_cuenta = fila[2]
                dict_nro_cuentas[nro_cuenta] = fila[0] + ' ' +fila[4]
                list_nro_cuentas.append(nro_cuenta)

                sql = ''' (SELECT
                    MAX(cp.id) AS id,
                    ci.fecha AS fecha_orden,
                    to_char(ci.fecha, 'DD/MM/YYYY') AS fecha_operacion_bancaria,
                    ci.numero_operacion AS numero_operacion_bancaria,
                    (CASE WHEN MAX(dgm.abreviatura)='USD' THEN MAX(ci.monto) ELSE ROUND(MAX(ci.monto)/MAX(dgtcs.tipo_cambio_venta),2) END) AS monto_dolares,
                    (CASE WHEN MAX(dgm.abreviatura)='USD' THEN ROUND(MAX(ci.monto)*MAX(dgtcs.tipo_cambio_venta),2) ELSE MAX(ci.monto) END) AS monto_soles,
                    STRING_AGG(cc.razon_social, '\n') AS empresas,
                    STRING_AGG(CONCAT('FACTURA: ', dgsc.serie, '-', cvf.numero_factura), '\n') AS documentos,
                    STRING_AGG(to_char(cvf.fecha_emision, 'DD/MM/YYYY'), '\n') AS fecha_documentos,
                    STRING_AGG(CAST((CASE WHEN dgm.abreviatura='USD' THEN cp.monto ELSE ROUND(cp.monto/dgtcs.tipo_cambio_venta,2) END) AS TEXT), '\n') AS pago_dolares,
                    STRING_AGG(CAST((CASE WHEN dgm.abreviatura='USD' THEN ROUND(cp.monto*dgtcs.tipo_cambio_venta,2) ELSE cp.monto END) AS TEXT), '\n') AS pago_soles
                    FROM cobranza_pago cp
                    LEFT JOIN cobranza_ingreso ci
                        ON ci.id=cp.id_registro AND cp.content_type_id='%s'
                    LEFT JOIN datos_globales_tipocambiosunat dgtcs
                        ON ci.fecha=dgtcs.fecha
                    LEFT JOIN datos_globales_cuentabancariasociedad dgcbs
                        ON dgcbs.id=ci.cuenta_bancaria_id AND dgcbs.estado='1' AND dgcbs.efectivo=False
                    LEFT JOIN datos_globales_moneda dgm
                        ON dgm.id=dgcbs.moneda_id
                    LEFT JOIN cobranza_deuda cd
                        ON cd.id=cp.deuda_id
                    LEFT JOIN comprobante_venta_facturaventa cvf
                        ON cvf.id=cd.id_registro AND cd.content_type_id='%s' AND cvf.estado='4'
                    LEFT JOIN datos_globales_seriescomprobante dgsc
                        ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
                    LEFT JOIN datos_globales_banco dgb
                        ON dgb.id = dgcbs.banco_id
                    LEFT JOIN clientes_cliente cc
                        ON cc.id=cvf.cliente_id
                    WHERE dgcbs.numero_cuenta='%s' AND '%s' <= ci.fecha AND ci.fecha <= '%s' AND cvf.id IS NOT NULL
                    GROUP BY ci.numero_operacion, dgcbs.banco_id, ci.cuenta_bancaria_id, ci.fecha
                    ORDER BY ci.fecha ASC)
                    UNION
                    (SELECT
                    MAX(cp.id) AS id,
                    ci.fecha AS fecha_orden,
                    to_char(ci.fecha, 'DD/MM/YYYY') AS fecha_operacion_bancaria,
                    ci.numero_operacion AS numero_operacion_bancaria,
                    (CASE WHEN MAX(dgm.abreviatura)='USD' THEN MAX(ci.monto) ELSE ROUND(MAX(ci.monto)/MAX(dgtcs.tipo_cambio_venta),2) END) AS monto_dolares,
                    (CASE WHEN MAX(dgm.abreviatura)='USD' THEN ROUND(MAX(ci.monto)*MAX(dgtcs.tipo_cambio_venta),2) ELSE MAX(ci.monto) END) AS monto_soles,
                    STRING_AGG(cc.razon_social, '\n') AS empresas,
                    STRING_AGG(CONCAT('BOLETA: ', dgsc.serie, '-', cvb.numero_boleta), '\n') AS documentos,
                    STRING_AGG(to_char(cvb.fecha_emision, 'DD/MM/YYYY'), '\n') AS fecha_documentos,
                    STRING_AGG(CAST((CASE WHEN dgm.abreviatura='USD' THEN cp.monto ELSE ROUND(cp.monto/dgtcs.tipo_cambio_venta,2) END) AS TEXT), '\n') AS pago_dolares,
                    STRING_AGG(CAST((CASE WHEN dgm.abreviatura='USD' THEN ROUND(cp.monto*dgtcs.tipo_cambio_venta,2) ELSE cp.monto END) AS TEXT), '\n') AS pago_soles
                    FROM cobranza_pago cp
                    LEFT JOIN cobranza_ingreso ci
                        ON ci.id=cp.id_registro AND cp.content_type_id='%s'
                    LEFT JOIN datos_globales_tipocambiosunat dgtcs
                        ON ci.fecha=dgtcs.fecha
                    LEFT JOIN datos_globales_cuentabancariasociedad dgcbs
                        ON dgcbs.id=ci.cuenta_bancaria_id AND dgcbs.estado='1' AND dgcbs.efectivo=False
                    LEFT JOIN datos_globales_moneda dgm
                        ON dgm.id=dgcbs.moneda_id
                    LEFT JOIN cobranza_deuda cd
                        ON cd.id=cp.deuda_id
                    LEFT JOIN comprobante_venta_boletaventa cvb
                        ON cvb.id=cd.id_registro AND cd.content_type_id='%s' AND cvb.estado='4'
                    LEFT JOIN datos_globales_seriescomprobante dgsc
                        ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
                    LEFT JOIN datos_globales_banco dgb
                        ON dgb.id = dgcbs.banco_id
                    LEFT JOIN clientes_cliente cc
                        ON cc.id=cvb.cliente_id
                    WHERE dgcbs.numero_cuenta='%s' AND '%s' <= ci.fecha AND ci.fecha <= '%s' AND cvb.id IS NOT NULL
                    GROUP BY ci.numero_operacion, dgcbs.banco_id, ci.cuenta_bancaria_id, ci.fecha
                    ORDER BY ci.fecha ASC )
                    ORDER BY 2 ; ''' %(DICT_CONTENT_TYPE['cobranza | ingreso'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], nro_cuenta, global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['cobranza | ingreso'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], nro_cuenta, global_fecha_inicio, global_fecha_fin)
                query_info = Pago.objects.raw(sql)

                info_depositos = []
                for dato_fila in query_info:
                    lista_datos = []
                    lista_datos.append(dato_fila.fecha_operacion_bancaria)
                    lista_datos.append(dato_fila.numero_operacion_bancaria)
                    lista_datos.append(dato_fila.monto_dolares)
                    lista_datos.append(dato_fila.monto_soles)
                    lista_datos.append(dato_fila.empresas)
                    lista_datos.append(dato_fila.documentos)
                    lista_datos.append(dato_fila.fecha_documentos)
                    lista_datos.append(dato_fila.pago_dolares)
                    lista_datos.append(dato_fila.pago_soles)
                    info_depositos.append(lista_datos)

                count_cuenta += 1
                count_mes = 0
                list_general = []
                list_mes = []
                for deposito in info_depositos:
                    if count_mes != 0:
                        año_mes_actual = deposito[0][6:] + deposito[0][3:5]
                        año_mes_anterior = info_depositos[count_mes-1][0][6:] + info_depositos[count_mes-1][0][3:5]
                        if año_mes_actual != año_mes_anterior:
                            list_general.append(list_mes)
                            list_mes = []
                    list_mes.append(deposito)
                    try:
                        deposito[2] = float(deposito[2])
                        deposito[3] = float(deposito[3])
                    except:
                        ""
                    count_mes += 1
                if list_mes != []:
                    list_general.append(list_mes)

    # ******************************************************************************************************************************
                count = 0
                for list_mes_deposito in list_general:
                    mes = list_mes_deposito[0][0][3:5]
                    año = list_mes_deposito[0][0][6:]
                    name_sheet = DICT_MESES[str(mes)] + ' - ' + str(año)
                    # print('*************************************')
                    # print(str(mes), name_sheet, list_temp_hojas)
                    # print('*************************************')
                    if count != 0:
                        if name_sheet not in list_temp_hojas:
                            hoja = wb.create_sheet(name_sheet)
                            list_temp_hojas.append(name_sheet)
                        else:
                            hoja = wb[name_sheet]
                    else:
                        hoja = wb.active
                        hoja.title = name_sheet
                        if name_sheet not in list_temp_hojas:
                            list_temp_hojas.append(name_sheet)
                        count += 1
                        # count_cuenta += 1

                    hoja.append(('', ''))
                    hoja.append(('', ''))
                    hoja.append(('BANCO:', fila[0]))
                    hoja.append(('EMPRESA:', fila[1]))
                    hoja.append(('CUENTA:', fila[2]))
                    hoja.append(('CCI:', fila[3]))
                    hoja.append(('MONEDA:', fila[4]))
                    # wb.active = hoja
                    hoja.append(('FECHA', 'REFERENCIA', 'MONTO (US$)', 'MONTO (S/)', 'EMPRESAS', 'DOCUMENTOS', 'FECHA DOCUMENTOS', 'PAGOS (US$)', 'PAGOS (S/)'))
                    col_range = hoja.max_column
                    nueva_fila = hoja.max_row

                    for col in range(1, col_range + 1):
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = color_relleno
                        cell_header.font = NEGRITA
                        for count_fila in range(1,6):
                            cell_header = hoja.cell(nueva_fila-count_fila, col)
                            cell_header.fill = color_relleno

                    total_mes_cuenta_dolares = 0
                    total_mes_cuenta_soles = 0
                    for fila_deposito in list_mes_deposito:
                        if fila_deposito[4] != "" and fila_deposito[4] != None:
                            if 'Nota Credito:' not in fila_deposito[1]:
                                total_mes_cuenta_dolares += float(fila_deposito[2])
                                try:
                                    total_mes_cuenta_soles += float(fila_deposito[3])
                                except:
                                    pass
                        hoja.append(fila_deposito)
                    cuenta_banco_ingreso = fila[0] + ' ' + fila[4]
                    dict_totales_cuentas[cuenta_banco_ingreso + '|' + name_sheet] = str(round(total_mes_cuenta_soles,2)) + '|' + str(round(total_mes_cuenta_dolares,2))

                    for i in range(hoja.max_row):
                        if i >= nueva_fila-1:
                            row = list(hoja.rows)[i]
                            for col in range(hoja.max_column):
                                row[col].border = BORDE_DELGADO
                                if col == 2:
                                    row[col].alignment = ALINEACION_DERECHA
                                    row[col].number_format = FORMATO_DOLAR
                                if col == 3:
                                    row[col].alignment = ALINEACION_DERECHA
                                    row[col].number_format = FORMATO_SOLES
                                if col == 7 or col == 8:
                                    row[col].alignment = ALINEACION_DERECHA
                                if 4 <= col <= 8:
                                    row[col].alignment = AJUSTAR_TEXTO

                    ajustarColumnasSheet(hoja)

    # ******************************************************************************************************************************
            # print(dict_totales_cuentas)
            for k,v in dict_totales_cuentas.items():
                print(k, v)

            ''' BCP DOLARES|ENERO - 2022 107347.17|27469.59
            BCP DOLARES|FEBRERO - 2022 62893.94|16654.11
            BCP SOLES|ENERO - 2022 67272.71|17324.28
            BCP SOLES|FEBRERO - 2022 37920.94|10013.3
            BBVA SOLES|ENERO - 2022 2230.6|579.12
            BBVA SOLES|FEBRERO - 2022 3884.22|1016.74
            BBVA DOLARES|ENERO - 2022 140658.24|36246.8
            BBVA DOLARES|FEBRERO - 2022 69518.89|18147.0 '''
            print()

            hoja = wb.create_sheet('Resumen Ingresos')
            # hoja.append(('', ''))
            hoja.append(('ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
            hoja.append(('', ''))
            if global_sociedad == '2':
                hoja.append(('SOLES', '', '', '', '', 'DOLARES'))
                hoja.merge_cells(start_row = 3, start_column = 1, end_row = 3, end_column = 4)
                hoja.merge_cells(start_row = 3, start_column = 6, end_row = 3, end_column = 9)
                celda_multiple = hoja['A3']
                celda_multiple.alignment = ALINEACION_CENTRO
                celda_multiple = hoja['F3']
                celda_multiple.alignment = ALINEACION_CENTRO
            if global_sociedad == '1':
                hoja.append(('SOLES', '', '', '', '', '', '', 'DOLARES'))
                hoja.merge_cells(start_row = 3, start_column = 1, end_row = 3, end_column = 6)
                hoja.merge_cells(start_row = 3, start_column = 8, end_row = 3, end_column = 13)
                celda_multiple = hoja['A3']
                celda_multiple.alignment = ALINEACION_CENTRO
                celda_multiple = hoja['H3']
                celda_multiple.alignment = ALINEACION_CENTRO

            for i in range(len(list_nro_cuentas)):
                list_nro_cuentas[i] = dict_nro_cuentas[list_nro_cuentas[i]]
            list_encabezado = [''] + list_nro_cuentas + ['TOTAL','',''] + list_nro_cuentas + ['TOTAL']
            hoja.append(tuple(list_encabezado))

            col_range = hoja.max_column
            nueva_fila = hoja.max_row
            for col in range(1, col_range + 1):
                if global_sociedad == '2':
                    if col != 5:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = color_relleno
                        cell_header.font = NEGRITA
                        for count_fila in range(1,2):
                            cell_header = hoja.cell(nueva_fila-count_fila, col)
                            cell_header.fill = color_relleno
                if global_sociedad == '1':
                    if col != 7:
                        cell_header = hoja.cell(nueva_fila, col)
                        cell_header.fill = color_relleno
                        cell_header.font = NEGRITA
                        for count_fila in range(1,2):
                            cell_header = hoja.cell(nueva_fila-count_fila, col)
                            cell_header.fill = color_relleno

            def insertarResumenDataAnterior(hoja, data):
                for fila in data:
                    hoja.append(fila)
                # return hoja

            # rpta = mensajeDialogo('pregunta','Resumen Ingresos','¿Desea Agregar la Información en el Resumen de los años anteriores?')
            # if rpta == 'Yes':
            if global_fecha_inicio <= '2022-01-01':
                if global_sociedad == '2':
                    insertarResumenDataAnterior(hoja, list_resumen_ingresos_sis_anterior_mpl)
                if global_sociedad == '1':
                    insertarResumenDataAnterior(hoja, list_resumen_ingresos_sis_anterior_mca)

            for mes in list_temp_hojas:
                list_temp_fila = []
                # list_temp.append(mes)
                list_temp_soles = []
                list_temp_dolares = []
                total_soles = 0
                total_dolares = 0
                for k,value in dict_totales_cuentas.items():
                    if mes in k:
                        # print(k)
                        monto_soles = float(value[:value.find("|")])
                        monto_dolares = float(value[value.find("|")+1:])
                        total_soles += monto_soles
                        total_dolares += monto_dolares
                        list_temp_soles.append(monto_soles)
                        list_temp_dolares.append(monto_dolares)
                if global_sociedad == '2':
                    if len(list_temp_soles) < 2:
                        if len(list_temp_soles) == 1:
                            list_temp_soles.extend([''])
                            list_temp_dolares.extend([''])
                if global_sociedad == '1':
                    if len(list_temp_soles) < 4:
                        if len(list_temp_soles) == 1:
                            list_temp_soles.extend(['','',''])
                            list_temp_dolares.extend(['','',''])
                        elif len(list_temp_soles) == 2:
                            list_temp_soles.extend(['',''])
                            list_temp_dolares.extend(['',''])
                        elif len(list_temp_soles) == 3:
                            list_temp_soles.extend([''])
                            list_temp_dolares.extend([''])
                list_temp_soles.append(total_soles)
                list_temp_dolares.append(total_dolares)
                list_temp_fila = [mes] + list_temp_soles + ['', mes] + list_temp_dolares
                hoja.append(tuple(list_temp_fila))

            i = 0
            for row in hoja.rows:
                if i >= 2:
                    for col in range(hoja.max_column):
                        if global_sociedad == '2':
                            if col != 4:
                                row[col].border = BORDE_DELGADO
                            if 1 <= col <= 3:
                                row[col].number_format = FORMATO_SOLES
                            if col >= 6:
                                row[col].number_format = FORMATO_DOLAR
                        if global_sociedad == '1':
                            if col != 6:
                                row[col].border = BORDE_DELGADO
                            if 1 <= col <= 5:
                                row[col].number_format = FORMATO_SOLES
                            if col >= 8:
                                row[col].number_format = FORMATO_DOLAR
                i += 1
            ajustarColumnasSheet(hoja)

            def extraer_resumen_bloc_de_notas(hoja):

                if global_sociedad == '2':
                    nro_col = 9
                    file_mpl = open('resumen_ingresos_mpl.txt', "w")
                    file = file_mpl
                if global_sociedad == '1':
                    nro_col = 13
                    file_mca = open('resumen_ingresos_mca.txt', "w")
                    file = file_mca

                list_temp = []
                for i in range(5, hoja.max_row + 1): # Nro de filas del excel
                    celda = hoja.cell(row = i, column = nro_col)
                    if celda.value == None:
                        valor_celda = ''
                    else:
                        valor_celda = celda.value
                    list_temp.append(valor_celda)
                # print(list_temp)

                for pos in range(len(list_temp)):
                    list_temp[pos] = str(round(list_temp[pos],2))

                info = '\n'.join(list_temp)
                file.write(info)
                file.close()

            extraer_resumen_bloc_de_notas(hoja)

            def grafico_resumen_ingresos(ws):
                max_fila = ws.max_row

                chart = LineChart()
                # print(help(chart))
                chart.height = 15 # default is 7.5
                chart.width = 30 # default is 15
                chart.y_axis.title = 'INGRESOS'
                chart.x_axis.title = 'MESES'
                chart.legend.position = 'b' #bottom
                # chart.style = 12

                count = 1
                fila_base = 5
                year_base = 2018
                if global_sociedad == '2':
                    chart.title = 'RESUMEN DE INGRESOS - MULTIPLAY'
                    data_col = 9 # montos en dolares
                    # data_col = 4 # montos en soles
                if global_sociedad == '1':
                    chart.title = 'RESUMEN DE INGRESOS - MULTICABLE'
                    data_col = 13 # montos en dolares
                    # data_col = 6 # montos en soles
                while max_fila >= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    series = Series(values, title = "Ingresos del " + str(year_base + count))
                    chart.append(series)
                    max_fila -= 12
                    count += 1
                if 1 <= max_fila <= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    series = Series(values, title = "Ingresos del " + str(year_base + count))
                    chart.append(series)

                meses = Reference(ws, min_col = 1, min_row = 1, max_col = 12, max_row = 1)
                chart.set_categories(meses)
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                chart.dataLabels.dLblPos = 't' # top

                chart.plot_area.dTable = DataTable()
                chart.plot_area.dTable.showHorzBorder = True
                chart.plot_area.dTable.showVertBorder = True
                chart.plot_area.dTable.showOutline = True
                chart.plot_area.dTable.showKeys = True

                ws.add_chart(chart)

            grafico_resumen_ingresos(hoja)

            return wb
            # wb.save('reporte3_prueba.xlsx')

        def reporte_depositos_cuentas_resumen():
            file_mpl = open('resumen_ingresos_mpl.txt', "r")
            file_mca = open('resumen_ingresos_mca.txt', "r")
            contenido_mpl = file_mpl.read()
            contenido_mca = file_mca.read()
            file_mpl.close()
            file_mca.close()
            list_mpl = contenido_mpl.split("\n")
            list_mca = contenido_mca.split("\n")
            for pos in range(len(list_mpl)):
                list_mpl[pos] = float(list_mpl[pos])
            for pos in range(len(list_mca)):
                list_mca[pos] = float(list_mca[pos])
            # list_ambas_empresas = np.add(list_mpl, list_mca)
            list_ambas_empresas = [x + y for x, y in zip(list_mpl, list_mca)]

            wb = Workbook()
            hoja = wb.active
            hoja.append(('','ENERO', 'FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SETIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'))
            for dato in list_ambas_empresas:
                hoja.append((dato,''))

            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if col == 0:
                        row[col].number_format = FORMATO_DOLAR

            def grafico_resumen_ingresos(ws):
                max_fila = ws.max_row

                chart = LineChart()
                # print(help(chart))
                chart.height = 15 # default is 7.5
                chart.width = 30 # default is 15
                chart.y_axis.title = 'INGRESOS'
                chart.x_axis.title = 'MESES'
                chart.legend.position = 'b' #bottom
                chart.title = 'RESUMEN DE INGRESOS - MULTICABLE + MULTIPLAY'
                # chart.style = 12

                count = 1
                fila_base = 2
                year_base = 2018
                data_col = 1 # montos en dolares
                while max_fila >= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    series = Series(values, title = "Ingresos del " + str(year_base + count))
                    chart.append(series)
                    max_fila -= 12
                    count += 1
                if 1 <= max_fila <= 12:
                    values = Reference(ws, min_col = data_col, min_row = fila_base + 12*(count-1), max_col = data_col, max_row = 12*count + fila_base - 1)
                    series = Series(values, title = "Ingresos del " + str(year_base + count))
                    chart.append(series)

                meses = Reference(ws, min_col = 2, min_row = 1, max_col = 13, max_row = 1)
                chart.set_categories(meses)
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                chart.dataLabels.dLblPos = 't' # top

                chart.plot_area.dTable = DataTable()
                chart.plot_area.dTable.showHorzBorder = True
                chart.plot_area.dTable.showVertBorder = True
                chart.plot_area.dTable.showOutline = True
                chart.plot_area.dTable.showKeys = True

                ws.add_chart(chart)

            grafico_resumen_ingresos(hoja)

            return wb
        

        if global_sociedad:
            wb=reporte_depositos_cuentas()
            query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
            abreviatura = query_sociedad.abreviatura
        else:
            wb=reporte_depositos_cuentas_resumen()
            abreviatura = "MPL-MCA"
        nombre_archivo = "Reporte Depositos Cuentas - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteClientesProductos(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def consulta_resumen_cliente():
            sql = ''' (SELECT
                MAX(cvfd.id) AS id,
                cvf.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv) - SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvfd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvfd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvfd.factura_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvf.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvfd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_facturaventadetalle cvfd
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN material_material mm
                    ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
                WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s' AND cvf.estado='4'
                GROUP BY cvf.cliente_id, cvfd.content_type_id, cvfd.id_registro
                ORDER BY 3, 5)
                UNION
                (SELECT
                MAX(cvbd.id) AS id,
                cvb.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv) - SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvbd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvbd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvbd.boleta_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvb.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvbd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_boletaventadetalle cvbd
                LEFT JOIN comprobante_venta_boletaventa cvb
                    ON cvb.id=cvbd.boleta_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN material_material mm
                    ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
                WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s' AND cvb.estado='4'
                GROUP BY cvb.cliente_id, cvbd.content_type_id, cvbd.id_registro
                ORDER BY 3, 5)
                ORDER BY 3, 5 ; ''' %(DICT_CONTENT_TYPE['material | material'], global_sociedad, global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['material | material'], global_sociedad, global_fecha_inicio, global_fecha_fin)
            query_info = FacturaVentaDetalle.objects.raw(sql)
            
            info = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.codigo_cliente)
                lista_datos.append(dato_fila.cliente_denominacion)
                lista_datos.append(dato_fila.materiales)
                lista_datos.append(dato_fila.texto_material)
                lista_datos.append(dato_fila.precio_cotizado_sin_igv)
                lista_datos.append(dato_fila.igv_cotizado)
                lista_datos.append(dato_fila.precio_cotizado)
                lista_datos.append(dato_fila.cantidades)
                lista_datos.append(dato_fila.total_cantidades)
                lista_datos.append(dato_fila.nro_compras)
                lista_datos.append(dato_fila.nro_compras_en_oferta)
                lista_datos.append(dato_fila.fecha_registro)
                info.append(lista_datos)

            return info

        def consulta_resumen_producto():
            sql = ''' (SELECT
                MAX(cvfd.id) AS id,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                cvf.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv) - SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvfd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvfd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvfd.factura_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvf.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvfd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_facturaventadetalle cvfd
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN material_material mm
                    ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
                WHERE cvf.sociedad_id='%s' AND '%s' <= cvf.fecha_emision AND cvf.fecha_emision <= '%s' AND cvf.estado='4'
                GROUP BY cvfd.content_type_id, cvfd.id_registro, cvf.cliente_id
                ORDER BY 3, 5)
                UNION
                (SELECT
                MAX(cvbd.id) AS id,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                cvb.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv) - SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvbd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvbd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvbd.boleta_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvb.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvbd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_boletaventadetalle cvbd
                LEFT JOIN comprobante_venta_boletaventa cvb
                    ON cvb.id=cvbd.boleta_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN material_material mm
                    ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
                WHERE cvb.sociedad_id='%s' AND '%s' <= cvb.fecha_emision AND cvb.fecha_emision <= '%s' AND cvb.estado='4'
                GROUP BY cvbd.content_type_id, cvbd.id_registro, cvb.cliente_id
                ORDER BY 3, 5)
                ORDER BY 3, 5 ; ''' %(DICT_CONTENT_TYPE['material | material'], global_sociedad, global_fecha_inicio, global_fecha_fin, DICT_CONTENT_TYPE['material | material'], global_sociedad, global_fecha_inicio, global_fecha_fin)
            query_info = FacturaVentaDetalle.objects.raw(sql)
            
            info = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.materiales)
                lista_datos.append(dato_fila.texto_material)
                lista_datos.append(dato_fila.codigo_cliente)
                lista_datos.append(dato_fila.cliente_denominacion)
                lista_datos.append(dato_fila.precio_cotizado_sin_igv)
                lista_datos.append(dato_fila.igv_cotizado)
                lista_datos.append(dato_fila.precio_cotizado)
                lista_datos.append(dato_fila.cantidades)
                lista_datos.append(dato_fila.total_cantidades)
                lista_datos.append(dato_fila.nro_compras)
                lista_datos.append(dato_fila.nro_compras_en_oferta)
                lista_datos.append(dato_fila.fecha_registro)
                info.append(lista_datos)

            return info

        def procesar_consulta_resumen(info, pos=0):
            count = 0
            list_general = []
            list_temp = []
            for fila in info:
                fila[4] = float(fila[4])
                fila[5] = float(fila[5])
                fila[6] = float(fila[6])
                fila[8] = float(fila[8])
                fila[9] = float(fila[9])
                fila[10] = float(fila[10])
                if count != 0:
                    if fila[pos] != info[count-1][pos]:
                        list_general.append(list_temp)
                        list_temp = []
                list_temp.append(fila)
                count += 1
            list_general.append(list_temp)
            # print(list_general)
            return list_general


        def consulta_general_cliente_productos():
            sql = ''' (SELECT
                MAX(cvfd.id) AS id,
                SUBSTRING(to_char(cvfd.created_at, 'YYYY-MM-DD'),1,7) AS fecha_orden,
                cvf.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv) - SUM(cvfd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvfd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvfd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvfd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvfd.factura_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvf.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvfd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_facturaventadetalle cvfd
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvf.cliente_id
                LEFT JOIN material_material mm
                    ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
                WHERE cvf.sociedad_id='%s' AND cvf.estado='4'
                GROUP BY cvf.cliente_id, fecha_orden, cvfd.content_type_id, cvfd.id_registro
                ORDER BY 2, 4, 6)
                UNION
                (SELECT
                MAX(cvbd.id) AS id,
                SUBSTRING(to_char(cvbd.created_at, 'YYYY-MM-DD'),1,7) AS fecha_orden,
                cvb.cliente_id as codigo_cliente,
                MAX(cc.razon_social) AS cliente_denominacion,
                STRING_AGG(DISTINCT CAST(mm.id AS TEXT), ' | ') AS materiales,
                MAX(mm.descripcion_corta) as texto_material,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS precio_cotizado_sin_igv,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv) - SUM(cvbd.precio_final_con_igv/1.18), 2) AS TEXT) AS igv_cotizado,
                CAST(ROUND(SUM(cvbd.precio_final_con_igv), 2) AS TEXT) AS precio_cotizado,
                STRING_AGG(CAST(ROUND(cvbd.cantidad,3) AS TEXT), ' | ') AS cantidades,
                ROUND(SUM(cvbd.cantidad),3) AS total_cantidades,
                COUNT(DISTINCT cvbd.factura_venta_id) AS nro_compras,
                SUM(CASE WHEN (cvb.descuento_global > 0.00) THEN 1 ELSE 0 END) AS nro_compras_en_oferta,
                to_char(MAX(cvbd.created_at), 'YYYY-MM-DD') as fecha_registro
                FROM comprobante_venta_facturaventadetalle cvbd
                LEFT JOIN comprobante_venta_facturaventa cvb
                    ON cvb.id=cvbd.factura_venta_id
                LEFT JOIN clientes_cliente cc
                    ON cc.id=cvb.cliente_id
                LEFT JOIN material_material mm
                    ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
                WHERE cvb.sociedad_id='%s' AND cvb.estado='4'
                GROUP BY cvb.cliente_id, fecha_orden, cvbd.content_type_id, cvbd.id_registro
                ORDER BY 2, 4, 6)
                ORDER BY 2, 4, 6; ''' %(DICT_CONTENT_TYPE['material | material'], global_sociedad, DICT_CONTENT_TYPE['material | material'], global_sociedad)
            query_info = FacturaVentaDetalle.objects.raw(sql)
            
            info = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.codigo_cliente)
                lista_datos.append(dato_fila.cliente_denominacion)
                lista_datos.append(dato_fila.materiales)
                lista_datos.append(dato_fila.texto_material)
                lista_datos.append(dato_fila.precio_cotizado_sin_igv)
                lista_datos.append(dato_fila.igv_cotizado)
                lista_datos.append(dato_fila.precio_cotizado)
                lista_datos.append(dato_fila.cantidades)
                lista_datos.append(dato_fila.total_cantidades)
                lista_datos.append(dato_fila.nro_compras)
                lista_datos.append(dato_fila.nro_compras_en_oferta)
                lista_datos.append(dato_fila.fecha_registro)
                info.append(lista_datos)

            return info

        def procesar_consulta_cliente_productos(info):
            list_general = []
            list_mes = []
            list_cliente = []
            # list_temp = []
            # count = 0
            # for fila in info:
            #     fila[4] = float(fila[4])
            #     fila[5] = float(fila[5])
            #     fila[6] = float(fila[6])
            #     fila[8] = float(fila[8])
            #     fila[9] = float(fila[9])
            #     fila[10] = float(fila[10])
            #     # list_cliente.append(fila)
            #     if count != 0:
            #         if fila[11][:7] != info[count-1][11][:7]: # verifica el mes
            #             # list_mes.append(list_cliente)
            #             list_general.append(list_mes)
            #             list_mes = []
            #             # list_cliente = []
            #         if fila[0] != info[count-1][0]: # verifica el cliente
            #             list_mes.append(list_cliente)
            #             list_cliente = []
            #     # list_mes.append(list_cliente)
            #     list_cliente.append(fila)
            #     count += 1
            # if list_cliente != []:
            #     list_mes.append(list_cliente)
            # if list_mes != []:
            #     list_general.append(list_mes)
            # # print(list_general)

            count = 0
            for fila in info:
                fila[4] = float(fila[4])
                fila[5] = float(fila[5])
                fila[6] = float(fila[6])
                fila[8] = float(fila[8])
                fila[9] = float(fila[9])
                fila[10] = float(fila[10])
                if count != 0:
                    mes_actual = fila[11][:7]
                    mes_anterior = info[count-1][11][:7]
                    cliente_actual = fila[0]
                    cliente_anterior = info[count-1][0]
                    if (mes_actual == mes_anterior) and (cliente_actual != cliente_anterior):
                        list_mes.append(list_cliente)
                        list_cliente = []
                        list_cliente.append(fila)
                    elif (mes_actual != mes_anterior) and (cliente_actual == cliente_anterior):
                        list_mes.append(list_cliente)
                        list_general.append(list_mes)
                        list_mes = []
                        list_cliente = []
                        list_cliente.append(fila)
                    elif (mes_actual == mes_anterior) and (cliente_actual == cliente_anterior):
                        list_cliente.append(fila)
                        # list_mes.append(list_cliente)
                    elif (mes_actual != mes_anterior) and (cliente_actual != cliente_anterior):
                        list_mes.append(list_cliente)
                        list_general.append(list_mes)
                        list_mes = []
                        list_cliente = []
                        list_cliente.append(fila)
                else:
                    list_cliente = []
                    list_cliente.append(fila)
                count += 1
            if list_mes != []:
                list_general.append(list_mes)

            return list_general

            # sql = ''' SELECT COUNT(DISTINCT df.Cod_Material) AS producto, cf.Cod_Cliente, mc.Razon_Social, GROUP_CONCAT(df.Cod_Material SEPARATOR ' | ') AS materiales, cm.Texto_Breve,
            #         ROUND(SUM(df.Precio_Final)/1.18,2) AS precio_cot_sin_igv,
            #         (SUM(df.Precio_Final) - ROUND(SUM(df.Precio_Final)/1.18,2)) AS igv_cotizado,
            #         SUM(df.Precio_Final) AS precio_cotizado,
            #         GROUP_CONCAT(df.Precio_Final SEPARATOR ' | ') as precios,
            #         GROUP_CONCAT(df.Cantidad SEPARATOR ' | ') AS cantidades, SUM(df.Cantidad), COUNT(DISTINCT df.Nro_Facturacion) as nro_compras,
            #         GROUP_CONCAT(df.Fecha_Reg SEPARATOR ' | ') AS fechas_compra,
            #         SUM(IF(cf.Descuento_Global > 0.00,1,0)) AS nro_compras_en_oferta
            #     	FROM `TAB_VENTA_010_Detalle_Facturacion` df
            #     	LEFT JOIN `TAB_VENTA_009_Cabecera_Facturacion` cf ON cf.Cod_Soc=df.Cod_Soc AND cf.Año=df.Año AND cf.Tipo_Comprobante=df.Tipo_Comprobante AND cf.Serie=df.Serie AND cf.Nro_Facturacion=df.Nro_Facturacion
            #         LEFT JOIN `TAB_COM_001_Maestro Clientes` mc ON mc.Cod_Cliente=cf.Cod_Cliente
            #         LEFT JOIN `TAB_MAT_001_Catalogo_Materiales` cm ON cm.Cod_Soc=df.Cod_Soc AND cm.Cod_Mat=df.Cod_Material
            #         WHERE df.Fecha_Reg<=CAST('2021-11-30' AS date) AND df.Fecha_Reg>=CAST('2021-11-01' AS date)
            #         GROUP BY df.Cod_Material, cf.Cod_Cliente
            #         HAVING COUNT(DISTINCT df.Cod_Material)>1
            #         ORDER BY mc.Razon_Social, cm.Texto_Breve ASC ;'''

        def reporte_cliente_productos():
            list_general = procesar_consulta_cliente_productos(consulta_general_cliente_productos())
            # print(list_general)
            # print()
            list_resumen_cliente = procesar_consulta_resumen(consulta_resumen_cliente())
            list_resumen_producto = procesar_consulta_resumen(consulta_resumen_producto())
            # print(list_resumen_cliente)
            # print()
            # print(list_resumen_producto)

            wb = Workbook() #  7429

            count = 0
            for info in list_general:
                mes = info[0][0][11][5:7]
                año = info[0][0][11][:4]
                mes_inicial = global_fecha_inicio[5:7]
                año_inicial = global_fecha_inicio[:4]
                mes_final = global_fecha_fin[5:7]
                año_final = global_fecha_fin[:4]
                año_mes = año + mes
                año_mes_inicial = año_inicial + mes_inicial
                año_mes_final = año_final + mes_final

                if float(año_mes) >= float(año_mes_inicial) and float(año_mes) <= float(año_mes_final):
                    name_sheet = DICT_MESES[str(info[0][0][11][5:7])] + ' - ' + str(info[0][0][11][:4])
                    if count != 0:
                        hoja = wb.create_sheet(name_sheet)
                        # wb.active = hoja
                    else:
                        hoja = wb.active
                        hoja.title = name_sheet
                        count += 1

                    for data_cliente in info:
                        if data_cliente != []:
                            hoja.append((data_cliente[0][1],''))
                            hoja.append(('PRODUCTO', 'SUB TOTAL (US$)', 'IGV (US$)', 'TOTAL (US$)', 'CANTIDADES', 'CANT. TOTAL', 'NRO. COMPRAS', 'NRO. COMPRAS EN OFERTA')) # Crea la fila del encabezado con los títulos
                            color_relleno = rellenoSociedad(global_sociedad)

                            col_range = hoja.max_column
                            nueva_fila = hoja.max_row
                            for col in range(1, col_range + 1):
                                cell_header = hoja.cell(nueva_fila, col)
                                cell_header.fill = color_relleno
                                cell_header.font = NEGRITA
                            for producto in data_cliente:
                                fila = producto[3:-1]
                                hoja.append(fila)

                            for i in range(hoja.max_row):
                                if i >= nueva_fila-1:
                                    row = list(hoja.rows)[i]
                                    for col in range(hoja.max_column):
                                        row[col].border = BORDE_DELGADO
                                        if 1 <= col <= 3:
                                            row[col].number_format = FORMATO_DOLAR
                                        elif col == 4:
                                            row[col].alignment = ALINEACION_DERECHA
                                        elif col == 5:
                                            row[col].number_format = FORMATO_NUMERO
                        hoja.append(('',''))
                    ajustarColumnasSheet(hoja)

            def crear_sheet_resumen(wb, name_sheet, list_resumen, nombre_columna):
                hoja = wb.create_sheet(name_sheet)
                for data_cliente in list_resumen:
                    if data_cliente != []:
                        hoja.append((data_cliente[0][1],''))
                        hoja.append((nombre_columna, 'SUB TOTAL (US$)', 'IGV (US$)', 'TOTAL (US$)', 'CANTIDADES', 'CANT. TOTAL', 'NRO. COMPRAS', 'NRO. COMPRAS EN OFERTA')) # Crea la fila del encabezado con los títulos
                        color_relleno = rellenoSociedad(global_sociedad)

                        col_range = hoja.max_column
                        nueva_fila = hoja.max_row
                        for col in range(1, col_range + 1):
                            cell_header = hoja.cell(nueva_fila, col)
                            cell_header.fill = color_relleno
                            cell_header.font = NEGRITA
                        for producto in data_cliente:
                            fila = producto[3:-1]
                            hoja.append(fila)

                        for i in range(hoja.max_row):
                            if i >= nueva_fila-1:
                                row = list(hoja.rows)[i]
                                for col in range(hoja.max_column):
                                    row[col].border = BORDE_DELGADO
                                    if 1 <= col <= 3:
                                        row[col].number_format = FORMATO_DOLAR
                                    elif col == 4:
                                        row[col].alignment = ALINEACION_DERECHA
                                    elif col == 5:
                                        row[col].number_format = FORMATO_NUMERO
                    hoja.append(('',''))
                ajustarColumnasSheet(hoja)

            crear_sheet_resumen(wb, 'Resumen Cliente', list_resumen_cliente, 'PRODUCTO')
            crear_sheet_resumen(wb, 'Resumen Productos', list_resumen_producto, 'CLIENTE')
            ajustarColumnasSheet(hoja)

            return wb

        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        wb=reporte_cliente_productos()
        nombre_archivo = "Reporte Cliente vs Productos - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteDeudas(TemplateView):
    def get(self,request, *args,**kwargs):
        DICT_CLIENTE = {}
        query_cliente = Cliente.objects.all()
        for dato in query_cliente:
            c_id = str(dato.id)
            DICT_CLIENTE[c_id] = dato.razon_social
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')
        
        sql_productos = ''' (SELECT
            MAX(cvf.id) AS id,
            to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            CONCAT(MAX(dgs.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT), 6, '0')) AS numero_comprobante,
            STRING_AGG(mm.descripcion_corta, ' | ') AS texto_materiales
            FROM comprobante_venta_facturaventa cvf
            LEFT JOIN datos_globales_seriescomprobante dgs
                ON dgs.id = cvf.serie_comprobante_id AND dgs.tipo_comprobante_id='%s'
            LEFT JOIN comprobante_venta_facturaventadetalle cvfd
                ON cvfd.factura_venta_id=cvf.id
            LEFT JOIN material_material mm
                ON cvfd.content_type_id='%s' AND mm.id=cvfd.id_registro
            WHERE cvf.estado='4'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura)
            UNION
            (SELECT
            MAX(cvb.id) AS id,
            to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            CONCAT(MAX(dgs.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT), 6, '0')) AS numero_comprobante,
            STRING_AGG(mm.descripcion_corta, ' | ') AS texto_materiales
            FROM comprobante_venta_boletaventa cvb
            LEFT JOIN datos_globales_seriescomprobante dgs
                ON dgs.id = cvb.serie_comprobante_id AND dgs.tipo_comprobante_id='%s'
            LEFT JOIN comprobante_venta_boletaventadetalle cvbd
                ON cvbd.boleta_venta_id=cvb.id
            LEFT JOIN material_material mm
                ON cvbd.content_type_id='%s' AND mm.id=cvbd.id_registro
            WHERE cvb.estado='4'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta)''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['material | material'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['material | material'])
        query_productos = FacturaVenta.objects.raw(sql_productos)
        
        info = []
        for fila in query_productos:
            lista_datos = []
            lista_datos.append(fila.fecha_emision_comprobante)
            lista_datos.append(fila.numero_comprobante)
            lista_datos.append(fila.texto_materiales)
            info.append(lista_datos)

        dict_productos = {}
        for fila in info:
            dict_productos[fila[0]+'|'+fila[1]] = fila[2]

        objeto_sociedad = Sociedad.objects.get(id=global_sociedad)

        # Letras según fechas pactadas al cotizar
        sql_letras = ''' (SELECT
            MAX(cvf.id) AS id, 
            to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'FACTURA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            (CASE WHEN MAX(cvf.tipo_venta)='2'
            THEN (
                STRING_AGG(CONCAT(
                    to_char(cobc.fecha, 'DD/MM/YYYY'),
                    ' $ ',
                    CAST(ROUND(cobc.monto,2) AS TEXT)
                    ), '\n' ORDER BY cobc.fecha)
            ) ELSE (
                ''
            ) END) AS letras
            FROM comprobante_venta_facturaventa cvf
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvf.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
            LEFT JOIN cobranza_cuota cobc
                ON cobc.deuda_id=cd.id
            WHERE cvf.tipo_venta='2' AND cvf.sociedad_id='%s' AND cd.id IS NOT NULL AND cvf.estado='4'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
            ORDER BY cliente_denominacion ASC, letras ASC)
            UNION
            (SELECT
            MAX(cvb.id) AS id,
            to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            'BOLETA' AS tipo_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            (CASE WHEN MAX(cvb.tipo_venta)='2'
            THEN (
                STRING_AGG(CONCAT(
                    to_char(cobc.fecha, 'DD/MM/YYYY'),
                    ' $ ',
                    CAST(ROUND(cobc.monto,2) AS TEXT)
                    ), '\n' ORDER BY cobc.fecha)
            ) ELSE (
                ''
            ) END) AS letras
            FROM comprobante_venta_boletaventa cvb
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvb.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
            LEFT JOIN cobranza_cuota cobc
                ON cobc.deuda_id=cd.id
            WHERE cvb.tipo_venta='2' AND cvb.sociedad_id='%s' AND cd.id IS NOT NULL AND cvb.estado='4'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
            ORDER BY cliente_denominacion ASC, letras ASC)
            ORDER BY cliente_denominacion ASC, letras ASC ;''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)                                                               
        query_letras = FacturaVenta.objects.raw(sql_letras)

        info = []
        for fila in query_letras:
            lista_datos = []
            lista_datos.append(fila.fecha_emision_comprobante)
            lista_datos.append(fila.tipo_comprobante)
            lista_datos.append(fila.nro_comprobante)
            lista_datos.append(fila.cliente_denominacion)
            lista_datos.append(fila.letras)
            info.append(lista_datos)

        dict_letras = {}
        for fila in info:
            dict_letras[fila[0]+'|'+fila[1]+'|'+fila[2]] = fila[4]

        # facturas por jarrones -- mejorar esto..
        DICT_FACT_INVALIDAS = {}

        list_fact_invalidas_mpl = [
            'F001-000231',
            'F001-000233',
            'F001-000245',
            'F001-000298',
            ]
        list_fact_invalidas_mca = [
            'F001-002098',
            'F001-002099',
            ]
        DICT_FACT_INVALIDAS['4'] = list_fact_invalidas_mpl
        DICT_FACT_INVALIDAS['3'] = list_fact_invalidas_mca
        DICT_FACT_INVALIDAS['2'] = list_fact_invalidas_mpl
        DICT_FACT_INVALIDAS['1'] = list_fact_invalidas_mca

        sql = ''' (SELECT
            MAX(cn.id) AS id,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
            SUM(cp.monto) as monto_nota_credito,
            MAX(cc.razon_social) AS cliente_denominacion,
            'FACTURA' AS tipo_comprobante
            FROM cobranza_nota cn
            LEFT JOIN cobranza_pago cp
                ON cp.id_registro=cn.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_deuda cd
                ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
            LEFT JOIN comprobante_venta_facturaventa cvf
                ON cd.id_registro=cvf.id
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvf.cliente_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=cn.moneda_id
            WHERE cvf.sociedad_id='%s'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
            ORDER BY 4)
            UNION
            (SELECT
            MAX(cn.id) AS id,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
            SUM(cp.monto) as monto_nota_credito,
            MAX(cc.razon_social) AS cliente_denominacion,
            'BOLETA' AS tipo_comprobante
            FROM cobranza_nota cn
            LEFT JOIN cobranza_pago cp
                ON cp.id_registro=cn.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_deuda cd
                ON cp.deuda_id=cd.id  AND cd.content_type_id='%s'
            LEFT JOIN comprobante_venta_boletaventa cvb
                ON cd.id_registro=cvb.id
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvb.cliente_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=cn.moneda_id
            WHERE cvb.sociedad_id='%s'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
            ORDER BY 4) ; ''' % (DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], global_sociedad, DICT_CONTENT_TYPE['cobranza | nota'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], global_sociedad)
        query_info = Nota.objects.raw(sql)

        info = []
        for fila in query_info:
            lista_datos = []
            lista_datos.append(fila.nro_comprobante)
            lista_datos.append(fila.monto_nota_credito)
            lista_datos.append(fila.cliente_denominacion)
            lista_datos.append(fila.tipo_comprobante)
            info.append(lista_datos)
        
        dict_cobranza_nota = {}
        for fila in info:
            dict_cobranza_nota[fila[0]+'|'+fila[3]] = fila[1]

        sql = ''' (SELECT 
            MAX(cvf.id) AS id,
            to_char(MAX(cvf.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            MAX(cvf.total) AS monto_facturado,
            SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_amortizado,
            MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_pendiente,
            (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) AS estado_cobranza,
            to_char(MAX(cvf.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
            MAX(cvf.fecha_vencimiento) - MAX(cvf.fecha_emision) AS dias_credito,
            MAX(cvf.fecha_vencimiento) AS dias_vencimiento,
            '' AS letras,
            '' AS productos,
            'FACTURA' AS tipo_comprobante,
            MAX(cvf.fecha_emision) AS fecha_orden
            FROM comprobante_venta_facturaventa cvf
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvf.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvf.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvf.id
            LEFT JOIN cobranza_pago cp
                ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_ingreso ci
                ON ci.id=cp.id_registro
            LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                ON dgcb.id=ci.cuenta_bancaria_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=dgcb.moneda_id
            LEFT JOIN cobranza_redondeo cr
                ON cr.deuda_id=cd.id
            WHERE cvf.sociedad_id='%s' AND cvf.estado='4' AND cd.id IS NOT NULL AND cc.id = '%s'
            GROUP BY cvf.sociedad_id, cvf.tipo_comprobante, cvf.serie_comprobante_id, cvf.numero_factura
            HAVING (CASE WHEN MAX(cvf.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) = 'PENDIENTE' AND CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvf.numero_factura) AS TEXT),6,'0')) NOT IN %s
            ORDER BY cliente_denominacion ASC, fecha_orden ASC)
            UNION
            (SELECT
            MAX(cvb.id) AS id,
            to_char(MAX(cvb.fecha_emision), 'DD/MM/YYYY') AS fecha_emision_comprobante,
            CONCAT(MAX(dgsc.serie), '-', lpad(CAST(MAX(cvb.numero_boleta) AS TEXT),6,'0')) AS nro_comprobante,
            MAX(cc.razon_social) AS cliente_denominacion,
            MAX(cvb.total) AS monto_facturado,
            SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) + SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_amortizado,
            MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - SUM(CASE WHEN cr.monto IS NOT NULL THEN (cr.monto) ELSE 0.00 END) AS monto_pendiente,
            (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) AS estado_cobranza,
            to_char(MAX(cvb.fecha_vencimiento), 'DD/MM/YYYY') AS fecha_vencimiento_comprobante,
            MAX(cvb.fecha_vencimiento) - MAX(cvb.fecha_emision) AS dias_credito,
            MAX(cvb.fecha_vencimiento) AS dias_vencimiento,
            '' AS letras,
            '' AS productos,
            'BOLETA' AS tipo_comprobante,
            MAX(cvb.fecha_emision) AS fecha_orden
            FROM comprobante_venta_boletaventa cvb
            LEFT JOIN datos_globales_seriescomprobante dgsc
                ON dgsc.tipo_comprobante_id='%s' AND dgsc.id=cvb.serie_comprobante_id
            LEFT JOIN clientes_cliente cc
                ON cc.id=cvb.cliente_id
            LEFT JOIN cobranza_deuda cd
                ON cd.content_type_id='%s' AND cd.id_registro=cvb.id
            LEFT JOIN cobranza_pago cp
                ON cp.deuda_id=cd.id AND cp.content_type_id='%s'
            LEFT JOIN cobranza_ingreso ci
                ON ci.id=cp.id_registro
            LEFT JOIN datos_globales_cuentabancariasociedad dgcb
                ON dgcb.id=ci.cuenta_bancaria_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=dgcb.moneda_id
            LEFT JOIN cobranza_redondeo cr
                ON cr.deuda_id=cd.id
            WHERE cvb.sociedad_id='%s' AND cvb.estado='4' AND cd.id IS NOT NULL AND cc.id = '%s'
            GROUP BY cvb.sociedad_id, cvb.tipo_comprobante, cvb.serie_comprobante_id, cvb.numero_boleta
            HAVING (CASE WHEN MAX(cvb.total) - SUM(CASE WHEN dgm.abreviatura='PEN' THEN ROUND(cp.monto/cp.tipo_cambio,2) ELSE cp.monto END) - (CASE WHEN MAX(cr.monto) IS NOT NULL THEN MAX(cr.monto) ELSE 0.00 END) <= 0.00
                THEN (
                    'CANCELADO'
                ) ELSE (
                    'PENDIENTE'
                ) END) = 'PENDIENTE'
            ORDER BY cliente_denominacion ASC, fecha_orden ASC)
            ORDER BY cliente_denominacion ASC, fecha_orden ASC ; ''' %(DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['comprobante_venta | facturaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad, global_cliente, tuple(DICT_FACT_INVALIDAS[global_sociedad]), DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['comprobante_venta | boletaventa'], DICT_CONTENT_TYPE['cobranza | ingreso'], global_sociedad, global_cliente)
        query_info = FacturaVenta.objects.raw(sql)

        list_general = []
        for fila in query_info:
            lista_datos = []
            lista_datos.append(fila.fecha_emision_comprobante)
            lista_datos.append(fila.nro_comprobante)
            lista_datos.append(fila.cliente_denominacion)
            lista_datos.append(fila.monto_facturado)
            lista_datos.append(fila.monto_amortizado)
            lista_datos.append(fila.monto_pendiente)
            lista_datos.append(fila.estado_cobranza)
            lista_datos.append(fila.fecha_vencimiento_comprobante)
            lista_datos.append(fila.dias_credito)
            lista_datos.append(str(fila.dias_vencimiento))
            lista_datos.append(fila.letras)
            lista_datos.append(fila.productos)
            lista_datos.append(fila.tipo_comprobante)
            list_general.append(lista_datos)

        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        for fila in list_general:
            fila[3] = float(fila[3])
            if fila[4] == None:
                fila[4] = '0.00'
            if fila[5] == None:
                fila[5] = fila[3]
            fila[4] = float(fila[4])
            fila[5] = float(fila[5])

            if fila[1] + '|' + fila[12] in dict_cobranza_nota:
                fila[4] += float(dict_cobranza_nota[fila[1] + '|' + fila[12]])
                fila[5] = fila[3] - fila[4]

            if fila[9] != '':
                fecha1 = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                fecha2 = datetime.strptime(fila[9], '%Y-%m-%d')
                dias = (fecha1 - fecha2) / timedelta(days=1)
                fila[9] = str(dias)
            if float(dias) > float(0):
                fila[6] = 'VENCIDO'
            if fila[11] == None:
                fila[11] = ''
            try:
                fila[10] = dict_letras[fila[0]+'|'+fila[12]+'|'+fila[1]]
                div = fila[10].split('\n')
                rest = float(fila[4])
                list_resumen_letra = []
                for sub_div in div:
                    list_fecha_monto = sub_div.split(' $ ')
                    fecha_letra = list_fecha_monto[0]
                    monto_letra = float(list_fecha_monto[1])
                    rest = rest - monto_letra
                    if rest >= float(0):
                        estado_letra = "CANCELADO"
                    else:
                        fecha_base = datetime.strptime(fecha_hoy, '%Y-%m-%d')
                        fecha_letra_dt = datetime.strptime(fecha_letra, '%d/%m/%Y')
                        dias = (fecha_base - fecha_letra_dt) / timedelta(days=1)
                        if float(dias) > float(0):
                            estado_letra = 'VENCIDO'
                        else:
                            estado_letra = "PENDIENTE"
                    fila_letra = fecha_letra + ' $ ' + str(monto_letra) + ' ' + estado_letra
                    list_resumen_letra.append(fila_letra)
                fila[10] = '\n'.join(list_resumen_letra)
                    
            except:
                fila[10] = ''

        fecha_texto = formatoFechaTexto(StrToDate(fecha_hoy))
        fecha_invertida = datetime.now().strftime("%d-%m-%Y")

        color = DICT_SOCIEDAD[global_sociedad].color
        #####
        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        #####
        titulo = "Reporte de Deudas - " + abreviatura + " - " + DICT_CLIENTE[global_cliente] + " - " + FECHA_HOY
        vertical = False
        alinear = 'right'
        logo = [[objeto_sociedad.logo.url, alinear]]
        pie_pagina = objeto_sociedad.pie_pagina
        list_texto = []
        # texto = '''Lima, %s''' % str(fecha_texto) + '\n''\n' + '''SR. ''' + DICT_CLIENTE[global_cliente] + '\n' + '''Estimado cliente, se le remite la deuda actualizada al día de hoy <strong>%s</strong>, cuyos detalles son los siguientes:''' % (str(fecha_hoy))
        text = Paragraph('''<para align=center><strong>%s</strong></para>''' %("REPORTE DE DEUDAS"), styleSheet["ComicNeue-Bold-14"])
        list_texto.append(text)
        text = '''Lima, %s''' % str(fecha_texto) + '\n''\n'
        list_texto.append(text)
        text = Paragraph('''<b>SR. ''' + DICT_CLIENTE[global_cliente] + '''</b>''')
        list_texto.append(text)
        text = Paragraph('''Estimado cliente, se le remite la deuda actualizada al día de hoy <strong>%s</strong>, cuyos detalles son los siguientes:''' % (str(fecha_invertida)))
        list_texto.append(text)

        TablaEncabezado = [
            'FECHA',
            'NRO. COMPROBANTE',
            'FACTURADO',
            'AMORTIZADO',
            'PENDIENTE',
            'ESTADO',
            'FEC. VENC.',
            'CRÉDITO',
            'DIAS VENC.',
            'LETRAS',
            'PRODUCTOS',
            ]
        suma_deuda_total = 0

        TablaDatos = []
        for lista in list_general:
            if lista[5] > float(0):
                fila = []
                fila.append(lista[0])
                fila.append(lista[1])
                fila.append(lista[3])
                fila.append(lista[4])
                fila.append(lista[5])
                suma_deuda_total += float(lista[5])
                fila.append(lista[6])
                fila.append(lista[7])
                fila.append(lista[8])
                fila.append(lista[9])
                fila.append(lista[10])
                fila.append(dict_productos[lista[0]+'|'+lista[1]])
                TablaDatos.append(fila)

        TablaDatos.append(["","","", "Deuda Total:", round(suma_deuda_total,2) ,"","","","","","",""])

        # texto = '''Agradeceremos pueda realizar los pagos a nombre de <strong>%s</strong> y confirmarnos el pago en cualquiera de las siguientes cuentas:''' % DICT_SOCIEDAD[global_sociedad].razon_social
        # list_texto.append(texto)

        text = Paragraph('''Agradeceremos pueda realizar los pagos a nombre de <strong>%s</strong> y confirmarnos el pago en cualquiera de las siguientes cuentas:''' % DICT_SOCIEDAD[global_sociedad].razon_social)
        list_texto.append(text)

        sql_cuentas_bancarias = '''SELECT
            dgcb.id,
            dgb.nombre_comercial AS nombre_banco,
            dgcb.numero_cuenta AS cuenta_banco,
            dgcb.numero_cuenta_interbancaria AS cuenta_cci_banco,
            dgm.nombre AS moneda_descripcion
            FROM datos_globales_cuentabancariasociedad dgcb
            LEFT JOIN datos_globales_banco dgb
                ON dgb.id=dgcb.banco_id
            LEFT JOIN datos_globales_moneda dgm
                ON dgm.id=dgcb.moneda_id
            LEFT JOIN sociedad_sociedad ss
                ON ss.id=dgcb.sociedad_id
            WHERE dgcb.sociedad_id='%s' AND dgcb.estado='1' AND dgcb.efectivo=False''' %(global_sociedad)
        query_info_cuentas = CuentaBancariaSociedad.objects.raw(sql_cuentas_bancarias)

        info_cuentas = []
        for fila in query_info_cuentas:
            lista_datos = []

            lista_datos.append(fila.nombre_banco)
            lista_datos.append(fila.cuenta_banco)
            lista_datos.append(fila.cuenta_cci_banco)
            lista_datos.append(fila.moneda_descripcion)
            info_cuentas.append(lista_datos)


            list_cuenta_dolares, list_cuenta_soles = [], []
            for fila in info_cuentas:
                if 'SOL' in fila[3]: # SOLES
                    list_temp = []
                    list_temp.extend([fila[0], fila[1], fila[2]])
                    list_cuenta_soles.append(list_temp)
                if 'DÓLAR' in fila[3]: # DOLARES
                    list_temp = []
                    list_temp.extend([fila[0], fila[1], fila[2]])
                    list_cuenta_dolares.append(list_temp)

        buf = generarReporteDeudas(titulo, vertical, logo, pie_pagina, list_texto, TablaEncabezado, TablaDatos, color, list_cuenta_dolares, list_cuenta_soles)

        respuesta = HttpResponse(buf.getvalue(), content_type='application/pdf')
        respuesta.headers['content-disposition']='inline; filename=%s.pdf' % titulo

        return respuesta


class ReporteCobranza(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        titulo = "Reporte de Cobranza - " + abreviatura + " - " + FECHA_HOY
        # DICT_FACTURAS_SORTEO = {
        #     '1000': [
        #         ],
        #     '2000': [
        #         'F001-002098',
        #         'F001-002099',
        #         ],
        # }

        reporte_cobranza(self.request, titulo)

        buf = generar_reporte_cobranza(global_sociedad, titulo)
        respuesta = HttpResponse(buf.getvalue(), content_type='application/pdf')
        respuesta.headers['content-disposition']='inline; filename=%s.pdf' % titulo
            
        return respuesta


class ReporteRotacion(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def consulta_rotacion():
            sql_descrip = ''' SELECT
                mm.id AS id,
                mm.id AS id_producto,
                mm.id AS cod_mat,
                mf.nombre AS familia_material,
                mm.descripcion_corta,
                mm.descripcion_venta
                FROM material_material mm
                LEFT JOIN material_subfamilia msf
                    ON mm.subfamilia_id=msf.id
                LEFT JOIN material_familia mf
                    ON mf.id=msf.familia_id
                ORDER BY mf.nombre, mm.descripcion_corta, mm.descripcion_venta ; '''
            query_info = Material.objects.raw(sql_descrip)
            
            info_descrip = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.id_producto)
                lista_datos.append(dato_fila.cod_mat)
                lista_datos.append(dato_fila.familia_material)
                lista_datos.append(dato_fila.descripcion_corta)
                lista_datos.append(dato_fila.descripcion_venta)
                info_descrip.append(lista_datos)

            dict_descrip = {}
            for fila in info_descrip:
                list_temp = []
                list_temp.extend([fila[1], fila[2], fila[3], fila[4]])
                dict_descrip[fila[0]] = list_temp


            sql_stock = ''' SELECT
                MAX(mam.id) AS id,
                mm.id AS id_producto,
                mm.descripcion_corta,
                SUM(ROUND(cantidad, 3) * signo_factor_multiplicador) AS stock
                FROM movimiento_almacen_movimientosalmacen mam
                LEFT JOIN material_material mm
                    ON mm.id=mam.id_registro_producto
                LEFT JOIN movimiento_almacen_tipostock mats
                    ON mam.tipo_stock_id=mats.id
                WHERE mam.content_type_producto_id='%s' AND mats.codigo IN (
                    3, 5, 36)
                GROUP BY mm.id
                ORDER BY mm.descripcion_corta ; ''' %(DICT_CONTENT_TYPE['material | material'])
            query_info = MovimientosAlmacen.objects.raw(sql_stock)
            
            info_stock = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.id_producto)
                lista_datos.append(dato_fila.descripcion_corta)
                lista_datos.append(dato_fila.stock)
                info_stock.append(lista_datos)

            dict_stock = {}
            for fila in info_stock:
                dict_stock[fila[0]] = fila[2]

            sql_pedidos = ''' SELECT
                MAX(ccpid.id) AS id,
                MAX(mm.id) AS id_producto,
                MAX(mm.descripcion_corta) AS nombre_producto,
                ROUND(MAX(ccpid.precio_final_con_igv),2) AS precio_pedido,
                to_char(MAX(ccpi.fecha_comprobante), 'DD/MM/YYYY') AS fecha_pedido,
                MAX(ccpi.fecha_comprobante) AS fecha_orden
                FROM comprobante_compra_comprobantecomprapidetalle ccpid
                LEFT JOIN orden_compra_ordencompradetalle ocod
                    ON ocod.id=ccpid.orden_compra_detalle_id
                LEFT JOIN comprobante_compra_comprobantecomprapi ccpi
                    ON ccpi.id=ccpid.comprobante_compra_id
                LEFT JOIN material_material mm
                    ON mm.id=ocod.id_registro AND ocod.content_type_id='%s'
                GROUP BY mm.id
                ORDER BY fecha_orden DESC ; ''' %(DICT_CONTENT_TYPE['material | material'])
            query_info = ComprobanteCompraPIDetalle.objects.raw(sql_pedidos)
            
            info_pedidos = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.id_producto)
                lista_datos.append(dato_fila.nombre_producto)
                lista_datos.append(dato_fila.precio_pedido)
                lista_datos.append(dato_fila.fecha_pedido)
                lista_datos.append(dato_fila.fecha_orden)
                info_pedidos.append(lista_datos)

            dict_precios, dict_fecha_pedido = {}, {}
            list_mat_pedidos = []
            for fila in info_pedidos:
                dict_precios[fila[0]] = fila[2]
                if fila[3]:
                    dict_fecha_pedido[fila[0]] = fila[3]
                if fila[0] not in list_mat_pedidos:
                    list_mat_pedidos.append(fila[0])

            # select date_sub('2018-01-01', interval -1 day);
            sql_rotacion_6ultimos = ''' SELECT
                MAX(cvfd.id) AS id,
                MAX(mm.id) AS id_producto,
                CONCAT(CAST(ROUND(SUM(cvfd.cantidad),2) AS TEXT), ' / ', CAST(ROUND(SUM(cvfd.cantidad)/6,2) AS TEXT)) AS venta_6_meses,
                MAX(mf.nombre) as familia_nombre,
                MAX(mm.descripcion_corta) AS nombre_material
                FROM comprobante_venta_facturaventadetalle cvfd
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN material_material mm
                    ON cvfd.id_registro=mm.id AND content_type_id='%s'
                LEFT JOIN material_subfamilia msf
                    ON mm.subfamilia_id=msf.id
                LEFT JOIN material_familia mf
                    ON mf.id=msf.familia_id
                WHERE cvfd.created_at >= CURRENT_DATE - INTERVAL '6 months' AND cvf.estado = '4'
                GROUP BY mm.id
                ORDER BY familia_nombre;''' %(DICT_CONTENT_TYPE['material | material'])
            query_info = FacturaVentaDetalle.objects.raw(sql_rotacion_6ultimos)
            
            info_rotacion_6ultimos = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.id_producto)
                lista_datos.append(dato_fila.venta_6_meses)
                lista_datos.append(dato_fila.familia_nombre)
                lista_datos.append(dato_fila.nombre_material)
                info_rotacion_6ultimos.append(lista_datos)

            dict_6ultimos = {}
            for fila in info_rotacion_6ultimos:
                dict_6ultimos[fila[0]] = fila[1]

            sql_rotacion = ''' SELECT
                MAX(cvfd.id) AS id,
                mm.id AS id_producto,
                mm.id AS cod_mat,
                MAX(mf.nombre) AS familia_nombre,
                MAX(mm.descripcion_corta) AS nombre_material_breve,
                MAX(mm.descripcion_venta) AS nombre_material_venta,
                '' AS precio,
                '' AS stock,
                ROUND(SUM(cvfd.cantidad),3) AS venta_total,
                ROUND(SUM(cvfd.cantidad) / CAST(
                        CASE WHEN
                            EXTRACT(year FROM AGE(CURRENT_DATE, MIN(cvfd.created_at)))*12 + EXTRACT(month FROM AGE(CURRENT_DATE, MIN(cvfd.created_at))) != 0
                        THEN
                            EXTRACT(year FROM AGE(CURRENT_DATE, MIN(cvfd.created_at)))*12 + EXTRACT(month FROM AGE(CURRENT_DATE, MIN(cvfd.created_at)))
                        ELSE
                            SUM(cvfd.cantidad)
                        END
                    AS NUMERIC) ,3) AS venta_mensual,
                '' AS venta_6_meses,
                '' AS venta_ultimo_ingreso,
                '' AS tiempo_duracion,
                '' AS pedido,
                '' AS sugerencia
                FROM comprobante_venta_facturaventadetalle cvfd
                LEFT JOIN comprobante_venta_facturaventa cvf
                    ON cvf.id=cvfd.factura_venta_id
                LEFT JOIN material_material mm
                    ON cvfd.id_registro=mm.id AND content_type_id='%s'
                LEFT JOIN material_subfamilia msf
                    ON mm.subfamilia_id=msf.id
                LEFT JOIN material_familia mf
                    ON mf.id=msf.familia_id
                WHERE cvf.estado = '4' AND mm.id IS NOT NULL
                GROUP BY mm.id
                ORDER BY familia_nombre, nombre_material_breve, nombre_material_venta ; ''' %(DICT_CONTENT_TYPE['material | material'])
            query_info = FacturaVentaDetalle.objects.raw(sql_rotacion)
            
            info_rotacion = []
            for dato_fila in query_info:
                lista_datos = []
                lista_datos.append(dato_fila.id_producto)
                lista_datos.append(dato_fila.cod_mat)
                lista_datos.append(dato_fila.familia_nombre)
                lista_datos.append(dato_fila.nombre_material_breve)
                lista_datos.append(dato_fila.nombre_material_venta)
                lista_datos.append(dato_fila.precio)
                lista_datos.append(dato_fila.stock)
                lista_datos.append(dato_fila.venta_total)
                lista_datos.append(dato_fila.venta_mensual)
                lista_datos.append(dato_fila.venta_6_meses)
                lista_datos.append(dato_fila.venta_ultimo_ingreso)
                lista_datos.append(dato_fila.tiempo_duracion)
                lista_datos.append(dato_fila.pedido)
                lista_datos.append(dato_fila.sugerencia)
                info_rotacion.append(lista_datos)

            list_general = []
            for fila in info_rotacion:
                if fila[0]!='163': # servicio de activacion
                    list_temp = []
                    if fila[0] not in dict_fecha_pedido:
                        list_temp.extend(['2000-01-01', fila[0]])
                    else:
                        list_temp.extend([dict_fecha_pedido[fila[0]], fila[0]])
                    list_general.append(list_temp)

            def funcion_temporal(list_temp):
                fecha_pedido = list_temp[0]
                id_producto = list_temp[1]
                sql = ''' SELECT
                    MAX(cvfd.id) AS id,
                    mm.id AS id_producto,
                    mm.id AS cod_mat,
                    MAX(mf.nombre) AS familia_nombre,
                    MAX(mm.descripcion_corta) AS nombre_material_breve,
                    MAX(mm.descripcion_venta) AS nombre_material_venta,
                    CONCAT(CAST(ROUND(SUM(cvfd.cantidad),3) AS TEXT), ' / ',
                        CAST(ROUND(SUM(cvfd.cantidad) / CAST(
                                CASE WHEN
                                    EXTRACT(year FROM AGE(CURRENT_DATE, CAST('%s' AS DATE)))*12 + EXTRACT(month FROM AGE(CURRENT_DATE, CAST('%s' AS DATE))) != 0
                                THEN
                                    EXTRACT(year FROM AGE(CURRENT_DATE, CAST('%s' AS DATE)))*12 + EXTRACT(month FROM AGE(CURRENT_DATE, CAST('%s' AS DATE)))
                                ELSE
                                    SUM(cvfd.cantidad)
                                END
                            AS NUMERIC) ,2) AS TEXT) ) AS venta_desde_ultimo_pedido
                    FROM comprobante_venta_facturaventadetalle cvfd
                    LEFT JOIN comprobante_venta_facturaventa cvf
                        ON cvf.id=cvfd.factura_venta_id
                    LEFT JOIN material_material mm
                        ON cvfd.id_registro=mm.id AND content_type_id='%s'
                    LEFT JOIN material_subfamilia msf
                        ON mm.subfamilia_id=msf.id
                    LEFT JOIN material_familia mf
                        ON mf.id=msf.familia_id
                    WHERE mm.id='%s' AND cvfd.created_at>=CAST('%s' AS DATE) AND cvf.estado = '4'
                    GROUP BY mm.id
                    ORDER BY familia_nombre, nombre_material_breve, nombre_material_venta ''' %(fecha_pedido, fecha_pedido, fecha_pedido, fecha_pedido, DICT_CONTENT_TYPE['material | material'], id_producto, fecha_pedido)
                return '(' + sql + ')\n'
                # CONCAT(SUM(df.Cantidad), ' / ', ROUND(SUM(df.Cantidad)/(TIMESTAMPDIFF(DAY, '%s', CURDATE())/30),2))

            sql_general = ''
            for lista in list_general:
                if list_general.index(lista) == 0:
                    sql_general = funcion_temporal(lista)
                else:
                    sql_general += 'UNION\n' + funcion_temporal(lista)
            if sql_general != '':
                sql_general += 'ORDER BY 3,4,5 ASC'

                query_info = FacturaVentaDetalle.objects.raw(sql_general)
            
                info = []
                for dato_fila in query_info:
                    lista_datos = []
                    lista_datos.append(dato_fila.id_producto)
                    lista_datos.append(dato_fila.cod_mat)
                    lista_datos.append(dato_fila.familia_nombre)
                    lista_datos.append(dato_fila.nombre_material_breve)
                    lista_datos.append(dato_fila.nombre_material_venta)
                    lista_datos.append(dato_fila.venta_desde_ultimo_pedido)
                    info.append(lista_datos)

                print(info)
                dict_venta_ultimo_ingreso = {}
                for fila in info:
                    dict_venta_ultimo_ingreso[fila[0]] = fila[5]


            print(100*'-')
            list_temp = []
            list_mat_vendidos = []
            for fila in info_rotacion:
                if fila[0] != '163': # Servicio de Activación
                    try:
                        fila[5] = float(dict_precios[fila[0]])
                    except:
                        pass
                else:
                    fila[5] = float('0.00')
                if fila[0] in dict_stock: # Servicio de Activación
                    try:
                        fila[6] = dict_stock[fila[0]]
                    except:
                        pass
                else:
                    fila[6] = '0.00'
                if fila[0] in dict_6ultimos:
                    fila[9] = dict_6ultimos[fila[0]]
                else:
                    fila[9] = '0.00 / 0.00'
                if fila[0] in dict_venta_ultimo_ingreso:
                    fila[10] = dict_venta_ultimo_ingreso[fila[0]]
                else:
                    list_temp.append(fila[0])
                    fila[10] = '0.00 / 0.00'
                promedio_1 = fila[9][fila[9].find(' / ')+3:]
                try:
                    promedio_2 = fila[10][fila[10].find(' / ')+3:]
                except:
                    promedio_2 = '0.000'
                if float(promedio_2) >= float(promedio_1):
                    try:
                        cant_meses = float(fila[6]) / float(promedio_2)
                    except ZeroDivisionError:
                        cant_meses = 0
                    fila[12] = formatearDecimal(str(round(float(promedio_2)*5,0)),'2')
                else:
                    try:
                        cant_meses = float(fila[6]) / float(promedio_1)
                    except ZeroDivisionError:
                        cant_meses = 0
                    fila[12] = formatearDecimal(str(round(float(promedio_1)*5,0)),'2')
                fila[11] = formatearDecimal(str(cant_meses),'2') + ' mes(es)'
                if cant_meses >= 5:
                    fila[13] = 'NO TRAER'
                else:
                    fila[13] = 'EVALUAR'
                if fila[0] not in list_mat_vendidos:
                    list_mat_vendidos.append(fila[0])
                # print(fila)

            print('NO SE TIENE VENTAS DESDE SU ÚLTIMO INGRESO DE LOS PRODUCTOS:',list_temp)


            list_productos_no_vendidos = []
            for mat in list_mat_pedidos:
                if mat not in list_mat_vendidos and mat != None:
                    print('PRODUCTO NO VENDIDO DESDE SU LLEGADA:',mat)
                    list_temp = []
                    stock_actual = '0.00' if mat not in dict_stock else dict_stock[mat]
                    list_temp.extend([mat, dict_descrip[mat][0], dict_descrip[mat][1], dict_descrip[mat][2], dict_descrip[mat][3], float(dict_precios[mat]), stock_actual, '0.00', '0.00', '0.00 / 0.00', '0.00 / 0.00', ' - ', '0.00', 'NO SE VENDIÓ'])
                    list_productos_no_vendidos.append(list_temp)

            info = info_rotacion + list_productos_no_vendidos
            info_ordenado = sorted(info, key=lambda x: x[2])
            return info_ordenado

        def reporte_rotacion():
            wb = Workbook()
            hoja = wb.active
            hoja.append(('ID MAT.', 'COD. MATERIAL', 'FAMILIA', 'NOMBRE', 'DESCRIPCIÓN', 'PRECIO', 'STOCK', 'VENTA TOTAL', 'VENTAS MENSUALES DEL TOTAL', 'VENTA DESDE ÚLTIMOS 6 MESES (total/promedio)', 'VENTA DESDE ULTIMO INGRESO (total/promedio)', 'TIEMPO DURACION (APROX.)', 'PEDIDO PARA 5 MESES', 'SUGERENCIA'))
            
            color_relleno = rellenoSociedad(global_sociedad)

            col_range = hoja.max_column  # get max columns in the worksheet
            # cabecera de la tabla
            for col in range(1, col_range + 1):
                cell_header = hoja.cell(1, col)
                cell_header.fill = color_relleno
                cell_header.font = NEGRITA

            info = consulta_rotacion()
            for producto in info:
                hoja.append(producto) # Crea la fila del encabezado con los títulos

            # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11, M=12, N=13
            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    # if 8 <= col <=11:
                    if col == 5:
                        row[col].alignment = ALINEACION_DERECHA
                        row[col].number_format = FORMATO_DOLAR
                    elif 6 <= col <= 8 or col == 11 or col == 12:
                        row[col].alignment = ALINEACION_DERECHA
                    elif 9 <= col <= 10:
                        row[col].alignment = ALINEACION_CENTRO
                    elif col == 13:
                        if row[col].value == 'EVALUAR':
                            row[col].font =  COLOR_AZUL
                        if row[col].value == 'NO SE VENDIÓ':
                            row[col].font =  COLOR_ROJO
                    #     row[col].number_format = self.formato_soles
                    # elif col == 15:
                    #     if row[col].value != 'LINK':
                    #         row[col].hyperlink =  row[col].value
                    #         row[col].font =  COLOR_AZUL
            hoja.freeze_panes = 'A2'
            ajustarColumnasSheet(hoja)
            return wb

        abreviatura = "MPL-MCA"
        wb=reporte_rotacion()
        nombre_archivo = "Reporte Rotacion - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteResumenStockProductosExcel(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        def consulta_stock_productos():

            sql_stock_productos = ''' SELECT
                MAX(mam.id) AS id,
                mm.id,
                mm.descripcion_corta,
                ROUND(SUM(CASE WHEN (mats.codigo='3') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_disponible,
                ROUND(SUM(CASE WHEN (mats.codigo='5') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_sin_qa,
                ROUND(SUM(CASE WHEN (mats.codigo='6') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_por_qa,
                ROUND(SUM(CASE WHEN (mats.codigo NOT IN (3,5,6)) THEN mam.cantidad*mam.signo_factor_multiplicador ELSE (0.00) END),3) AS stock_otros,
                ROUND(SUM(mam.cantidad*mam.signo_factor_multiplicador),3) as total_stock
                FROM movimiento_almacen_movimientosalmacen mam
                LEFT JOIN material_material mm
                    ON mm.id=mam.id_registro_producto AND mam.content_type_producto_id='%s'
                LEFT JOIN movimiento_almacen_tipostock mats
                    ON mam.tipo_stock_id=mats.id
                WHERE mam.sociedad_id='%s' AND mats.codigo NOT IN (
                    1, 2, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25)
                GROUP BY mm.id
                ORDER BY mm.descripcion_corta ; ''' %(DICT_CONTENT_TYPE['material | material'], global_sociedad)
            query_info = MovimientosAlmacen.objects.raw(sql_stock_productos)

            info = []
            for fila in query_info:
                lista_datos = []
                lista_datos.append(fila.id)
                lista_datos.append(fila.descripcion_corta)
                lista_datos.append(fila.stock_disponible)
                lista_datos.append(fila.stock_sin_qa)
                lista_datos.append(fila.stock_por_qa)
                lista_datos.append(fila.stock_otros)
                lista_datos.append(fila.total_stock)
                info.append(lista_datos)

            return info

        def generar_reporte():

            info = consulta_stock_productos()

            list_encabezado = [
                'COD. MAT.',
                'DESCRIPCIÓN',
                'ALMACÉN #1',
                'ALMACÉN #2',
                'ALMACÉN #3',
                'ALMACÉN #4',
                'ALMACÉN #5',
                'SUMA CONTEO',
                'DISPONIBLE',
                'BLOQ. SIN QA',
                'BLOQ. POR QA',
                'BLOQ. DESG.',
                'SUMATORIA',
                'DIFERENCIA',
                ]

            color_relleno = rellenoSociedad(global_sociedad)

            wb = Workbook()
            hoja = wb.active
            hoja.append(tuple(list_encabezado))

            col_range = hoja.max_column  # get max columns in the worksheet
            # cabecera de la tabla
            for col in range(1, col_range + 1):
                if col == 8 or col == 13:
                    # color_celda_cabecera = PatternFill(start_color='8C4966', end_color='8C4966', fill_type='solid')
                    color_celda_cabecera = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                elif col == 14:
                    color_celda_cabecera = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                else:
                    color_celda_cabecera = color_relleno
                cell_header = hoja.cell(1, col)
                cell_header.fill = color_celda_cabecera
                cell_header.font = NEGRITA
            # if info == []:
            #     return False
            # for bloque in info:
            # for fila in bloque:
            for fila in info:
                fila[2] = float(fila[2])
                fila[3] = float(fila[3])
                fila[4] = float(fila[4])
                fila[5] = float(fila[5])
                fila[6] = float(fila[6])
                nueva_fila = []
                nueva_fila.extend([
                    fila[0],
                    fila[1],
                    '',     # ALMACEN 1
                    '',     # ALMACEN 2
                    '',     # ALMACEN 3
                    '',     # ALMACEN 4
                    '',     # ALMACEN 5
                    '',     # SUMA CONTEO
                    fila[2],
                    fila[3],
                    fila[4],
                    fila[5],
                    fila[6],
                    '',     # DIFERENCIA
                    ])
                hoja.append(nueva_fila)

            for row in hoja.rows:
                for col in range(hoja.max_column):
                    row[col].border = BORDE_DELGADO
                    if col >= 2:
                        row[col].number_format = FORMATO_NUMERO

            hoja.freeze_panes = 'C2'
            ajustarColumnasSheet(hoja)
            return wb
    
    
        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        wb=generar_reporte()
        nombre_archivo = "Reporte Resumen Stock Productos - " + abreviatura + " - " + FECHA_HOY + ".xlsx"
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(nombre_archivo)
        respuesta['content-disposition']= content
        wb.save(respuesta)
        return respuesta


class ReporteResumenStockProductosPDF(TemplateView):
    def get(self,request, *args,**kwargs):
        global_sociedad = self.request.GET.get('filtro_sociedad')
        global_fecha_inicio = self.request.GET.get('filtro_fecha_inicio')
        global_fecha_fin = self.request.GET.get('filtro_fecha_fin')
        global_cliente = self.request.GET.get('filtro_cliente')

        sql_stock_productos = ''' SELECT
            MAX(mam.id) AS id,
            mm.id,
            mm.descripcion_corta,
            ROUND(SUM(CASE WHEN (mats.codigo='3') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_disponible,
            ROUND(SUM(CASE WHEN (mats.codigo='5') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_sin_qa,
            ROUND(SUM(CASE WHEN (mats.codigo='6') THEN (mam.cantidad*mam.signo_factor_multiplicador) ELSE (0.00) END),3) AS stock_por_qa,
            ROUND(SUM(CASE WHEN (mats.codigo NOT IN (3,5,6)) THEN mam.cantidad*mam.signo_factor_multiplicador ELSE (0.00) END),3) AS stock_otros,
            ROUND(SUM(mam.cantidad*mam.signo_factor_multiplicador),3) as total_stock
            FROM movimiento_almacen_movimientosalmacen mam
            LEFT JOIN material_material mm
                ON mm.id=mam.id_registro_producto AND mam.content_type_producto_id='%s'
            LEFT JOIN movimiento_almacen_tipostock mats
                ON mam.tipo_stock_id=mats.id
            WHERE mam.sociedad_id='%s' AND mats.codigo NOT IN (
                1, 2, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25)
            GROUP BY mm.id
            ORDER BY mm.descripcion_corta ; ''' %(DICT_CONTENT_TYPE['material | material'], global_sociedad)
        query_info = MovimientosAlmacen.objects.raw(sql_stock_productos)

        info = []
        for fila in query_info:
            lista_datos = []
            lista_datos.append(fila.id)
            lista_datos.append(fila.descripcion_corta)
            lista_datos.append(fila.stock_disponible)
            lista_datos.append(fila.stock_sin_qa)
            lista_datos.append(fila.stock_por_qa)
            lista_datos.append(fila.stock_otros)
            lista_datos.append(fila.total_stock)
            info.append(lista_datos)

        objeto_sociedad = Sociedad.objects.get(id=global_sociedad)

        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        fecha_texto = formatoFechaTexto(StrToDate(fecha_hoy))

        color = DICT_SOCIEDAD[global_sociedad].color
        #####
        query_sociedad = Sociedad.objects.filter(id = int(global_sociedad))[0]
        abreviatura = query_sociedad.abreviatura
        #####
        titulo = "Reporte Resumen de Stock Productos - " + abreviatura + " - " + FECHA_HOY
        vertical = False
        alinear = 'right'
        logo = [[objeto_sociedad.logo.url, alinear]]
        pie_pagina = objeto_sociedad.pie_pagina
        list_texto = []
        # texto = '''Lima, %s''' % str(fecha_texto) + '\n''\n' + '''Facturas por cobrar en la semana: '''
        # list_texto.append(texto)
        TablaEncabezado = [
            'COD. MAT',
            'DESCRIPCIÓN DEL MATERIAL',
            'DISPONIBLE',
            'BLOQ. SIN QA',
            'BLOQ. POR QA',
            'BLOQ. DESG.',
            'SUMATORIA',
            ]

        TablaDatos = []
        for lista in info:
            fila = []
            fila.append(lista[0])
            fila.append(lista[1])
            fila.append(lista[2])
            fila.append(lista[3])
            fila.append(lista[4])
            fila.append(lista[5])
            fila.append(lista[6])
            TablaDatos.append(fila)


        buf = generarReporteResumenStockProductos(titulo, vertical, logo, pie_pagina, list_texto, TablaEncabezado, TablaDatos, color)

        respuesta = HttpResponse(buf.getvalue(), content_type='application/pdf')
        respuesta.headers['content-disposition']='inline; filename=%s.pdf' % titulo

        return respuesta


class ReporteStockSociedadPdf(TemplateView):
    template_name = 'reportes/reporte stock sociedad.html'

    def post(self,request, *args,**kwargs):
        formulario = self.request.POST.get('formulario')
        if formulario == 'formulario1':
            sociedad = Sociedad.objects.get(id=self.request.POST.get('sociedad'))
            tipo = self.request.POST.get('tipo')

            vertical = True
            logo = [sociedad.logo.url]
            pie_pagina = sociedad.pie_pagina
            color = sociedad.color

            if tipo == '1':
                titulo = f'Reporte Stock Disponible por Sociedad - {sociedad.abreviatura}'
                buf = generarReporteStockSociedad(titulo, vertical, logo, pie_pagina, sociedad, color)
            elif tipo == '2':
                titulo = f'Reporte Stock Malogrado por Sociedad - {sociedad.abreviatura}'
                buf = generarReporteStockMalogradoSociedad(titulo, vertical, logo, pie_pagina, sociedad, color)
        elif formulario == 'formulario2':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()
            departamento = self.request.POST.get('departamento')
            tipo = self.request.POST.get('tipo')

            vertical = True
            vertical = False
            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]
            pie_pagina = PIE_DE_PAGINA_DEFAULT
            color = COLOR_DEFAULT

            if tipo == '1':
                if departamento:
                    titulo = f'Reporte Ventas por Departamento - {departamento}'
                else:
                    titulo = f'Reporte Ventas por Departamento - TODOS'

                buf = generarReporteVentasDepartamento(titulo, vertical, logo, pie_pagina, fecha_inicio, fecha_fin, departamento, color)
            
        respuesta = HttpResponse(buf.getvalue(), content_type='application/pdf')
        respuesta.headers['content-disposition']='inline; filename=%s.pdf' % titulo

        return respuesta

    
    def get_context_data(self, **kwargs):
        context = super(ReporteStockSociedadPdf, self).get_context_data(**kwargs)
        context['form'] = ReporteStockSociedadPdfForm()
        context['form2'] = ReporteVentasDepartamentoPdfForm()
        return context

####################################################  REPORTES CRM  ####################################################   

class ReportesCRM(TemplateView):
    template_name = 'reportes/reportes_crm.html'

    def post(self,request, *args,**kwargs):
        formulario = self.request.POST.get('formulario')
        if formulario == 'formulario1':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()
            departamento = self.request.POST.get('departamento')

            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]

            if departamento:
                titulo = f'Reporte Ventas por Departamento - {departamento} del {fecha_inicio} al {fecha_fin}'
            else:
                titulo = f'Reporte Ventas por Departamento - TODOS del {fecha_inicio} al {fecha_fin}'

            wb = ReporteVentasDepartamento(titulo, fecha_inicio, fecha_fin, departamento)
            
        elif formulario == 'formulario2':
            
            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]
            titulo = 'Reporte Cliente CRM' + '_' + str(fecha_doc)

            wb = ReporteClienteCRMExcel()

        elif formulario == 'formulario3':
            cliente = self.request.POST.get('cliente')

            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]
            titulo = 'Reporte Comportamiento Cliente' + '_' + str(fecha_doc)

            wb = ReporteComportamientoCliente(cliente)
        
        elif formulario == 'formulario4':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()

            titulo = f'Reporte Tasa Conversión a Cliente Final del {fecha_inicio} al {fecha_fin}'

            wb = ReporteTasaConversionCliente(fecha_inicio, fecha_fin)
            
        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(titulo + '.xlsx')
        respuesta['content-disposition']= content

        wb.save(respuesta)
        return respuesta

    def get_context_data(self, **kwargs):
        context = super(ReportesCRM, self).get_context_data(**kwargs)
        context['form1'] = ReporteVentasDepartamentoExcelForm()
        context['form3'] = ReporteComportamientoClienteExcelForm()
        context['form4'] = ReporteTasaConversionClienteForm()

        return context
    
####################################################  REPORTES GERENCIA  ####################################################   

class ReportesGerencia(TemplateView):
    template_name = 'reportes/reportes_gerencia.html'

    def post(self,request, *args,**kwargs):
        formulario = self.request.POST.get('formulario')
        if formulario == 'formulario1':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()
            asesor_comercial = self.request.POST.get('asesor_comercial')

            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]

            if asesor_comercial:
                titulo = f'Reporte Facturación por Asesor Comercial - {asesor_comercial} del {fecha_inicio} al {fecha_fin}'
            else:
                titulo = f'Reporte Facturación por Asesor Comercial - TODOS del {fecha_inicio} al {fecha_fin}'

            wb = ReporteFacturacionAsesorComercial(fecha_inicio, fecha_fin, asesor_comercial)
        
        elif formulario == 'formulario2':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d")

            sociedad_MPL = Sociedad.objects.get(abreviatura='MPL')
            # sociedad_MCA = Sociedad.objects.get(abreviatura='MCA')
            # logo = [sociedad_MPL.logo.url, sociedad_MCA.logo.url]
            logo = [sociedad_MPL.logo.url]
            titulo = 'Reporte Facturación General' + '_' + str(fecha_fin.date())

            wb = ReporteFacturacionGeneral(fecha_inicio, fecha_fin)

        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(titulo + '.xlsx')
        respuesta['content-disposition']= content

        wb.save(respuesta)
        return respuesta

    def get_context_data(self, **kwargs):
        context = super(ReportesGerencia, self).get_context_data(**kwargs)
        context['form1'] = ReporteFacturacionAsesorComercialExcelForm()
        context['form2'] = ReporteFacturacionGeneralExcelForm()

        return context
    
####################################################  REPORTES CORREGIDOS  ####################################################   

class ReportesCorregidos(TemplateView):
    template_name = 'reportes/reportes_corregidos.html'

    def post(self,request, *args,**kwargs):
        formulario = self.request.POST.get('formulario')
        if formulario == 'formulario1':
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()
            sociedad = Sociedad.objects.get(id=self.request.POST.get('sociedad'))
            titulo = f'Reporte Contador - {sociedad.abreviatura} del {fecha_inicio} al {fecha_fin}'
            wb = ReporteContadorCorregido(sociedad, fecha_inicio, fecha_fin)

        elif formulario == 'formulario2':
            sociedad = None
            if self.request.POST.get('sociedad'):
                sociedad = Sociedad.objects.get(id=self.request.POST.get('sociedad'))
                titulo = f'Reporte de Rotación - {sociedad.abreviatura} al {date.today()}'
            else:
                titulo = f'Reporte de Rotación - TOTAL al {date.today()}'
            wb = ReporteRotacionCorregido(sociedad)

        elif formulario == 'formulario3':
            sociedad = None
            fecha_inicio = datetime.strptime(self.request.POST.get('fecha_inicio'),"%Y-%m-%d").date()
            fecha_fin = datetime.strptime(self.request.POST.get('fecha_fin'),"%Y-%m-%d").date()

            if self.request.POST.get('sociedad'):
                sociedad = Sociedad.objects.get(id=self.request.POST.get('sociedad'))
                titulo = f'Reporte Depositos Cuentas Bancarias ~ {sociedad.abreviatura} ~ {fecha_inicio} al {fecha_fin}'
            else:
                titulo = f'Reporte Depositos Cuentas Bancarias ~ General ~ {fecha_inicio} al {fecha_fin}'

            wb = ReporteDepositoCuentasBancariasCorregido(sociedad, fecha_inicio, fecha_fin)

        elif formulario == 'formulario4':
            sede = None

            if self.request.POST.get('sede'):
                sede = Sede.objects.get(id=self.request.POST.get('sede'))
                titulo = f'Reporte Resumen Stock Productos {sede.nombre} ~ ' + str(fecha_doc)
            else:
                titulo = f'Reporte Resumen Stock Productos ~ General ~ ' + str(fecha_doc)

            wb = ReporteResumenStockProductosCorregido(sede)

        respuesta = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename ={0}".format(titulo + '.xlsx')
        respuesta['content-disposition']= content
        
        wb.save(respuesta)
        return respuesta

    def get_context_data(self, **kwargs):
        context = super(ReportesCorregidos, self).get_context_data(**kwargs)
        context['form1'] = ReportesContadorForm()
        context['form2'] = ReportesRotacionForm()
        context['form3'] = ReporteDepositoCuentasBancariasForm()
        context['form4'] = ReporteResumenStockProductosForm()

        return context